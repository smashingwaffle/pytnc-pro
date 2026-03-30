#!/usr/bin/env python3
r"""
PyTNC Pro - APRS Transceiver with Map Display

Features:
- APRS RX/TX with real-time map display
- APRS-IS gateway integration
- Beacon transmission with AFSK modulator
- GPS support (NMEA via COM port)
- PTT control (Serial RTS/DTR)
- VARA FM support
- EmComm layers (Weather, Earthquakes, Fires, AQI, Hospitals)
"""

__version__ = "0.1.4-beta"
VERSION = __version__

import sys
import time
import re
import math
import json
import queue
import threading
import http.server
import socket
import socketserver
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple
from functools import partial
import urllib.parse

import numpy as np

try:
    import sounddevice as sd
    HAS_SOUNDDEVICE = True
except ImportError:
    HAS_SOUNDDEVICE = False
    sd = None
    print("Warning: sounddevice not installed - RF AFSK transmit disabled")

from PIL import Image, ImageDraw, ImageFont

try:
    import serial
    import serial.tools.list_ports
    HAS_SERIAL = True
except ImportError:
    HAS_SERIAL = False
    print("Warning: pyserial not installed - PTT/GPS control disabled")

from PyQt6.QtGui import QFont, QPixmap, QColor, QIcon
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QUrl, pyqtSlot, QRunnable, QObject, QSize
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QComboBox, QTextEdit, QTextBrowser, QLabel, QGroupBox, QSplitter,
    QProgressBar, QFrame, QGridLayout, QSlider, QMessageBox, QTabWidget,
    QLineEdit, QSpinBox, QDoubleSpinBox, QCheckBox, QScrollArea, QFileDialog,
    QListWidget, QListWidgetItem, QSizePolicy
)
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWebEngineCore import QWebEngineSettings, QWebEnginePage

from ax25_demod import AX25Demodulator
from hdlc_bitbybit import BitByBitHDLC
from ax25_parser import AX25Parser

# Import from config module
from pytnc_config import (
    BASE_DIR, CACHE_DIR, ICON_CACHE_DIR, HESSU_SYMBOLS_DIR, LEAFLET_JS_PATH, LEAFLET_CSS_PATH,
    SETTINGS_FILE, LUT_FILENAME,
    SAMPLE_RATE, TX_SAMPLE_RATE, HTTP_PORT,
    TOCALL_DEVICES, get_device_from_tocall,
    TILE_CACHE_DIR, USER_DATA_DIR, BUNDLE_DIR
)


# =============================================================================
# TNC Module - AFSK Modulator, AX.25 Protocol, VARA FM, Map
# =============================================================================

from tnc import AFSKModulator, APRSPacketBuilder, apply_cosine_ramp
from tnc.vara import VARAFMInterface
from tnc.map import write_map_html
from tnc.ptt import PTTMixin
from tnc.igate import IGateMixin
from tnc.aprs_is import APRSISMixin
from tnc.monitors import MonitorsMixin

# Alias for compatibility
VARAInterface = VARAFMInterface


# =============================================================================
# Local HTTP Server for serving map files
# =============================================================================

# In-memory tile cache for hot tiles (shared across handler instances)
_tile_memory_cache = {}

# 1x1 transparent PNG for missing tiles (prevents 404 black squares)
_TRANSPARENT_TILE = bytes([
    0x89, 0x50, 0x4E, 0x47, 0x0D, 0x0A, 0x1A, 0x0A, 0x00, 0x00, 0x00, 0x0D,
    0x49, 0x48, 0x44, 0x52, 0x00, 0x00, 0x00, 0x01, 0x00, 0x00, 0x00, 0x01,
    0x08, 0x06, 0x00, 0x00, 0x00, 0x1F, 0x15, 0xC4, 0x89, 0x00, 0x00, 0x00,
    0x0A, 0x49, 0x44, 0x41, 0x54, 0x78, 0x9C, 0x63, 0x00, 0x01, 0x00, 0x00,
    0x05, 0x00, 0x01, 0x0D, 0x0A, 0x2D, 0xB4, 0x00, 0x00, 0x00, 0x00, 0x49,
    0x45, 0x4E, 0x44, 0xAE, 0x42, 0x60, 0x82
])

class QuietHandler(http.server.SimpleHTTPRequestHandler):
    """HTTP handler that serves tiles from cache, or fetches and caches them."""
    
    def __init__(self, *args, directory=None, **kwargs):
        # Use BASE_DIR as primary (where map HTML is written)
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)
    
    def log_message(self, format, *args):
        pass  # Suppress logging
    
    def translate_path(self, path):
        """Translate URL path to filesystem path - check multiple locations"""
        if '?' in path:
            path = path.split('?')[0]
        
        # Tile cache - serve from USER_DATA_DIR
        if '/tile_cache/' in path:
            tile_part = path[path.index('/tile_cache/'):]
            tile_path = USER_DATA_DIR / tile_part[1:]
            return str(tile_path)
        
        # Get the relative path
        rel_path = path.lstrip('/')
        
        # Check BASE_DIR first (where map HTML is written)
        base_file = BASE_DIR / rel_path
        if base_file.exists():
            return str(base_file)
        
        # Check BUNDLE_DIR (PyInstaller _MEIPASS for bundled files like leaflet.js)
        bundle_file = BUNDLE_DIR / rel_path
        if bundle_file.exists():
            return str(bundle_file)
        
        # Check _internal directory (PyInstaller 6.x)
        internal_file = BASE_DIR / "_internal" / rel_path
        if internal_file.exists():
            return str(internal_file)
        
        # Default - return BASE_DIR path (will 404 if not found)
        return str(base_file)
    
    def do_GET(self):
        """Handle GET - serve cached tiles or fetch from OSM"""
        try:
            if '/tile_cache/' in self.path:
                return self._handle_tile_request()
            return super().do_GET()
        except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
            pass  # Client disconnected, silently ignore
    
    def _handle_tile_request(self):
        """Handle tile request - serve from cache, or fetch from OSM.
        
        Key principles:
        1. NEVER return 404 - always return valid image (prevents black squares)
        2. Memory cache for hot tiles
        3. Fetch from OSM if not cached, but return transparent on failure
        """
        import urllib.request
        
        try:
            # Parse tile path: /tile_cache/z/x/y.png
            path = self.path
            if '?' in path:
                path = path.split('?')[0]
            
            # Extract z/x/y from path
            parts = path.replace('/tile_cache/', '').replace('.png', '').split('/')
            if len(parts) != 3:
                self._send_transparent_tile()
                return
            
            try:
                z, x, y = int(parts[0]), int(parts[1]), int(parts[2])
            except ValueError:
                self._send_transparent_tile()
                return
            
            tile_key = f"{z}/{x}/{y}"
            
            # 1. Check memory cache (fastest)
            if tile_key in _tile_memory_cache:
                try:
                    self.send_response(200)
                    self.send_header('Content-Type', 'image/png')
                    self.send_header('X-Tile-Source', 'memory')
                    self.end_headers()
                    self.wfile.write(_tile_memory_cache[tile_key])
                    return
                except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
                    return
            
            # 2. Check disk cache
            tile_path = TILE_CACHE_DIR / str(z) / str(x) / f"{y}.png"
            
            if tile_path.exists():
                try:
                    with open(tile_path, 'rb') as f:
                        tile_data = f.read()
                    
                    # Add to memory cache (500 tiles ≈ 25MB)
                    if len(_tile_memory_cache) < 2000:
                        _tile_memory_cache[tile_key] = tile_data
                    elif len(_tile_memory_cache) >= 2000:
                        # Evict oldest entry (simple FIFO)
                        try:
                            oldest = next(iter(_tile_memory_cache))
                            del _tile_memory_cache[oldest]
                            _tile_memory_cache[tile_key] = tile_data
                        except:
                            pass
                    
                    self.send_response(200)
                    self.send_header('Content-Type', 'image/png')
                    self.send_header('X-Tile-Source', 'disk')
                    self.end_headers()
                    self.wfile.write(tile_data)
                    return
                except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
                    return
                except Exception:
                    pass  # Fall through to OSM fetch
            
            # 3. Not in cache - fetch from OSM (threaded server handles concurrency)
            osm_url = f"https://tile.openstreetmap.org/{z}/{x}/{y}.png"
            try:
                req = urllib.request.Request(osm_url, headers={'User-Agent': 'PyTNC-Pro/1.0'})
                with urllib.request.urlopen(req, timeout=2) as resp:  # Reduced timeout
                    tile_data = resp.read()
                    
                    # Save to disk cache for next time
                    try:
                        tile_path.parent.mkdir(parents=True, exist_ok=True)
                        with open(tile_path, 'wb') as f:
                            f.write(tile_data)
                    except:
                        pass
                    
                    # Add to memory cache
                    if len(_tile_memory_cache) < 2000:
                        _tile_memory_cache[tile_key] = tile_data
                    
                    self.send_response(200)
                    self.send_header('Content-Type', 'image/png')
                    self.send_header('X-Tile-Source', 'osm')
                    self.end_headers()
                    self.wfile.write(tile_data)
                    return
            except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
                return
            except Exception:
                # OSM fetch failed - return transparent (NOT 404!)
                # Tile will load on next pan/zoom when cached
                self._send_transparent_tile()
                return
            
        except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
            return
        except Exception:
            self._send_transparent_tile()
    
    def _send_transparent_tile(self):
        """Send a transparent 1x1 PNG - prevents black squares from 404s"""
        try:
            self.send_response(200)
            self.send_header('Content-Type', 'image/png')
            self.send_header('X-Tile-Source', 'transparent-fallback')
            self.end_headers()
            self.wfile.write(_TRANSPARENT_TILE)
        except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
            pass
    
    def end_headers(self):
        try:
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Cache-Control', 'max-age=86400')
            super().end_headers()
        except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
            pass  # Client disconnected, silently ignore


def start_http_server(port: int) -> socketserver.TCPServer:
    """Start a local HTTP server in a background thread."""
    handler = partial(QuietHandler, directory=str(BASE_DIR))
    
    # Multi-threaded server for concurrent tile requests
    class ThreadedTCPServer(socketserver.ThreadingMixIn, socketserver.TCPServer):
        allow_reuse_address = True
        daemon_threads = True  # Threads die when main thread exits
        
        def handle_error(self, request, client_address):
            """Suppress connection errors from client disconnects"""
            import sys
            exc_type = sys.exc_info()[0]
            # Silently ignore connection errors (client closed connection)
            if exc_type in (ConnectionAbortedError, ConnectionResetError, BrokenPipeError, OSError):
                return
            # For other errors, use default handling
            super().handle_error(request, client_address)
    
    server = ThreadedTCPServer(("127.0.0.1", port), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server


from aprs_parser import parse_nmea, aprs_classify


# =============================================================================
# APRS icons
# =============================================================================

def _find_grid_lines(gray: np.ndarray, axis: int, threshold: float = 40.0):
    mean = gray.mean(axis=axis)
    idxs = np.where(mean < threshold)[0]
    if len(idxs) < 10:
        raise RuntimeError("Grid detection failed")
    mids, start, prev = [], idxs[0], idxs[0]
    for i in idxs[1:]:
        if i != prev + 1:
            mids.append((start + prev) // 2)
            start = i
        prev = i
    mids.append((start + prev) // 2)
    return mids


def _symbol_block(a: int) -> Tuple[int, int]:
    ranges = [(0x21, 0x30), (0x31, 0x40), (0x41, 0x50), (0x51, 0x60), (0x61, 0x70), (0x71, 0x7E)]
    for i, (lo, hi) in enumerate(ranges):
        if lo <= a <= hi:
            return (i, a - lo)
    raise ValueError("Out of range")


def build_icon_cache():
    ICON_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    lut_path = BASE_DIR / LUT_FILENAME
    if not lut_path.exists():
        raise FileNotFoundError(f"Missing {LUT_FILENAME}")

    img = Image.open(lut_path).convert("RGBA")
    gray = np.array(img.convert("L"))
    v, h = _find_grid_lines(gray, 0), _find_grid_lines(gray, 1)
    
    col_b = [(v[i], v[i+1]) for i in range(len(v)-1)]
    row_b = [(h[i], h[i+1]) for i in range(len(h)-1)]

    for a in range(0x21, 0x7F):
        blk, col = _symbol_block(a)
        for row_off, prefix in [(1, "primary"), (2, "secondary")]:
            x0, x1 = col_b[1 + col]
            y0, y1 = row_b[blk * 3 + row_off]
            cell = img.crop((x0+6, y0+6, x1-6, y1-6))
            cell.save(ICON_CACHE_DIR / f"{prefix}_{a:03d}.png")


def haversine_meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Return distance in meters between two lat/lon points."""
    import math
    R = 6_371_000  # Earth radius in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def icon_path(table: str, sym: str) -> Tuple[Path, Optional[str]]:
    """
    Get the icon path for an APRS symbol.
    
    Uses Hessu's official symbols if available, otherwise falls back to extracted cache.
    Hessu's naming: symbols/primary/XX.png where XX = ord(sym) - 33 (00-93)
    """
    # Determine table type
    if table == "/":
        folder = "primary"
        overlay = None
    elif table == "\\":
        folder = "secondary"
        overlay = None
    else:
        # Overlay symbol: table char is the overlay, use secondary icons
        folder = "secondary"
        overlay = table
    
    # Get symbol code (ASCII 33-126 maps to 00-93)
    a = ord(sym) if 0x21 <= ord(sym) <= 0x7E else ord("?")
    hessu_num = a - 33  # Hessu uses 0-based numbering from '!'
    
    # Try Hessu's symbols first (cleaner, official)
    hessu_path = HESSU_SYMBOLS_DIR / folder / f"{hessu_num:02d}.png"
    if hessu_path.exists():
        return hessu_path, overlay
    
    # Fall back to our extracted cache
    cache_path = ICON_CACHE_DIR / f"{folder}_{a:03d}.png"
    if cache_path.exists():
        return cache_path, overlay
    
    # Ultimate fallback - return a default icon path
    default_hessu = HESSU_SYMBOLS_DIR / "primary" / "29.png"  # '>' car symbol
    if default_hessu.exists():
        return default_hessu, None
    
    return ICON_CACHE_DIR / "primary_062.png", None  # '>' in old format


def make_overlay(base_path: Path, char: str) -> Path:
    out = ICON_CACHE_DIR / f"overlay_{base_path.stem}_{ord(char):03d}.png"
    if out.exists():
        return out
    img = Image.open(base_path).convert("RGBA")
    draw = ImageDraw.Draw(img)
    
    # Use a font sized to fit the icon
    font_size = max(12, img.size[1] // 2)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except (OSError, IOError):
        font = ImageFont.load_default()  # Font file not found - use default
    
    # Get text bounding box to center it
    bbox = draw.textbbox((0, 0), char, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    
    # Center the text on the icon
    x = (img.size[0] - text_width) // 2 - bbox[0]
    y = (img.size[1] - text_height) // 2 - bbox[1]
    
    # Draw with outline for visibility
    draw.text((x, y), char, font=font, fill=(255,255,255,255), stroke_width=2, stroke_fill=(0,0,0,220))
    img.save(out)
    return out


# =============================================================================
# Callsign Matching Helper
# =============================================================================

def callsigns_match(call1: str, call2: str) -> bool:
    """
    Flexible APRS callsign comparison that handles:
    - Padding (spaces)
    - Case differences
    - SSID variations (-0 same as no SSID)
    - Base callsign matching (KO6IKR matches KO6IKR-9)
    
    Returns True if callsigns refer to same station.
    """
    if not call1 or not call2:
        return False
    
    # Normalize: strip whitespace and uppercase
    c1 = call1.strip().upper()
    c2 = call2.strip().upper()
    
    # Exact match
    if c1 == c2:
        return True
    
    # Extract base callsign and SSID
    def parse_call(c):
        if '-' in c:
            base, ssid_str = c.rsplit('-', 1)
            try:
                ssid = int(ssid_str)
            except ValueError:
                ssid = 0
        else:
            base = c
            ssid = 0
        return base, ssid
    
    base1, ssid1 = parse_call(c1)
    base2, ssid2 = parse_call(c2)
    
    # Base callsigns must match
    if base1 != base2:
        return False
    
    # If either has no SSID (or SSID 0), match any SSID from same base
    # This handles "KO6IKR" matching "KO6IKR-9"
    if ssid1 == 0 or ssid2 == 0:
        return True
    
    # SSIDs must match
    return ssid1 == ssid2


# =============================================================================
# APRS Comment Cleanup Helper
# =============================================================================

def clean_aprs_comment(text: str, max_len: int = 120) -> str:
    """
    Clean APRS comment/status text for display.
    Removes weather tokens, altitude data, control chars, and truncates.
    """
    if not text:
        return ""
    
    # First, remove non-printable and non-ASCII characters (telemetry garbage)
    text = ''.join(c for c in text if c.isprintable() and ord(c) < 128)

    # Truncate at first run of garbage — 3+ consecutive non-printable-friendly chars
    # catches mixed comments like "13.1V 75F Simi Club 2024???????`??l???"
    import re as _re
    garbage_match = _re.search(r'[^\w\s.,!\-_/:()@#\[\]+=\'"]{3,}', text)
    if garbage_match:
        text = text[:garbage_match.start()].strip()

    # Whole-string garbage check — if >60% non-alphanumeric it's pure binary
    if len(text) > 8:
        alpha_count = sum(1 for c in text if c.isalnum() or c in ' .,!?-_/:()')
        if alpha_count / len(text) < 0.4:
            return ""
    
    # Remove altitude data: /A=xxxxxx (exactly 6 digits per APRS spec, optional leading slash)
    text = re.sub(r'/?A=-?\d{6}', '', text)
    
    # Weather tokens pattern - matches individual weather data fields
    # Allow 1-6 digits to handle variations (g0, g005, b10156, etc.)
    weather_tokens = r'(?:[cgstprPLl][\d.]{1,6}|h[\d.]{1,3}|b[\d.]{4,6}|#[\d.]{1,5})'
    
    # Remove OpenTracker version strings: V###OTW# (e.g., V118OTW1)
    text = re.sub(r'V\d+OTW\d*', '', text)
    
    # Remove positionless weather format: _MMDDHHMM followed by weather data
    text = re.sub(r'^_\d{8}' + weather_tokens + r'+\.?', '', text)
    
    # Remove .../SSS or DDD/SSS at start (wind direction/speed)
    text = re.sub(r'^\.{0,3}/[\d.]{3}', '', text)
    text = re.sub(r'^[\d.]{3}/[\d.]{3}', '', text)
    
    # Remove concatenated weather tokens (anywhere in string)
    # This catches: g0t055P000h48b10156
    text = re.sub(weather_tokens + r'{2,}', '', text)
    
    # Remove individual weather tokens at start
    text = re.sub(r'^' + weather_tokens + r'+', '', text)
    
    # Remove standalone weather tokens elsewhere
    text = re.sub(r'(?<!\w)' + weather_tokens + r'(?!\w)', '', text)
    
    # Remove Davis weather station suffix (.DsVP, .DsIP, etc.)
    text = re.sub(r'\.Ds[A-Z]{2,3}', '', text)
    
    # Clean up leading dots/slashes/underscores
    text = re.sub(r'^[./_]+', '', text)
    
    # Collapse whitespace and strip
    text = ' '.join(text.split())
    
    # Remove potentially dangerous chars for HTML (just strip them)
    text = text.replace("'", "").replace('"', '').replace('<', '').replace('>', '')
    
    # If what remains is very short or just numbers/punctuation, discard it
    if len(text) < 3 or not re.search(r'[a-zA-Z]{2}', text):
        return ""
    
    # Truncate
    if len(text) > max_len:
        text = text[:max_len]
    
    return text


# =============================================================================
# Precompiled APRS Message Regex (spec-compliant fixed-width)
# =============================================================================

# APRS message format: :ADDRESSEE:message (addressee is exactly 9 chars, space-padded)
_APRS_MSG_RE = re.compile(r'^:(?P<addressee>.{9}):(?P<text>.*)$')

# =============================================================================
# Audio receiver
# =============================================================================

class AudioReceiver(QThread):
    packet_received = pyqtSignal(object, int)
    audio_level = pyqtSignal(float)
    status_update = pyqtSignal(str)
    error_occurred = pyqtSignal(str)

    def __init__(self, device_id: int, gain: float = 1.0):
        super().__init__()
        self.device_id = device_id
        self.gain = gain
        self.running = False
        self.demod = AX25Demodulator(SAMPLE_RATE, 1200, 1200, 2200, "A")
        self.hdlc = [BitByBitHDLC(0, 0, i, self.demod) for i in range(self.demod.num_slicers)]
        for h in self.hdlc:
            h.set_frame_callback(self._on_frame)
        self.parser = AX25Parser()
        self.queue = queue.Queue(maxsize=200)

    def set_gain(self, g): self.gain = g

    def audio_cb(self, data, frames, time_info, status):
        if status: self.status_update.emit(f"Audio: {status}")
        try: self.queue.put(data.copy(), block=False)
        except queue.Full: pass

    def _on_frame(self, data, ch, sub, sl):
        try:
            pkt = self.parser.parse(data)
            if pkt: self.packet_received.emit(pkt, sl)
        except Exception as e:
            # Log parse errors but don't crash the audio thread
            pass  # AX.25 parse errors are common with noise - silent is OK

    def run(self):
        self.running = True
        
        # Check if sounddevice is available
        if not HAS_SOUNDDEVICE:
            self.error_occurred.emit("sounddevice not installed - RX disabled")
            return
        
        try:
            stream = sd.InputStream(device=self.device_id, channels=1, samplerate=SAMPLE_RATE,
                                    callback=self.audio_cb, blocksize=2400)
            stream.start()
            self.status_update.emit("Audio started")
            while self.running:
                try:
                    chunk = self.queue.get(timeout=1.0)
                    audio = np.clip(chunk.flatten().astype(np.float32) * self.gain, -1, 1)
                    self.audio_level.emit(float(np.abs(audio).mean()))
                    for s in audio:
                        for sl, bit, q in self.demod.process_sample(int(s * 32767)):
                            self.hdlc[sl].process_bit(bit, q)
                except queue.Empty: pass
                except Exception as e: self.error_occurred.emit(str(e))
            stream.stop()
            stream.close()
        except Exception as e:
            self.error_occurred.emit(f"Audio error: {e}")

    def stop(self):
        self.running = False
        self.wait()


# =============================================================================
# GUI
# =============================================================================

class FetchWorkerSignals(QObject):
    """Signals for NetworkFetchWorker"""
    finished = pyqtSignal(object)  # Emits the result data
    error = pyqtSignal(str)        # Emits error message


class NetworkFetchWorker(QRunnable):
    """
    Worker for non-blocking network fetches.
    Uses QThreadPool instead of blocking UI thread.
    Handles both JSON and CSV/text responses.
    """
    def __init__(self, url, headers=None, timeout=30, expect_json=True):
        super().__init__()
        self.url = url
        self.headers = headers or {'User-Agent': 'PyTNC-Pro/1.0'}
        self.timeout = timeout
        self.expect_json = expect_json
        self.signals = FetchWorkerSignals()
    
    def run(self):
        """Execute the fetch in background thread"""
        try:
            import requests
            response = requests.get(
                self.url, 
                timeout=self.timeout, 
                headers=self.headers
            )
            response.raise_for_status()
            
            # Try JSON first, fall back to text
            try:
                data = response.json()
                self.signals.finished.emit(data)
            except (ValueError, json.JSONDecodeError):
                # Not JSON - return raw text (for CSV, etc.)
                self.signals.finished.emit(response.text)
                
        except ImportError:
            # Fallback to urllib
            try:
                import urllib.request
                import ssl
                import json as json_module
                context = ssl._create_unverified_context()
                req = urllib.request.Request(self.url, headers=self.headers)
                with urllib.request.urlopen(req, timeout=self.timeout, context=context) as response:
                    raw = response.read().decode()
                    # Try JSON first
                    try:
                        data = json_module.loads(raw)
                        self.signals.finished.emit(data)
                    except (ValueError, json.JSONDecodeError):
                        # Not JSON - return raw text
                        self.signals.finished.emit(raw)
            except Exception as e:
                self.signals.error.emit(f"urllib fallback: {e}")
        except Exception as e:
            self.signals.error.emit(str(e))


class AudioMeter(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        lbl = QLabel("Audio Level")
        lbl.setStyleSheet("color:#a0c4ff;font-weight:bold;")
        layout.addWidget(lbl)
        self.bar = QProgressBar()
        self.bar.setMaximum(100)
        self.bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #1e3a5f;
                border-radius: 6px;
                background: #0a1929;
                color: #a0c4ff;
                text-align: center;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1976d2, stop:0.5 #42a5f5, stop:1 #1976d2);
                border-radius: 4px;
            }
        """)
        layout.addWidget(self.bar)
    def set_level(self, v): self.bar.setValue(min(max(int(v*100),0),100))


class LogPage(QWebEnginePage):
    def __init__(self, log_fn, parent=None):
        super().__init__(parent)
        self.log = log_fn
    
    def javaScriptConsoleMessage(self, lv, msg, line, src):
        # Check for external URL open request
        if msg.startswith('OPEN_EXTERNAL:'):
            url = msg.replace('OPEN_EXTERNAL:', '')
            import webbrowser
            webbrowser.open(url)
            self.log(f"🌐 Opening: {url}")
        else:
            self.log(f"[JS] {msg}")
    
    def createWindow(self, window_type):
        """Handle target=_blank links by opening in system browser"""
        # Create a temporary page to capture the URL
        temp_page = QWebEnginePage(self)
        temp_page.urlChanged.connect(self._handle_new_window_url)
        return temp_page
    
    def _handle_new_window_url(self, url):
        """Open the URL from target=_blank in system browser"""
        url_str = url.toString()
        if url_str and url_str != 'about:blank':
            import webbrowser
            webbrowser.open(url_str)
            self.log(f"🌐 Opening: {url_str}")
    
    def acceptNavigationRequest(self, url, nav_type, is_main_frame):
        """Handle link clicks - open external URLs in system browser"""
        url_str = url.toString()
        
        # If it's an external URL (not our local server), open in browser
        if url_str.startswith('https://') or url_str.startswith('http://'):
            if '127.0.0.1' not in url_str and 'localhost' not in url_str:
                # External link - open in system browser
                import webbrowser
                webbrowser.open(url_str)
                self.log(f"🌐 Opening: {url_str}")
                return False  # Don't navigate in the WebView
        
        # Allow local navigation
        return True


class MainWindow(PTTMixin, IGateMixin, APRSISMixin, MonitorsMixin, QMainWindow):
    # Signals for thread-safe UI updates
    aprs_is_connected_signal = pyqtSignal()
    aprs_is_disconnected_signal = pyqtSignal()
    gps_position_signal = pyqtSignal(float, float)  # lat, lon
    gps_status_signal = pyqtSignal(bool, float)     # has_fix, speed_mph
    aprs_is_packet_signal = pyqtSignal(str)
    aprs_is_error_signal = pyqtSignal(str)
    
    def __init__(self, http_port: int):
        super().__init__()
        
        # Connect APRS-IS signals
        self.aprs_is_connected_signal.connect(self._aprs_is_connected)
        self.aprs_is_disconnected_signal.connect(self._aprs_is_disconnected)
        self.aprs_is_packet_signal.connect(self._handle_aprs_is_packet)
        self.aprs_is_error_signal.connect(lambda msg: self._log(msg))
        
        # Connect GPS signals
        self.gps_position_signal.connect(self._update_gps_position)
        self.gps_status_signal.connect(self._update_gps_status)
        
        self.http_port = http_port
        self.receiver = None
        self.packets = 0
        self.dedup = {}
        self.map_ready = False
        self.pending_js = []
        self.map_checks = 0
        self.log_buf = []
        self.log_history = []  # Full history for filtering
        
        # Station status/info cache - stores last status message per callsign
        self.station_status = {}  # {callsign: "status text"}
        
        # Digipeater traffic tracking - which stations used which digis
        # Format: {digi_callsign: [(station, timestamp), ...]}
        self.digi_traffic = {}
        
        # TX in progress flag - prevents self-decode during transmit
        self.tx_in_progress = False
        self.tx_end_time = 0  # Timestamp when TX ended (for holdoff)
        
        # Telemetry coefficient cache per station
        # Format: {callsign: {"parm": [...], "unit": [...], "eqns": [...]}}
        self.telem_defs = {}
        
        # PTT serial connection (for RTS/DTR PTT control)
        self.ptt_serial = None
        # CI-V CAT serial connection (for Icom CI-V PTT)
        self.civ_serial = None
        self.civ_ptt_method = "RTS/DTR"  # "RTS/DTR", "CI-V CAT", or "CM108 GPIO"
        # CM108 GPIO PTT (DigiRig Lite / USB audio dongles)
        self.cm108_device = None   # open hid device handle
        
        # GPS serial connection
        self.gps_serial = None
        self.gps_timer = None
        self.gps_running = False
        self.gps_lat = None
        self.gps_lon = None
        self.gps_has_fix = False
        self.gps_buffer = ""  # Buffer for NMEA sentence assembly
        self.last_beacon_lat = None  # last position we beaconed from
        self.last_beacon_lon = None
        self.MIN_BEACON_DISTANCE_M = 25  # meters — don't beacon if moved less than this
        
        # APRS-IS connection
        self.aprs_is_socket = None
        self.aprs_is_thread = None
        self.aprs_is_running = False
        self.aprs_is_connected = False  # Connection status flag for UI

        # IGate state
        self.igate_rx_enabled = False   # RF → Internet
        self.igate_tx_enabled = False   # Internet → RF (off by default)
        self.igate_rx_count = 0
        self.igate_tx_count = 0
        self.igate_start_time = None
        self.igate_rf_heard = {}        # {callsign: timestamp} for TX gate eligibility
        self.igate_recent = []          # recent gated packets for display (max 200)
        self.igate_dedup = {}           # {(src, info): timestamp} gate-side dedup
        
        # VARA FM connection
        self.vara_connected = False
        self.vara_rx_thread = None
        self.vara_rx_running = False
        self.vara_kiss_socket = None
        self.vara_kiss_connected = False
        self.vara_kiss_rx_thread = None
        
        # VARA FM state flags
        self._vara_symbol_table = "/"
        self._vara_symbol_code = ">"
        self._vara_beacon_in_progress = False
        self._vara_is_connected_to_remote = False
        self._vara_ptt_active = False
        
        self._build_ui()
        self._init()
        self.load_devices()
        self.load_com_ports()
        self.load_settings()  # Load saved settings
        
        # Update cache status indicators
        QTimer.singleShot(500, self._update_cache_status)
        
        self._log("=" * 50)
        self._log(f"PyTNC Pro v{VERSION} - APRS Transceiver")
        self._log(f"HTTP server: http://127.0.0.1:{http_port}/")
        self._log(f"User data: {USER_DATA_DIR}")
        self._log("=" * 50)
        
        # Initialize connection status display
        QTimer.singleShot(100, self._sync_beacon_connection_status)

    def _init(self):
        global HESSU_SYMBOLS_DIR
        ICON_CACHE_DIR.mkdir(exist_ok=True)
        
        # Find icons - search common locations (PyInstaller puts them in different places)
        from pytnc_config import BASE_DIR, BUNDLE_DIR
        import pytnc_config
        
        search_paths = [
            BUNDLE_DIR / "aprs_symbols_48",  # PyInstaller _MEIPASS
            BASE_DIR / "aprs_symbols_48",     # Dev or portable
            BASE_DIR / "_internal" / "aprs_symbols_48",  # PyInstaller 6.x
        ]
        
        found_icons = None
        for p in search_paths:
            test = p / "primary" / "29.png"
            if test.exists():
                found_icons = p
                break
        
        if found_icons:
            if found_icons != HESSU_SYMBOLS_DIR:
                self._log(f"🎨 Found icons at: {found_icons}")
            # Update both local and module-level
            HESSU_SYMBOLS_DIR = found_icons
            pytnc_config.HESSU_SYMBOLS_DIR = found_icons
            
            primary_count = len(list((found_icons / "primary").glob("*.png")))
            secondary_count = len(list((found_icons / "secondary").glob("*.png")))
            self._log(f"🎨 Icons: {primary_count} primary + {secondary_count} secondary")
            
            # Rebuild symbol grids now that we found icons
            QTimer.singleShot(100, self._rebuild_symbol_grids)
        else:
            self._log(f"⚠️ Icons not found!")
            self._log(f"   Searched: {[str(p) for p in search_paths]}")
            if not (ICON_CACHE_DIR / "primary_033.png").exists():
                self._log("   Building fallback icons...")
                try:
                    build_icon_cache()
                except Exception as e:
                    self._log(f"   Icon error: {e}")
    
    def _rebuild_symbol_grids(self):
        """Rebuild symbol picker grids after finding icons"""
        if hasattr(self, 'symbol_grid_layout'):
            self._update_symbol_grid()
        if hasattr(self, 'vara_symbol_grid_layout'):
            self._vara_build_symbol_grid()
        
        for f, name in [(LEAFLET_JS_PATH, "leaflet.js"), (LEAFLET_CSS_PATH, "leaflet.css")]:
            if f.exists():
                self._log(f"{name}: {f.stat().st_size:,} bytes")
            else:
                self._log(f"MISSING: {name}")

    def _build_ui(self):
        self.setWindowTitle(f"PyTNC Pro v{VERSION} - APRS Transceiver")
        self.setGeometry(50, 50, 1400, 850)
        
        # Set application icon — check all possible locations including PyInstaller bundle
        icon_paths = [
            BUNDLE_DIR / "pytnc_pro.ico",
            BUNDLE_DIR / "tnc" / "icon.png",
            Path(__file__).parent / "pytnc_pro.ico",
            Path(__file__).parent / "tnc" / "icon.png",
            Path(sys.executable).parent / "pytnc_pro.ico",
        ]
        for _ip in icon_paths:
            if _ip.exists():
                _icon = QIcon(str(_ip))
                self.setWindowIcon(_icon)
                QApplication.instance().setWindowIcon(_icon)
                break
        
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(3)
        
        # Header - ultra compact
        hdr = QFrame()
        hdr.setFixedHeight(24)
        hdr.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #1a3a5c, stop:1 #0d2137);
            border-radius: 4px;
            border: 1px solid #2a5a8a;
        """)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(8, 2, 8, 2)
        hl.setSpacing(8)
        
        self.dot = QLabel("●")
        self.dot.setStyleSheet("font-size:14px;color:#ff6b6b")
        hl.addWidget(self.dot)
        
        self.status_lbl = QLabel("STOPPED")
        self.status_lbl.setStyleSheet("color:#e0e0e0;font-weight:bold;font-size:11px")
        hl.addWidget(self.status_lbl)
        
        hl.addSpacing(15)
        self.pkt_lbl = QLabel("Packets: 0")
        self.pkt_lbl.setStyleSheet("color:#64b5f6;font-weight:bold;font-size:11px")
        hl.addWidget(self.pkt_lbl)
        
        hl.addStretch()
        self.map_lbl = QLabel("Map: loading...")
        self.map_lbl.setStyleSheet("color:#ffd54f;font-size:11px")
        hl.addWidget(self.map_lbl)
        
        layout.addWidget(hdr)
        
        # Control bar - ultra compact single line with just buttons
        ctrl = QFrame()
        ctrl.setFixedHeight(32)
        ctrl.setStyleSheet("""
            QFrame {
                background: #0d2137;
                border: 1px solid #1e3a5f;
                border-radius: 4px;
            }
            QLabel { color: #b0bec5; font-size: 10px; }
        """)
        ctrl_layout = QHBoxLayout(ctrl)
        ctrl_layout.setContentsMargins(6, 2, 6, 2)
        ctrl_layout.setSpacing(6)
        
        # RX Audio meter - prominent red bar
        rx_label = QLabel("RX")
        rx_label.setStyleSheet("color: #ff6666; font-weight: bold; font-size: 12px;")
        ctrl_layout.addWidget(rx_label)
        
        self.meter = QProgressBar()
        self.meter.setMaximum(100)
        self.meter.setFixedWidth(200)
        self.meter.setFixedHeight(18)
        self.meter.setTextVisible(False)
        self.meter.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.meter.setStyleSheet("""
            QProgressBar {
                border: 2px solid #aa3333;
                border-radius: 4px;
                background: #1a0a0a;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #cc0000, stop:0.5 #ff3333, stop:0.8 #ff6600, stop:1 #ffcc00);
                border-radius: 2px;
            }
        """)
        ctrl_layout.addWidget(self.meter)
        
        # Callsign labels toggle
        self.rx_callsign_check = QCheckBox("🏷️ Callsigns")
        self.rx_callsign_check.setToolTip("Show callsign labels on map")
        self.rx_callsign_check.setStyleSheet("color: #ffd54f; font-size: 11px;")
        self.rx_callsign_check.setChecked(True)  # Default ON
        self.rx_callsign_check.stateChanged.connect(self._rx_toggle_callsigns)
        ctrl_layout.addWidget(self.rx_callsign_check)
        
        # Station trails toggle
        self.rx_trails_check = QCheckBox("〰️ Trails")
        self.rx_trails_check.setToolTip("Show movement trails for mobile stations")
        self.rx_trails_check.setStyleSheet("color: #ce93d8; font-size: 11px;")
        self.rx_trails_check.setChecked(True)  # Default ON
        self.rx_trails_check.stateChanged.connect(self._rx_toggle_trails)
        ctrl_layout.addWidget(self.rx_trails_check)
        
        # Hospital toggle on RX page
        self.rx_hospital_check = QCheckBox("🏥 Hospitals")
        self.rx_hospital_check.setToolTip("Show hospitals on map")
        self.rx_hospital_check.setStyleSheet("color: #4da6ff; font-size: 11px;")
        self.rx_hospital_check.stateChanged.connect(self._rx_toggle_hospitals)
        ctrl_layout.addWidget(self.rx_hospital_check)
        
        # Weather alerts toggle on RX page
        self.rx_weather_check = QCheckBox("⚠️ NOAA")
        self.rx_weather_check.setToolTip("Show NWS weather alerts")
        self.rx_weather_check.setStyleSheet("color: #ff9800; font-size: 11px;")
        self.rx_weather_check.stateChanged.connect(self._rx_toggle_weather)
        ctrl_layout.addWidget(self.rx_weather_check)
        
        # DARN emergency repeaters toggle
        self.rx_darn_check = QCheckBox("🔴 DARN")
        self.rx_darn_check.setToolTip("Show DARN emergency repeater network")
        self.rx_darn_check.setStyleSheet("color: #ff6b6b; font-size: 11px;")
        self.rx_darn_check.stateChanged.connect(self._rx_toggle_darn)
        ctrl_layout.addWidget(self.rx_darn_check)
        
        # Fire/wildfire toggle (NASA FIRMS)
        self.rx_fire_check = QCheckBox("🔥 Fires")
        self.rx_fire_check.setToolTip("Show NASA wildfire hotspots (requires API key in Settings)")
        self.rx_fire_check.setStyleSheet("color: #ff5722; font-size: 11px;")
        self.rx_fire_check.stateChanged.connect(self._rx_toggle_fires)
        ctrl_layout.addWidget(self.rx_fire_check)
        
        # Earthquake toggle
        self.rx_quake_check = QCheckBox("🌍 Quakes")
        self.rx_quake_check.setToolTip("Show recent earthquakes from USGS")
        self.rx_quake_check.setStyleSheet("color: #ce93d8; font-size: 11px;")
        self.rx_quake_check.stateChanged.connect(self._rx_toggle_quakes)
        ctrl_layout.addWidget(self.rx_quake_check)
        
        # AQI toggle
        self.rx_aqi_check = QCheckBox("💨 AQI")
        self.rx_aqi_check.setToolTip("Show Air Quality Index from AirNow")
        self.rx_aqi_check.setStyleSheet("color: #8bc34a; font-size: 11px;")
        self.rx_aqi_check.stateChanged.connect(self._rx_toggle_aqi)
        ctrl_layout.addWidget(self.rx_aqi_check)
        
        ctrl_layout.addStretch()
        
        # Hidden gain slider (still needed internally)
        self.gain = QSlider(Qt.Orientation.Horizontal)
        self.gain.setRange(1, 100)
        self.gain.setValue(10)
        self.gain.valueChanged.connect(self.on_gain)
        self.gain.hide()  # Hidden - controlled from Settings
        
        # Hidden device combo (still needed internally)
        self.dev_combo = QComboBox()
        self.dev_combo.hide()  # Hidden - controlled from Settings
        
        self.gain_lbl = QLabel("1.0x")
        self.gain_lbl.hide()
        
        # APRS-IS status label (shows server when connected)
        self.aprs_is_info_label = QLabel("")
        self.aprs_is_info_label.setStyleSheet("color: #64b5f6; font-size: 10px;")
        ctrl_layout.addWidget(self.aprs_is_info_label)
        
        # APRS-IS Connect button
        self.aprs_is_connect_btn = QPushButton("🌐 START IS")
        self.aprs_is_connect_btn.setFixedHeight(22)
        self.aprs_is_connect_btn.clicked.connect(self.toggle_aprs_is)
        self.aprs_is_connect_btn.setStyleSheet("""
            QPushButton {
                background: #0277bd;
                color: #fff;
                font-weight: bold;
                padding: 2px 10px;
                border-radius: 3px;
                border: 1px solid #0288d1;
                font-size: 10px;
            }
            QPushButton:hover { background: #0288d1; }
        """)
        ctrl_layout.addWidget(self.aprs_is_connect_btn)
        
        self.aprs_is_status = QLabel("●")
        self.aprs_is_status.setStyleSheet("color: #ff6b6b; font-size: 14px;")
        ctrl_layout.addWidget(self.aprs_is_status)
        
        ctrl_layout.addSpacing(10)
        
        # Start/Stop buttons (compact)
        self.start_btn = QPushButton("▶ START RF")
        self.start_btn.setFixedHeight(22)
        self.start_btn.clicked.connect(self.start)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background: #2e7d32;
                color: #fff;
                font-weight: bold;
                padding: 2px 10px;
                border-radius: 3px;
                border: 1px solid #4caf50;
                font-size: 10px;
            }
            QPushButton:hover { background: #43a047; }
            QPushButton:disabled { background: #1a3a5c; border-color: #2a5a8a; color: #607d8b; }
        """)
        ctrl_layout.addWidget(self.start_btn)
        
        self.stop_btn = QPushButton("■ STOP RF")
        self.stop_btn.setFixedHeight(22)
        self.stop_btn.clicked.connect(self.stop)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("""
            QPushButton {
                background: #c62828;
                color: #fff;
                font-weight: bold;
                padding: 2px 10px;
                border-radius: 3px;
                border: 1px solid #ef5350;
                font-size: 10px;
            }
            QPushButton:hover { background: #e53935; }
            QPushButton:disabled { background: #1a3a5c; border-color: #2a5a8a; color: #607d8b; }
        """)
        ctrl_layout.addWidget(self.stop_btn)
        
        # Main content - build the split view for RX tab
        split = QSplitter(Qt.Orientation.Horizontal)
        
        # Left - Log panel
        left = QWidget()
        lv = QVBoxLayout(left)
        lv.setContentsMargins(0, 0, 0, 0)
        
        log_grp = QGroupBox("📡 Live Feed")
        log_grp.setStyleSheet("""
            QGroupBox {
                color: #a0c4ff;
                font-weight: bold;
                border: 1px solid #1e3a5f;
                border-radius: 10px;
                margin-top: 5px;
                padding-top: 8px;
                background: #0d2137;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                background: #0d2137;
            }
        """)
        log_l = QVBoxLayout(log_grp)
        
        # Filter search box
        self.log_filter = QLineEdit()
        self.log_filter.setPlaceholderText("🔍 Filter by callsign...")
        self.log_filter.setClearButtonEnabled(True)
        self.log_filter.setStyleSheet("""
            QLineEdit {
                background: #0a1929;
                color: #ffd54f;
                border: 1px solid #1e3a5f;
                border-radius: 4px;
                padding: 4px 8px;
                font: 11px Consolas, monospace;
            }
            QLineEdit:focus {
                border-color: #42a5f5;
            }
        """)
        self.log_filter.textChanged.connect(self._filter_log)
        log_l.addWidget(self.log_filter)
        
        self.log_txt = QTextBrowser()
        self.log_txt.setReadOnly(True)
        self.log_txt.setOpenExternalLinks(False)  # Handle clicks ourselves
        self.log_txt.setOpenLinks(False)  # Don't navigate - prevents clearing content
        self.log_txt.anchorClicked.connect(self._log_link_clicked)
        self.log_txt.setFont(QFont("Consolas", 10))
        self.log_txt.setStyleSheet("""
            QTextBrowser {
                background: #0a1628;
                color: #7fff7f;
                border: 1px solid #1e3a5f;
                border-radius: 6px;
                padding: 8px;
                selection-background-color: #1976d2;
            }
            QTextBrowser a {
                color: #ffd54f;
                text-decoration: none;
            }
            QTextBrowser a:hover {
                text-decoration: underline;
            }
        """)
        log_l.addWidget(self.log_txt)
        lv.addWidget(log_grp)
        
        for line in self.log_buf:
            self.log_txt.append(line)
            self.log_history.append(line)
        self.log_buf.clear()
        
        split.addWidget(left)
        
        # Right - map
        right = QWidget()
        rv = QVBoxLayout(right)
        rv.setContentsMargins(0, 0, 0, 0)
        
        # Map controls row
        map_ctrl_row = QHBoxLayout()
        map_ctrl_row.setSpacing(4)
        
        map_label = QLabel("🗺️ APRS Map")
        map_label.setStyleSheet("color: #a0c4ff; font-weight: bold; font-size: 12px;")
        map_ctrl_row.addWidget(map_label)
        map_ctrl_row.addStretch()
        
        # Refresh button
        map_refresh_btn = QPushButton("🔄")
        map_refresh_btn.setFixedSize(28, 24)
        map_refresh_btn.setToolTip("Refresh map tiles")
        map_refresh_btn.setStyleSheet("""
            QPushButton { background: #1565c0; border: none; border-radius: 4px; color: white; }
            QPushButton:hover { background: #1976d2; }
        """)
        map_refresh_btn.clicked.connect(self._refresh_map)
        map_ctrl_row.addWidget(map_refresh_btn)

        # Map layer switcher
        self.map_layer_combo = QComboBox()
        self.map_layer_combo.addItem("🗺️ Street", "osm")
        self.map_layer_combo.addItem("⛰️ Topo", "topo")
        self.map_layer_combo.addItem("🛰️ Satellite", "satellite")
        self.map_layer_combo.addItem("🌙 Dark", "dark")
        self.map_layer_combo.setFixedHeight(24)
        self.map_layer_combo.setToolTip("Map Style")
        self.map_layer_combo.setStyleSheet("""
            QComboBox {
                background: #0d2137; color: #e0e0e0;
                border: 2px solid #42a5f5; border-radius: 4px;
                padding: 0 8px; font-size: 11px;
            }
            QComboBox:hover { border-color: #ffd54f; }
            QComboBox::drop-down { border: none; width: 18px; }
            QComboBox QAbstractItemView {
                background: #1a1a2e; color: #fff;
                selection-background-color: #1565c0;
            }
        """)
        self.map_layer_combo.currentIndexChanged.connect(self._on_map_layer_changed)
        map_ctrl_row.addWidget(self.map_layer_combo)

        # Load Locations button
        self.load_locations_btn = QPushButton("📍 Locations")
        self.load_locations_btn.setFixedHeight(24)
        self.load_locations_btn.setToolTip("Load locations from files or folder")
        self.load_locations_btn.setStyleSheet("""
            QPushButton { background: #6a1b9a; border: none; border-radius: 4px; color: white; padding: 0 8px; }
            QPushButton:hover { background: #7b1fa2; }
        """)
        self.load_locations_btn.clicked.connect(self._load_locations_menu)
        map_ctrl_row.addWidget(self.load_locations_btn)
        
        # Beacon Locations button (hidden until locations loaded)
        self.beacon_locations_btn = QPushButton("📡 Beacon")
        self.beacon_locations_btn.setFixedHeight(24)
        self.beacon_locations_btn.setToolTip("Beacon all locations as APRS objects")
        self.beacon_locations_btn.setStyleSheet("""
            QPushButton { background: #2e7d32; border: none; border-radius: 4px; color: white; padding: 0 8px; }
            QPushButton:hover { background: #388e3c; }
        """)
        self.beacon_locations_btn.clicked.connect(self._beacon_locations_menu)
        self.beacon_locations_btn.hide()
        map_ctrl_row.addWidget(self.beacon_locations_btn)
        
        # Clear Locations button (hidden until locations loaded)
        self.clear_locations_btn = QPushButton("🗑️")
        self.clear_locations_btn.setFixedSize(24, 24)
        self.clear_locations_btn.setToolTip("Clear all loaded locations")
        self.clear_locations_btn.setStyleSheet("""
            QPushButton { background: #c62828; border: none; border-radius: 4px; color: white; }
            QPushButton:hover { background: #e53935; }
        """)
        self.clear_locations_btn.clicked.connect(self._clear_locations)
        self.clear_locations_btn.hide()
        map_ctrl_row.addWidget(self.clear_locations_btn)
        
        # Location count label
        self.location_count_lbl = QLabel("")
        self.location_count_lbl.setStyleSheet("color: #ce93d8; font-size: 10px;")
        map_ctrl_row.addWidget(self.location_count_lbl)
        
        rv.addLayout(map_ctrl_row)
        
        # Map widget
        map_container = QFrame()
        map_container.setStyleSheet("""
            QFrame {
                border: 1px solid #1e3a5f;
                border-radius: 6px;
                background: #0d2137;
            }
        """)
        map_inner = QVBoxLayout(map_container)
        map_inner.setContentsMargins(2, 2, 2, 2)
        
        self.map = QWebEngineView()
        s = self.map.settings()
        s.setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessRemoteUrls, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.LocalContentCanAccessFileUrls, True)
        # GPU acceleration - needed for smooth rendering on high-res displays
        s.setAttribute(QWebEngineSettings.WebAttribute.WebGLEnabled, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.Accelerated2dCanvasEnabled, True)
        s.setAttribute(QWebEngineSettings.WebAttribute.ScrollAnimatorEnabled, False)
        s.setAttribute(QWebEngineSettings.WebAttribute.TouchIconsEnabled, False)
        s.setAttribute(QWebEngineSettings.WebAttribute.FocusOnNavigationEnabled, False)
        
        self.map.setPage(LogPage(self._log, self.map))
        self.map.loadFinished.connect(self._map_loaded)
        
        # Write HTML and load via HTTP
        try:
            write_map_html(BASE_DIR, self.http_port)
            url = f"http://127.0.0.1:{self.http_port}/aprs_map.html"
            self._log(f"Loading map from: {url}")
            self.map.load(QUrl(url))
        except Exception as e:
            self._log(f"Map error: {e}")
        
        map_inner.addWidget(self.map)
        rv.addWidget(map_container, 1)
        split.addWidget(right)
        
        split.setSizes([450, 950])
        
        # Create tab widget - goes right after header!
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #1e3a5f;
                background: #0d2137;
                border-radius: 6px;
            }
            QTabBar::tab {
                background: #0d2137;
                color: #607d8b;
                padding: 8px 20px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background: #1a3a5c;
                color: #64b5f6;
                border-bottom: 2px solid #42a5f5;
            }
            QTabBar::tab:hover:!selected {
                background: #152a40;
                color: #90caf9;
            }
        """)
        
        # Tab 1: Receive - includes RX audio controls AND split view
        rx_tab = QWidget()
        rx_layout = QVBoxLayout(rx_tab)
        rx_layout.setContentsMargins(0, 0, 0, 0)
        rx_layout.setSpacing(2)
        
        # Add RX control bar to Receive tab only (already includes APRS-IS button)
        rx_layout.addWidget(ctrl)
        rx_layout.addWidget(split, 1)  # stretch factor 1 = fill remaining space
        rx_layout.addWidget(self._branding_label())
        
        self.tabs.addTab(rx_tab, "🗺️ MAP")
        
        # Hidden APRS-IS settings (used by connect function) - initialize defaults
        self.aprs_is_server = QLineEdit("noam.aprs2.net")
        self.aprs_is_port = QSpinBox()
        self.aprs_is_port.setRange(1, 65535)  # Must set range before value!
        self.aprs_is_port.setValue(14580)
        self.aprs_is_filter = QLineEdit("r/34.05/-118.24/50")
        # These are hidden - actual settings come from Settings tab
        
        # Tab 2: Settings & Beacon
        self._build_settings_tab()

        layout.addWidget(self.tabs, 1)  # stretch factor 1 = fill remaining space
        
        # Main window dark cobalt blue theme
        self.setStyleSheet("""
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0d2137, stop:1 #091428);
            }
            QSplitter::handle {
                background: #1e3a5f;
                width: 3px;
            }
            QSplitter::handle:hover {
                background: #42a5f5;
            }
            QScrollBar:vertical {
                background: #0d2137;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #1e3a5f;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: #2a5a8a;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0;
            }
        """)

    def _build_settings_tab(self):
        """Build the Transmit & Beacon tab - simplified, connections in Settings"""
        settings_tab = QWidget()
        settings_layout = QHBoxLayout(settings_tab)
        settings_layout.setContentsMargins(10, 10, 10, 10)
        settings_layout.setSpacing(10)
        
        # Left panel - Beacon Settings only
        left_panel = QWidget()
        left_panel.setMaximumWidth(620)  # Constrain left panel
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(10)
        
        # Hidden port combos - needed for compatibility but not shown
        self.ptt_port_combo = QComboBox()
        self.ptt_port_combo.hide()
        self.tx_audio_combo = QComboBox()
        self.tx_audio_combo.hide()
        self.tx_level_slider = QSlider(Qt.Orientation.Horizontal)
        self.tx_level_slider.setRange(1, 100)
        self.tx_level_slider.setValue(10)
        self.tx_level_slider.hide()
        
        # Connection Status display
        status_grp = QGroupBox("🔌 Connection Status")
        status_grp.setStyleSheet(self._group_style())
        status_layout = QGridLayout(status_grp)
        status_layout.setSpacing(4)
        
        self.tx_ptt_status = QLabel("⚫ PTT: Not connected")
        self.tx_ptt_status.setStyleSheet("color: #607d8b;")
        status_layout.addWidget(self.tx_ptt_status, 0, 0)
        
        self.tx_gps_status = QLabel("⚫ GPS: Not connected")
        self.tx_gps_status.setStyleSheet("color: #607d8b;")
        status_layout.addWidget(self.tx_gps_status, 0, 1)
        
        self.tx_audio_status = QLabel("⚫ TX Audio: Not set")
        self.tx_audio_status.setStyleSheet("color: #607d8b;")
        status_layout.addWidget(self.tx_audio_status, 1, 0)
        
        self.tx_vara_status = QLabel("⚫ VARA: Not connected")
        self.tx_vara_status.setStyleSheet("color: #607d8b;")
        status_layout.addWidget(self.tx_vara_status, 1, 1)
        
        self.tx_rf_status = QLabel("⚫ RF: Not connected")
        self.tx_rf_status.setStyleSheet("color: #607d8b;")
        status_layout.addWidget(self.tx_rf_status, 2, 0)
        
        self.tx_aprs_is_status = QLabel("⚫ APRS-IS: Not connected")
        self.tx_aprs_is_status.setStyleSheet("color: #607d8b;")
        status_layout.addWidget(self.tx_aprs_is_status, 2, 1)
        
        goto_settings_btn = QPushButton("⚙️ Settings")
        goto_settings_btn.setFixedWidth(100)
        goto_settings_btn.clicked.connect(lambda: self.tabs.setCurrentIndex(4))
        status_layout.addWidget(goto_settings_btn, 3, 0)
        
        refresh_status_btn = QPushButton("🔄 Refresh")
        refresh_status_btn.setFixedWidth(100)
        refresh_status_btn.clicked.connect(self._sync_beacon_connection_status)
        status_layout.addWidget(refresh_status_btn, 3, 1)
        
        left_layout.addWidget(status_grp)
        
        # Beacon Settings Group
        beacon_grp = QGroupBox("📍 Beacon Settings")
        beacon_grp.setStyleSheet(self._group_style())
        beacon_layout = QGridLayout(beacon_grp)
        beacon_layout.setSpacing(8)
        
        # Row 0: Callsign and SSID
        beacon_layout.addWidget(QLabel("Callsign:"), 0, 0)
        call_ssid_layout = QHBoxLayout()
        self.callsign_edit = QLineEdit("N0CALL")
        self.callsign_edit.setMaxLength(6)
        self.callsign_edit.setFixedWidth(80)
        self.callsign_edit.setPlaceholderText("N0CALL")
        call_ssid_layout.addWidget(self.callsign_edit)
        call_ssid_layout.addWidget(QLabel("-"))
        self.ssid_combo = QComboBox()
        self.ssid_combo.addItem("0  Primary/Main", 0)
        self.ssid_combo.addItem("1  Secondary", 1)
        self.ssid_combo.addItem("2  Secondary radio", 2)
        self.ssid_combo.addItem("3  Additional", 3)
        self.ssid_combo.addItem("4  Additional", 4)
        self.ssid_combo.addItem("5  IGate/Gateway", 5)
        self.ssid_combo.addItem("6  Satellite/Special", 6)
        self.ssid_combo.addItem("7  Handheld (HT)", 7)
        self.ssid_combo.addItem("8  Boat/Maritime", 8)
        self.ssid_combo.addItem("9  Mobile/Vehicle", 9)
        self.ssid_combo.addItem("10 Internet/APRS-IS", 10)
        self.ssid_combo.addItem("11 Balloon/Aircraft", 11)
        self.ssid_combo.addItem("12 Portable/Field", 12)
        self.ssid_combo.addItem("13 Weather station", 13)
        self.ssid_combo.addItem("14 Truck/Large veh", 14)
        self.ssid_combo.addItem("15 Digipeater", 15)
        self.ssid_combo.setCurrentIndex(9)  # Default to mobile
        self.ssid_combo.currentIndexChanged.connect(self._sync_beacon_ssid_to_settings)
        self.ssid_combo.setToolTip(
            "SSID identifies your station type on the APRS network.\n"
            "Common values:\n"
            "  9 = Mobile/Vehicle\n"
            "  7 = Handheld (HT)\n"
            "  5 = IGate/Gateway\n"
            " 15 = Digipeater"
        )
        call_ssid_layout.addWidget(self.ssid_combo)
        call_ssid_layout.addStretch()
        beacon_layout.addLayout(call_ssid_layout, 0, 1)
        
        # Row 1: Location (read-only, controlled by Settings/GPS)
        self.location_label = QLabel("Location:")
        beacon_layout.addWidget(self.location_label, 1, 0)
        loc_layout = QHBoxLayout()
        self.lat_edit = QDoubleSpinBox()
        self.lat_edit.setRange(-90, 90)
        self.lat_edit.setDecimals(6)
        self.lat_edit.setValue(34.0522)
        self.lat_edit.setPrefix("Lat ")
        self.lat_edit.setFixedWidth(130)
        self.lat_edit.setEnabled(False)  # Read-only - controlled by Settings
        loc_layout.addWidget(self.lat_edit)
        self.lon_edit = QDoubleSpinBox()
        self.lon_edit.setRange(-180, 180)
        self.lon_edit.setDecimals(6)
        self.lon_edit.setValue(-118.2437)
        self.lon_edit.setPrefix("Lon ")
        self.lon_edit.setFixedWidth(140)
        self.lon_edit.setEnabled(False)  # Read-only - controlled by Settings
        loc_layout.addWidget(self.lon_edit)
        
        # GPS source indicator - prominent with large icon
        self.gps_source_label = QLabel("📍 Manual")
        self.gps_source_label.setStyleSheet("""
            color: #ffb74d; 
            font-weight: bold; 
            font-size: 14px;
            padding: 2px 8px;
            background: #2a2a1a;
            border-radius: 4px;
        """)
        self.gps_source_label.setToolTip("Location source controlled in Settings tab")
        loc_layout.addWidget(self.gps_source_label)
        loc_layout.addStretch()
        beacon_layout.addLayout(loc_layout, 1, 1)
        
        # Row 2: Symbol and Path on same row
        beacon_layout.addWidget(QLabel("Symbol:"), 2, 0)
        sym_path_layout = QHBoxLayout()
        # Hidden fields to store symbol data (used by symbol picker)
        self.symbol_table_combo = QComboBox()
        self.symbol_table_combo.addItems(["/", "\\"])
        self.symbol_table_combo.hide()
        self.symbol_code_edit = QLineEdit(">")
        self.symbol_code_edit.setMaxLength(1)
        self.symbol_code_edit.hide()
        # Icon preview
        self.symbol_preview = QLabel()
        self.symbol_preview.setFixedSize(28, 28)
        self.symbol_preview.setStyleSheet("background: #1a3a5c; border: 1px solid #42a5f5; border-radius: 4px;")
        self.symbol_preview.setToolTip("Select from Symbol Picker tab")
        sym_path_layout.addWidget(self.symbol_preview)
        sym_path_layout.addWidget(QLabel("Path:"))
        self.path_combo = QComboBox()
        self.path_combo.setEditable(True)
        self.path_combo.addItems([
            "WIDE1-1",
            "WIDE1-1,WIDE2-1", 
            "WIDE1-1,WIDE2-2",
            "WIDE2-1",
            "DIRECT",
        ])
        self.path_combo.setCurrentIndex(1)  # Default to WIDE1-1,WIDE2-1
        self.path_combo.setToolTip("Digipeater path (VARA FM requires DIRECT)")
        self.path_combo.setMinimumWidth(150)
        self.path_combo.currentTextChanged.connect(self._on_path_changed)
        sym_path_layout.addWidget(self.path_combo)
        sym_path_layout.addStretch()
        beacon_layout.addLayout(sym_path_layout, 2, 1)
        
        # Row 3: Radio
        beacon_layout.addWidget(QLabel("Radio:"), 3, 0)
        self.radio_combo = QComboBox()
        self.radio_combo.setEditable(True)
        self.radio_combo.addItems([
            "",
            "Yaesu FT-991A",
            "Yaesu FT-991",
            "Yaesu FT-891",
            "Yaesu FT-710",
            "Yaesu FTM-500D",
            "Yaesu FTM-400XD",
            "Yaesu FT-5D",
            "Yaesu FT-3D",
            "Icom IC-705",
            "Icom IC-7100",
            "Icom IC-7300",
            "Kenwood TM-D710",
            "Kenwood TH-D74",
        ])
        self.radio_combo.setToolTip("Your radio (optional)")
        beacon_layout.addWidget(self.radio_combo, 3, 1)
        
        # Row 4: Comment
        beacon_layout.addWidget(QLabel("Comment:"), 4, 0)
        self.comment_edit = QLineEdit("PyTNC Pro")
        beacon_layout.addWidget(self.comment_edit, 4, 1)
        
        # Hidden TX Audio combo (for compatibility - actual control in Settings tab)
        self.tx_audio_combo = QComboBox()
        self.tx_audio_combo.hide()
        
        # Hidden TX Level slider (for compatibility - actual control in Settings tab)
        self.tx_level_slider = QSlider(Qt.Orientation.Horizontal)
        self.tx_level_slider.setMinimum(1)
        self.tx_level_slider.setMaximum(100)
        self.tx_level_slider.setValue(100)  # Default to 100% - adjust in Settings
        self.tx_level_slider.hide()
        self.tx_level_label = QLabel("100%")
        self.tx_level_label.hide()
        
        left_layout.addWidget(beacon_grp)
        
        # Load TX output devices (for hidden combo compatibility)
        self.load_tx_devices()
        
        # Save Settings button
        self.save_settings_btn = QPushButton("💾 Save")
        self.save_settings_btn.setFixedWidth(80)
        self.save_settings_btn.setMinimumHeight(28)
        self.save_settings_btn.clicked.connect(self.save_settings)
        self.save_settings_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ff9800, stop:1 #f57c00);
                color: white;
                font-weight: bold;
                font-size: 10px;
                border: 1px solid #ffb74d;
                border-radius: 4px;
            }
            QPushButton:hover {
                background: #ffb74d;
            }
        """)
        left_layout.addWidget(self.save_settings_btn)
        
        # Beacon send buttons - 3 options
        beacon_grp = QGroupBox("📤 Send Beacon")
        beacon_grp.setStyleSheet(self._group_style())
        beacon_layout = QVBoxLayout(beacon_grp)
        beacon_layout.setSpacing(6)
        
        # APRS-IS button
        self.beacon_is_btn = QPushButton("🌐 Send via APRS-IS")
        self.beacon_is_btn.setMinimumHeight(40)
        self.beacon_is_btn.setToolTip("Send beacon via internet (APRS-IS)")
        self.beacon_is_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #1565c0, stop:1 #0d47a1);
                color: white;
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #42a5f5;
                border-radius: 6px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #1976d2, stop:1 #1565c0);
            }
        """)
        self.beacon_is_btn.clicked.connect(self._send_beacon_aprs_is)
        beacon_layout.addWidget(self.beacon_is_btn)
        
        # RF button
        self.beacon_btn = QPushButton("📡 Send via RF")
        self.beacon_btn.setMinimumHeight(40)
        self.beacon_btn.setToolTip("Send beacon via audio/RF")
        self.beacon_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #d32f2f, stop:1 #b71c1c);
                color: white;
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #ef5350;
                border-radius: 6px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #e53935, stop:1 #c62828);
            }
        """)
        self.beacon_btn.clicked.connect(self.send_beacon)
        beacon_layout.addWidget(self.beacon_btn)
        
        left_layout.addWidget(beacon_grp)
        
        # Auto-Beacon Group
        auto_beacon_grp = QGroupBox("⏱️ Auto-Beacon")
        auto_beacon_grp.setStyleSheet(self._group_style())
        auto_layout = QGridLayout(auto_beacon_grp)
        auto_layout.setSpacing(6)
        
        # Enable checkbox
        self.auto_beacon_enabled = QCheckBox("Enable auto-beacon")
        self.auto_beacon_enabled.setToolTip("Automatically send beacon at regular intervals")
        self.auto_beacon_enabled.stateChanged.connect(self._toggle_auto_beacon)
        auto_layout.addWidget(self.auto_beacon_enabled, 0, 0, 1, 2)
        
        # Interval
        auto_layout.addWidget(QLabel("Interval:"), 1, 0)
        interval_layout = QHBoxLayout()
        self.auto_beacon_interval = QSpinBox()
        self.auto_beacon_interval.setRange(1, 60)
        self.auto_beacon_interval.setValue(10)
        self.auto_beacon_interval.setSuffix(" min")
        self.auto_beacon_interval.setToolTip("Beacon interval in minutes")
        self.auto_beacon_interval.valueChanged.connect(self._update_auto_beacon_interval)
        interval_layout.addWidget(self.auto_beacon_interval)
        interval_layout.addStretch()
        auto_layout.addLayout(interval_layout, 1, 1)
        
        # Mode (RF, APRS-IS, or Both)
        auto_layout.addWidget(QLabel("Mode:"), 2, 0)
        self.auto_beacon_mode = QComboBox()
        self.auto_beacon_mode.addItem("APRS-IS only", "is")
        self.auto_beacon_mode.addItem("RF only", "rf")
        self.auto_beacon_mode.addItem("Both RF + APRS-IS", "both")
        self.auto_beacon_mode.setToolTip("How to send auto-beacons")
        auto_layout.addWidget(self.auto_beacon_mode, 2, 1)
        
        # Status/countdown label
        self.auto_beacon_status = QLabel("Auto-beacon: Off")
        self.auto_beacon_status.setStyleSheet("color: #607d8b;")
        auto_layout.addWidget(self.auto_beacon_status, 3, 0, 1, 2)
        
        left_layout.addWidget(auto_beacon_grp)
        
        # Initialize auto-beacon timer
        self.auto_beacon_timer = QTimer()
        self.auto_beacon_timer.timeout.connect(self._auto_beacon_tick)
        self.auto_beacon_countdown = 0
        
        left_layout.addStretch()
        
        settings_layout.addWidget(left_panel, 1)
        
        # Right panel - Symbol picker and TX Log
        right_panel = QWidget()
        right_panel.setMinimumWidth(480)  # Match VARA FM width
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(5)
        
        # Table selector at very top (outside group box)
        table_layout = QHBoxLayout()
        table_layout.addWidget(QLabel("Table:"))
        self.symbol_table_picker = QComboBox()
        self.symbol_table_picker.addItems(["/ Primary", "\\ Secondary"])
        self.symbol_table_picker.currentIndexChanged.connect(self._update_symbol_grid)
        table_layout.addWidget(self.symbol_table_picker)

        # Overlay character field
        table_layout.addSpacing(12)
        overlay_lbl = QLabel("Overlay:")
        overlay_lbl.setStyleSheet("color: #b0bec5; font-size: 11px;")
        table_layout.addWidget(overlay_lbl)
        self.symbol_overlay_edit = QLineEdit()
        self.symbol_overlay_edit.setMaxLength(1)
        self.symbol_overlay_edit.setPlaceholderText("A-Z")
        self.symbol_overlay_edit.setFixedWidth(36)
        self.symbol_overlay_edit.setToolTip(
            "Optional overlay letter (A-Z, 0-9)\n"
            "Example: I = IGate, D = Digipeater, W = Winlink\n"
            "Leave blank for standard primary/secondary symbol"
        )
        self.symbol_overlay_edit.setStyleSheet("""
            QLineEdit {
                background: #0a1929; color: #ffd54f; font-weight: bold;
                border: 1px solid #42a5f5; border-radius: 3px;
                padding: 2px 4px; font-size: 13px; text-align: center;
            }
        """)
        self.symbol_overlay_edit.textChanged.connect(self._on_overlay_changed)
        table_layout.addWidget(self.symbol_overlay_edit)
        table_layout.addStretch()
        right_layout.addLayout(table_layout)
        
        # Symbol Picker Group
        symbol_grp = QGroupBox("🎨 Symbol Picker")
        symbol_grp.setStyleSheet(self._group_style())
        symbol_layout = QVBoxLayout(symbol_grp)
        
        # Create scrollable grid for symbols
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea { border: 1px solid #1e3a5f; border-radius: 4px; background: #050d18; }
        """)
        
        self.symbol_grid_widget = QWidget()
        self.symbol_grid_widget.setStyleSheet("background: #050d18;")
        self.symbol_grid_layout = QGridLayout(self.symbol_grid_widget)
        self.symbol_grid_layout.setSpacing(1)
        self.symbol_grid_layout.setContentsMargins(2, 2, 2, 2)
        
        scroll.setWidget(self.symbol_grid_widget)
        symbol_layout.addWidget(scroll)
        
        right_layout.addWidget(symbol_grp, 1)  # stretch factor 1
        
        # TX Log - bigger, takes remaining space
        log_grp = QGroupBox("📝 TX Log")
        log_grp.setStyleSheet(self._group_style())
        log_layout = QVBoxLayout(log_grp)
        
        self.preset_log = QTextEdit()
        self.preset_log.setReadOnly(True)
        self.preset_log.setStyleSheet("""
            QTextEdit {
                background: #000000; color: #00ff00;
                font-family: 'Consolas', 'Courier New', monospace; font-size: 12px;
                border: 2px solid #1e3a5f; border-radius: 4px; padding: 8px;
            }
        """)
        log_layout.addWidget(self.preset_log)
        
        clear_log_btn = QPushButton("🗑️ Clear Log")
        clear_log_btn.setFixedWidth(100)
        clear_log_btn.clicked.connect(lambda: self.preset_log.clear())
        log_layout.addWidget(clear_log_btn)
        
        right_layout.addWidget(log_grp, 1)  # stretch factor 1 - same size as symbol picker
        right_layout.addWidget(self._branding_label())  # Inside right panel like VARA FM
        
        # Build initial symbol grid
        QTimer.singleShot(100, self._build_symbol_grid)
        
        settings_layout.addWidget(right_panel, 1)
        
        self.tabs.addTab(settings_tab, "📻 APRS")
        
        # Initialize message storage (for future use)
        self.conversations_file = BASE_DIR / "pytnc_conversations.json"
        self.conversations = {}  # {callsign: [{"from": x, "to": y, "text": z, "time": t, "acked": bool}]}
        self.current_conv = None
        self.msg_seq = 0  # Message sequence number for acks
        
        # Custom locations (loaded from CSV)
        self.custom_locations = []  # [{"name": x, "lat": y, "lon": z, "symbol": s, "comment": c}]
        
        # Dummy message UI elements (Messages tab removed, but code still references them)
        self.msg_status = QLabel()
        self.msg_to_edit = QLineEdit()
        self.msg_text_edit = QLineEdit()
        self.msg_header = QLabel()
        self.msg_history = QTextEdit()
        self.conv_list = QListWidget()
        
        # =====================================================================
        # SETTINGS TAB (build first to create widget)
        # =====================================================================
        self._build_config_tab()
        
        # =====================================================================
        # VARA FM TAB 
        # =====================================================================
        self._build_vara_tab()

        # =====================================================================
        # IGATE TAB
        # =====================================================================
        self._build_igate_tab()

        # =====================================================================
        # INFO TAB - Links & Resources
        # =====================================================================
        self._build_info_tab()
    
    def _build_info_tab(self):
        """Build the Info/Links tab with useful ham radio resources"""
        info_tab = QWidget()
        info_layout = QVBoxLayout(info_tab)
        info_layout.setContentsMargins(15, 15, 15, 15)
        info_layout.setSpacing(10)
        
        # Title
        title = QLabel("📻 PyTNC Pro - Ham Radio Resources")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #64b5f6;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_layout.addWidget(title)
        
        # Main content in horizontal layout
        content_layout = QHBoxLayout()
        
        # LEFT - Built-in links
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        links_widget = QWidget()
        self.links_layout = QVBoxLayout(links_widget)
        self.links_layout.setSpacing(6)
        
        # Load custom links
        self.custom_links_file = BASE_DIR / "pytnc_links.json"
        self.custom_links = self._load_custom_links()
        
        # Define built-in link categories - Emergency Comms FIRST
        builtin_categories = [
            ("🚨 Emergency Comms", [
                ("LAXNORTHEAST", "https://www.laxnortheast.org/"),
                ("LAXNORTHEAST Radio Plan", "https://docs.google.com/spreadsheets/d/1LGbFTBhhlHhICyrq31NAcdWQqQxdpF2E0W3g7aA2oxc/edit?gid=480461466#gid=480461466"),
                ("LA County EMS Agency", "https://dhs.lacounty.gov/emergency-medical-services-agency/"),
                ("FEMA", "https://www.fema.gov"),
                ("ARRL ARES", "http://www.arrl.org/ares"),
                ("Ready.gov", "https://www.ready.gov"),
            ]),
            ("🌪️ Weather & Alerts", [
                ("NWS Weather Alerts", "https://alerts.weather.gov"),
                ("USGS Earthquakes", "https://earthquake.usgs.gov"),
                ("SKYWARN Info", "https://www.weather.gov/skywarn"),
            ]),
            ("🗺️ APRS Resources", [
                ("aprs.fi - Live APRS Map", "https://aprs.fi"),
                ("APRS Direct", "https://www.aprsdirect.com"),
                ("FindU APRS", "https://www.findu.com"),
            ]),
            ("📡 Repeaters & Frequencies", [
                ("RadioReference", "https://www.radioreference.com"),
                ("ARRL Band Plan", "https://www.arrl.org/band-plan"),
            ]),
            ("📚 Learning & License", [
                ("ARRL - Ham Radio", "https://www.arrl.org"),
                ("QRZ Callsign Lookup", "https://www.qrz.com"),
                ("HamStudy.org", "https://hamstudy.org"),
            ]),
            ("🔧 Software & Tools", [
                ("Winlink", "https://winlink.org"),
                ("VARA FM", "https://rosmodem.wordpress.com"),
                ("Direwolf TNC", "https://github.com/wb2osz/direwolf"),
            ]),
        ]
        
        # Render all links
        self._render_info_links(builtin_categories)
        
        self.links_layout.addStretch()
        left_scroll.setWidget(links_widget)
        content_layout.addWidget(left_scroll, 2)
        
        # RIGHT - Add custom link panel
        right_panel = QWidget()
        right_panel.setFixedWidth(420)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(10, 0, 0, 0)
        right_layout.setSpacing(8)
        
        add_grp = QGroupBox("➕ Add Custom Link")
        add_grp.setStyleSheet(self._group_style())
        add_layout = QVBoxLayout(add_grp)
        add_layout.setSpacing(6)
        
        # Category dropdown
        cat_row = QHBoxLayout()
        cat_row.addWidget(QLabel("Category:"))
        self.link_category_combo = QComboBox()
        self.link_category_combo.addItems([
            "🚨 Emergency Comms",
            "🌪️ Weather & Alerts", 
            "🗺️ APRS Resources",
            "📡 Repeaters & Frequencies",
            "📚 Learning & License",
            "🔧 Software & Tools",
            "⭐ My Links"
        ])
        self.link_category_combo.setCurrentIndex(6)  # Default to "My Links"
        cat_row.addWidget(self.link_category_combo, 1)
        add_layout.addLayout(cat_row)
        
        # Name input
        add_layout.addWidget(QLabel("Name:"))
        self.link_name_edit = QLineEdit()
        self.link_name_edit.setPlaceholderText("e.g., My Club Website")
        add_layout.addWidget(self.link_name_edit)
        
        # URL input
        add_layout.addWidget(QLabel("URL:"))
        self.link_url_edit = QLineEdit()
        self.link_url_edit.setPlaceholderText("https://...")
        add_layout.addWidget(self.link_url_edit)
        
        # Add button
        add_btn = QPushButton("💾 Save Link")
        add_btn.setStyleSheet(self._button_style("#2e7d32", "#388e3c"))
        add_btn.clicked.connect(self._add_custom_link)
        add_layout.addWidget(add_btn)
        
        right_layout.addWidget(add_grp)
        
        # Custom links list
        custom_grp = QGroupBox("⭐ My Custom Links")
        custom_grp.setStyleSheet(self._group_style())
        custom_layout = QVBoxLayout(custom_grp)
        
        self.custom_links_list = QListWidget()
        self.custom_links_list.setMinimumHeight(350)
        self.custom_links_list.setStyleSheet("background: #0d1117; border: 1px solid #1e3a5f; font-size: 12px;")
        self.custom_links_list.itemDoubleClicked.connect(self._open_custom_link)
        custom_layout.addWidget(self.custom_links_list)
        
        del_btn = QPushButton("🗑️ Delete Selected")
        del_btn.setStyleSheet(self._button_style("#c62828", "#d32f2f"))
        del_btn.clicked.connect(self._delete_custom_link)
        custom_layout.addWidget(del_btn)
        
        right_layout.addWidget(custom_grp)
        right_layout.addStretch()
        
        content_layout.addWidget(right_panel)
        info_layout.addLayout(content_layout, 1)
        
        # Version info at bottom
        version_label = QLabel(f"PyTNC Pro by KO6IKR © 2026")
        version_label.setStyleSheet("color: #607d8b; font-size: 11px;")
        version_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        info_layout.addWidget(version_label)
        
        self.tabs.addTab(info_tab, "ℹ️ Info")
        
        # Add Settings tab last
        self.tabs.addTab(self._settings_tab_widget, "⚙️ Settings")
        
        # Populate custom links list
        self._refresh_custom_links_list()
    
    def _render_info_links(self, categories):
        """Render link categories"""
        for cat_name, links in categories:
            cat_label = QLabel(cat_name)
            cat_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #81c784; margin-top: 8px;")
            self.links_layout.addWidget(cat_label)
            
            for name, url in links:
                link_btn = QPushButton(f"  🔗 {name}")
                link_btn.setStyleSheet("""
                    QPushButton {
                        background: transparent;
                        color: #64b5f6;
                        text-align: left;
                        padding: 2px 8px;
                        border: none;
                        font-size: 11px;
                    }
                    QPushButton:hover {
                        color: #90caf9;
                        text-decoration: underline;
                    }
                """)
                link_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                link_btn.clicked.connect(lambda checked, u=url: self._open_url(u))
                self.links_layout.addWidget(link_btn)
        
        # Add custom links by category
        if self.custom_links:
            for cat, links in self.custom_links.items():
                if links:
                    cat_label = QLabel(cat)
                    cat_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #ffd54f; margin-top: 8px;")
                    self.links_layout.addWidget(cat_label)
                    
                    for link in links:
                        link_btn = QPushButton(f"  ⭐ {link['name']}")
                        link_btn.setStyleSheet("""
                            QPushButton {
                                background: transparent;
                                color: #ffd54f;
                                text-align: left;
                                padding: 2px 8px;
                                border: none;
                                font-size: 11px;
                            }
                            QPushButton:hover {
                                color: #ffecb3;
                                text-decoration: underline;
                            }
                        """)
                        link_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                        link_btn.clicked.connect(lambda checked, u=link['url']: self._open_url(u))
                        self.links_layout.addWidget(link_btn)
    
    def _load_custom_links(self):
        """Load custom links from file"""
        if self.custom_links_file.exists():
            try:
                import json
                with open(self.custom_links_file, 'r') as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError) as e:
                # Corrupted or unreadable file - start fresh
                pass
        return {}
    
    def _save_custom_links(self):
        """Save custom links to file"""
        try:
            import json
            with open(self.custom_links_file, 'w') as f:
                json.dump(self.custom_links, f, indent=2)
        except Exception as e:
            self._log(f"❌ Failed to save links: {e}")
    
    def _add_custom_link(self):
        """Add a custom link"""
        name = self.link_name_edit.text().strip()
        url = self.link_url_edit.text().strip()
        category = self.link_category_combo.currentText()
        
        if not name or not url:
            self._log("❌ Enter both name and URL")
            return
        
        if not url.startswith("http"):
            url = "https://" + url
        
        if category not in self.custom_links:
            self.custom_links[category] = []
        
        self.custom_links[category].append({"name": name, "url": url})
        self._save_custom_links()
        
        # Clear inputs
        self.link_name_edit.clear()
        self.link_url_edit.clear()
        
        self._refresh_custom_links_list()
        self._log(f"✅ Added link: {name}")
    
    def _refresh_custom_links_list(self):
        """Refresh the custom links list widget"""
        self.custom_links_list.clear()
        for cat, links in self.custom_links.items():
            for link in links:
                item = QListWidgetItem(f"{link['name']}")
                item.setData(Qt.ItemDataRole.UserRole, link['url'])
                item.setToolTip(link['url'])
                self.custom_links_list.addItem(item)
    
    def _open_custom_link(self, item):
        """Open a custom link on double-click"""
        url = item.data(Qt.ItemDataRole.UserRole)
        if url:
            self._open_url(url)
    
    def _delete_custom_link(self):
        """Delete selected custom link"""
        item = self.custom_links_list.currentItem()
        if not item:
            return
        
        name = item.text()
        url = item.data(Qt.ItemDataRole.UserRole)
        
        # Find and remove
        for cat, links in self.custom_links.items():
            for link in links:
                if link['name'] == name and link['url'] == url:
                    links.remove(link)
                    break
        
        self._save_custom_links()
        self._refresh_custom_links_list()
        self._log(f"🗑️ Deleted link: {name}")
    
    def _open_url(self, url):
        """Open URL in default browser"""
        try:
            import webbrowser
            webbrowser.open(url)
            self._log(f"🌐 Opening: {url}")
        except Exception as e:
            self._log(f"❌ Failed to open URL: {e}")
    
    def _build_config_tab(self):
        """Build the Settings tab with all connection settings"""
        settings_tab = QWidget()
        main_layout = QHBoxLayout(settings_tab)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)
        
        # LEFT COLUMN - Hardware connections
        left_col = QVBoxLayout()
        left_col.setSpacing(6)
        
        # === PTT PORT ===
        ptt_grp = QGroupBox("🎙️ PTT Control")
        ptt_grp.setStyleSheet(self._group_style())
        ptt_layout = QGridLayout(ptt_grp)
        ptt_layout.setSpacing(4)
        
        # Row 0: PTT Method selector
        ptt_method_widget = QWidget()
        ptt_method_layout = QHBoxLayout(ptt_method_widget)
        ptt_method_layout.setContentsMargins(0, 0, 0, 0)
        ptt_method_layout.addWidget(QLabel("Method:"))
        self.ptt_method_combo = QComboBox()
        self.ptt_method_combo.addItems(["RTS/DTR", "CI-V CAT", "CM108 GPIO"])
        self.ptt_method_combo.setToolTip(
            "RTS/DTR: standard serial PTT (DigiRig Mobile, etc.)\n"
            "CI-V CAT: Icom radio CAT control (IC-7100, IC-9700, etc.)\n"
            "CM108 GPIO: USB audio chip GPIO (DigiRig Lite, no COM port)"
        )
        self.ptt_method_combo.currentTextChanged.connect(self._on_ptt_method_changed)
        ptt_method_layout.addWidget(self.ptt_method_combo)
        ptt_method_layout.addStretch()
        ptt_layout.addWidget(ptt_method_widget, 0, 0, 1, 4)

        # Row 1: Serial port settings (RTS/DTR mode)
        self.ptt_serial_widget = QWidget()
        ptt_serial_layout = QHBoxLayout(self.ptt_serial_widget)
        ptt_serial_layout.setContentsMargins(0, 0, 0, 0)
        ptt_serial_layout.addWidget(QLabel("Port:"))
        self.settings_ptt_combo = QComboBox()
        self._populate_serial_combo(self.settings_ptt_combo)
        ptt_serial_layout.addWidget(self.settings_ptt_combo)
        self.settings_ptt_btn = QPushButton("Connect")
        self.settings_ptt_btn.setFixedWidth(70)
        self.settings_ptt_btn.clicked.connect(self._toggle_ptt)
        ptt_serial_layout.addWidget(self.settings_ptt_btn)
        self.settings_ptt_status = QLabel("⚫")
        ptt_serial_layout.addWidget(self.settings_ptt_status)
        ptt_layout.addWidget(self.ptt_serial_widget, 1, 0, 1, 4)

        # Row 2: RTS/DTR line settings
        self.ptt_lines_widget = QWidget()
        ptt_lines_layout = QHBoxLayout(self.ptt_lines_widget)
        ptt_lines_layout.setContentsMargins(0, 0, 0, 0)
        ptt_lines_layout.addWidget(QLabel("RTS:"))
        self.ptt_rts_combo = QComboBox()
        self.ptt_rts_combo.addItems(["Off", "High=TX", "Low=TX"])
        ptt_lines_layout.addWidget(self.ptt_rts_combo)
        ptt_lines_layout.addWidget(QLabel("DTR:"))
        self.ptt_dtr_combo = QComboBox()
        self.ptt_dtr_combo.addItems(["Off", "High=TX", "Low=TX"])
        self.ptt_dtr_combo.setCurrentIndex(1)  # Default DTR High=TX
        ptt_lines_layout.addWidget(self.ptt_dtr_combo)
        ptt_layout.addWidget(self.ptt_lines_widget, 2, 0, 1, 4)

        # Row 3: CI-V CAT config widget (hidden by default)
        self.civ_widget = QWidget()
        civ_layout = QGridLayout(self.civ_widget)
        civ_layout.setContentsMargins(0, 4, 0, 0)
        civ_layout.setSpacing(4)

        # Row 3a: Port + Baud + Connect
        civ_layout.addWidget(QLabel("CI-V Port:"), 0, 0)
        self.civ_port_combo = QComboBox()
        self._populate_serial_combo(self.civ_port_combo)
        civ_layout.addWidget(self.civ_port_combo, 0, 1)
        self.civ_baud_combo = QComboBox()
        self.civ_baud_combo.addItems(["4800", "9600", "19200", "38400"])
        self.civ_baud_combo.setCurrentText("19200")
        self.civ_baud_combo.setFixedWidth(70)
        self.civ_baud_combo.setToolTip("Baud rate — IC-7100 default: 19200")
        civ_layout.addWidget(self.civ_baud_combo, 0, 2)
        self.civ_connect_btn = QPushButton("Connect")
        self.civ_connect_btn.setFixedWidth(70)
        self.civ_connect_btn.clicked.connect(self._toggle_civ)
        civ_layout.addWidget(self.civ_connect_btn, 0, 3)
        self.civ_status = QLabel("⚫")
        civ_layout.addWidget(self.civ_status, 0, 4)

        # Row 3b: Data bits, Parity, Stop bits, CI-V address
        civ_layout.addWidget(QLabel("Data:"), 1, 0)
        self.civ_data_combo = QComboBox()
        self.civ_data_combo.addItems(["8", "7"])
        self.civ_data_combo.setFixedWidth(44)
        civ_layout.addWidget(self.civ_data_combo, 1, 1)
        self.civ_parity_combo = QComboBox()
        self.civ_parity_combo.addItems(["None", "Even", "Odd"])
        self.civ_parity_combo.setFixedWidth(60)
        self.civ_parity_combo.setToolTip("Parity — Icom default: None")
        civ_layout.addWidget(self.civ_parity_combo, 1, 2)
        self.civ_stop_combo = QComboBox()
        self.civ_stop_combo.addItems(["1", "2"])
        self.civ_stop_combo.setFixedWidth(44)
        self.civ_stop_combo.setToolTip("Stop bits — Icom default: 1")
        civ_layout.addWidget(self.civ_stop_combo, 1, 3)

        civ_addr_widget = QWidget()
        civ_addr_layout = QHBoxLayout(civ_addr_widget)
        civ_addr_layout.setContentsMargins(0, 0, 0, 0)
        civ_addr_layout.addWidget(QLabel("CI-V Addr (hex):"))
        self.civ_addr_edit = QLineEdit("88")
        self.civ_addr_edit.setFixedWidth(40)
        self.civ_addr_edit.setMaxLength(2)
        self.civ_addr_edit.setToolTip("CI-V address in hex — IC-7100: 88, IC-9700: A2, IC-7300: 94")
        civ_addr_layout.addWidget(self.civ_addr_edit)
        civ_addr_layout.addStretch()
        civ_layout.addWidget(civ_addr_widget, 1, 4)

        self.civ_widget.hide()
        ptt_layout.addWidget(self.civ_widget, 3, 0, 1, 4)

        # Row 4: CM108 GPIO config widget (hidden by default)
        self.cm108_widget = QWidget()
        cm108_layout = QHBoxLayout(self.cm108_widget)
        cm108_layout.setContentsMargins(0, 4, 0, 0)
        cm108_layout.setSpacing(6)
        cm108_layout.addWidget(QLabel("Device:"))
        self.cm108_device_combo = QComboBox()
        self.cm108_device_combo.setToolTip("CM108/CM119 HID device — click Scan to detect")
        cm108_layout.addWidget(self.cm108_device_combo, 1)
        self.cm108_scan_btn = QPushButton("🔍 Scan")
        self.cm108_scan_btn.setFixedWidth(64)
        self.cm108_scan_btn.clicked.connect(self._cm108_scan)
        cm108_layout.addWidget(self.cm108_scan_btn)
        self.cm108_connect_btn = QPushButton("Connect")
        self.cm108_connect_btn.setFixedWidth(70)
        self.cm108_connect_btn.clicked.connect(self._toggle_cm108)
        cm108_layout.addWidget(self.cm108_connect_btn)
        self.cm108_status = QLabel("⚫")
        cm108_layout.addWidget(self.cm108_status)
        self.cm108_widget.hide()
        ptt_layout.addWidget(self.cm108_widget, 4, 0, 1, 4)
        
        # Row 2: Test PTT button
        self.ptt_test_btn = QPushButton("🔴 Test PTT")
        self.ptt_test_btn.setStyleSheet("""
            QPushButton { background: #c62828; color: white; font-weight: bold; border-radius: 4px; padding: 4px; }
            QPushButton:hover { background: #e53935; }
            QPushButton:pressed { background: #b71c1c; }
        """)
        self.ptt_test_btn.pressed.connect(self._ptt_test_on)
        self.ptt_test_btn.released.connect(self._ptt_test_off)
        ptt_layout.addWidget(self.ptt_test_btn, 5, 0, 1, 4)
        left_col.addWidget(ptt_grp)
        
        # === GPS PORT ===
        gps_grp = QGroupBox("📍 GPS / Location")
        gps_grp.setStyleSheet(self._group_style())
        gps_layout = QGridLayout(gps_grp)
        gps_layout.setSpacing(4)
        
        # Row 0: Port, baud, connect button
        self.settings_gps_combo = QComboBox()
        self._populate_serial_combo(self.settings_gps_combo)
        gps_layout.addWidget(self.settings_gps_combo, 0, 0)
        
        self.gps_baud_combo = QComboBox()
        self.gps_baud_combo.setToolTip("Baud rate (most GPS: 4800)")
        self.gps_baud_combo.addItem("4800", 4800)
        self.gps_baud_combo.addItem("9600", 9600)
        self.gps_baud_combo.addItem("38400", 38400)
        self.gps_baud_combo.setCurrentIndex(0)
        self.gps_baud_combo.setFixedWidth(60)
        gps_layout.addWidget(self.gps_baud_combo, 0, 1)
        
        self.settings_gps_btn = QPushButton("Connect")
        self.settings_gps_btn.setFixedWidth(60)
        self.settings_gps_btn.clicked.connect(self._toggle_gps)
        gps_layout.addWidget(self.settings_gps_btn, 0, 2)
        
        self.settings_gps_status = QLabel("⚫")
        self.settings_gps_status.setFixedWidth(20)
        gps_layout.addWidget(self.settings_gps_status, 0, 3)
        
        # Row 1: GPS Lat/Lon display (read-only, shows current GPS position)
        gps_layout.addWidget(QLabel("GPS:"), 1, 0)
        self.settings_gps_latlon = QLabel("---, ---")
        self.settings_gps_latlon.setStyleSheet("color: #69f0ae; font-family: Consolas; font-size: 11px;")
        self.settings_gps_latlon.setToolTip("Current GPS coordinates")
        gps_layout.addWidget(self.settings_gps_latlon, 1, 1, 1, 3)
        
        # Row 2: Manual location (smaller)
        gps_layout.addWidget(QLabel("Manual:"), 2, 0)
        self.manual_location = QLineEdit()
        self.manual_location.setPlaceholderText("34.05, -118.24")
        self.manual_location.setFixedWidth(140)
        gps_layout.addWidget(self.manual_location, 2, 1, 1, 2)
        self.apply_location_btn = QPushButton("Set")
        self.apply_location_btn.setFixedWidth(35)
        self.apply_location_btn.clicked.connect(self._apply_manual_location)
        gps_layout.addWidget(self.apply_location_btn, 2, 3)
        
        left_col.addWidget(gps_grp)
        
        # === AUDIO ===
        audio_grp = QGroupBox("🔊 Audio")
        audio_grp.setStyleSheet(self._group_style())
        audio_layout = QGridLayout(audio_grp)
        audio_layout.setSpacing(4)
        
        audio_layout.addWidget(QLabel("RX:"), 0, 0)
        self.settings_rx_audio_combo = QComboBox()
        self._populate_audio_inputs(self.settings_rx_audio_combo)
        audio_layout.addWidget(self.settings_rx_audio_combo, 0, 1)
        
        audio_layout.addWidget(QLabel("TX:"), 1, 0)
        self.settings_tx_audio_combo = QComboBox()
        self._populate_audio_outputs(self.settings_tx_audio_combo)
        self.settings_tx_audio_combo.currentIndexChanged.connect(self._on_tx_audio_changed)
        audio_layout.addWidget(self.settings_tx_audio_combo, 1, 1)
        
        audio_layout.addWidget(QLabel("RX Gain:"), 2, 0)
        rx_gain_layout = QHBoxLayout()
        self.settings_rx_gain = QSlider(Qt.Orientation.Horizontal)
        self.settings_rx_gain.setRange(1, 100)
        self.settings_rx_gain.setValue(10)
        self.settings_rx_gain.valueChanged.connect(self._on_settings_rx_gain)
        rx_gain_layout.addWidget(self.settings_rx_gain)
        self.settings_rx_gain_label = QLabel("1.0x")
        self.settings_rx_gain_label.setFixedWidth(35)
        rx_gain_layout.addWidget(self.settings_rx_gain_label)
        audio_layout.addLayout(rx_gain_layout, 2, 1)
        
        # TX Level control
        audio_layout.addWidget(QLabel("TX Level:"), 3, 0)
        tx_level_layout = QHBoxLayout()
        self.settings_tx_level = QSlider(Qt.Orientation.Horizontal)
        self.settings_tx_level.setRange(1, 100)
        self.settings_tx_level.setValue(100)  # Default 100% - adjust on radio if needed
        self.settings_tx_level.setToolTip("TX audio output level (usually keep at 100%, adjust on radio)")
        self.settings_tx_level.valueChanged.connect(self._on_settings_tx_level)
        tx_level_layout.addWidget(self.settings_tx_level)
        self.settings_tx_level_label = QLabel("100%")
        self.settings_tx_level_label.setFixedWidth(35)
        tx_level_layout.addWidget(self.settings_tx_level_label)
        audio_layout.addLayout(tx_level_layout, 3, 1)
        
        left_col.addWidget(audio_grp)
        
        left_col.addStretch()
        main_layout.addLayout(left_col)
        
        # MIDDLE COLUMN - Network & Station
        mid_col = QVBoxLayout()
        mid_col.setSpacing(6)
        
        # === APRS RANGE (prominent) ===
        range_grp = QGroupBox("📡 APRS Range")
        range_grp.setStyleSheet(self._group_style())
        range_layout = QHBoxLayout(range_grp)
        range_layout.setSpacing(8)
        
        range_layout.addWidget(QLabel("Show stations within:"))
        self.settings_aprs_radius = QSpinBox()
        self.settings_aprs_radius.setRange(10, 500)
        self.settings_aprs_radius.setValue(100)
        self.settings_aprs_radius.setSuffix(" km")
        self.settings_aprs_radius.setToolTip("APRS-IS filter radius from your location")
        self.settings_aprs_radius.setFixedWidth(90)
        range_layout.addWidget(self.settings_aprs_radius)
        
        self.aprs_range_label = QLabel("(200 km diameter)")
        self.aprs_range_label.setStyleSheet("color: #888; font-size: 11px;")
        range_layout.addWidget(self.aprs_range_label)
        
        # Update diameter label when radius changes
        self.settings_aprs_radius.valueChanged.connect(
            lambda v: self.aprs_range_label.setText(f"({v*2} km diameter)")
        )
        
        range_layout.addStretch()
        mid_col.addWidget(range_grp)
        
        # === APRS-IS (compact) ===
        aprs_grp = QGroupBox("🌐 APRS-IS")
        aprs_grp.setStyleSheet(self._group_style())
        aprs_layout = QGridLayout(aprs_grp)
        aprs_layout.setSpacing(3)
        
        aprs_layout.addWidget(QLabel("Server:"), 0, 0)
        self.settings_aprs_server = QLineEdit("rotate.aprs2.net")
        self.settings_aprs_server.setFixedWidth(140)
        aprs_layout.addWidget(self.settings_aprs_server, 0, 1)
        self.settings_aprs_port = QSpinBox()
        self.settings_aprs_port.setRange(1, 65535)
        self.settings_aprs_port.setValue(14580)
        self.settings_aprs_port.setFixedWidth(65)
        aprs_layout.addWidget(self.settings_aprs_port, 0, 2)
        
        aprs_layout.addWidget(QLabel("Pass:"), 1, 0)
        self.settings_aprs_passcode = QLineEdit()
        self.settings_aprs_passcode.setPlaceholderText("-1")
        self.settings_aprs_passcode.setEchoMode(QLineEdit.EchoMode.Password)
        self.settings_aprs_passcode.setFixedWidth(60)
        aprs_layout.addWidget(self.settings_aprs_passcode, 1, 1)
        
        self.settings_aprs_connect_btn = QPushButton("Connect")
        self.settings_aprs_connect_btn.setFixedWidth(65)
        self.settings_aprs_connect_btn.clicked.connect(self._toggle_aprs_is_from_settings)
        aprs_layout.addWidget(self.settings_aprs_connect_btn, 1, 2)
        
        self.settings_aprs_status = QLabel("⚫ Disconnected")
        self.settings_aprs_status.setStyleSheet("color: #ef5350; font-size: 10px;")
        aprs_layout.addWidget(self.settings_aprs_status, 2, 0, 1, 3)
        mid_col.addWidget(aprs_grp)
        
        # === EARTHQUAKE MONITOR (compact) ===
        quake_grp = QGroupBox("🌋 Earthquakes")
        quake_grp.setStyleSheet(self._group_style())
        quake_layout = QGridLayout(quake_grp)
        quake_layout.setSpacing(3)
        
        self.quake_enabled = QCheckBox("Enable")
        self.quake_enabled.setChecked(False)
        self.quake_enabled.stateChanged.connect(self._toggle_earthquake_monitor)
        quake_layout.addWidget(self.quake_enabled, 0, 0)
        
        # Time range dropdown
        self.quake_time_range = QComboBox()
        self.quake_time_range.addItem("1 hour", "hour")
        self.quake_time_range.addItem("24 hours", "day")
        self.quake_time_range.addItem("7 days", "week")
        self.quake_time_range.addItem("30 days", "month")
        self.quake_time_range.setCurrentIndex(1)  # Default 24 hours
        self.quake_time_range.setFixedWidth(75)
        quake_layout.addWidget(self.quake_time_range, 0, 1)
        
        self.quake_refresh_btn = QPushButton("🔄")
        self.quake_refresh_btn.setFixedWidth(30)
        self.quake_refresh_btn.clicked.connect(self._fetch_earthquakes)
        self.quake_refresh_btn.setEnabled(False)
        quake_layout.addWidget(self.quake_refresh_btn, 0, 2)
        
        quake_layout.addWidget(QLabel("Radius:"), 1, 0)
        self.quake_radius = QSlider(Qt.Orientation.Horizontal)
        self.quake_radius.setRange(10, 1000)
        self.quake_radius.setValue(100)
        self.quake_radius.setFixedWidth(80)
        quake_layout.addWidget(self.quake_radius, 1, 1)
        self.quake_radius_label = QLabel("100mi")
        self.quake_radius_label.setFixedWidth(40)
        self.quake_radius.valueChanged.connect(lambda v: self.quake_radius_label.setText(f"{v}mi"))
        quake_layout.addWidget(self.quake_radius_label, 1, 2)
        
        quake_layout.addWidget(QLabel("Min M:"), 2, 0)
        self.quake_min_mag = QDoubleSpinBox()
        self.quake_min_mag.setRange(0.0, 9.0)
        self.quake_min_mag.setValue(2.0)
        self.quake_min_mag.setSingleStep(0.5)
        self.quake_min_mag.setDecimals(1)
        self.quake_min_mag.setFixedWidth(55)
        quake_layout.addWidget(self.quake_min_mag, 2, 1)
        
        self.quake_status = QLabel("⚫")
        self.quake_status.setStyleSheet("color: #888; font-size: 10px;")
        self.quake_status.setMinimumWidth(70)
        quake_layout.addWidget(self.quake_status, 2, 2)
        
        mid_col.addWidget(quake_grp)
        
        # === FIRE MONITOR (NASA FIRMS) ===
        fire_grp = QGroupBox("🔥 Wildfires (NASA)")
        fire_grp.setStyleSheet(self._group_style())
        fire_layout = QGridLayout(fire_grp)
        fire_layout.setSpacing(3)
        
        self.fire_enabled = QCheckBox("Enable")
        self.fire_enabled.setChecked(False)
        self.fire_enabled.stateChanged.connect(self._toggle_fire_monitor)
        fire_layout.addWidget(self.fire_enabled, 0, 0)
        
        # Time range dropdown (NASA FIRMS supports 24h, 48h, 7d)
        self.fire_time_range = QComboBox()
        self.fire_time_range.addItem("24 hours", "24h")
        self.fire_time_range.addItem("48 hours", "48h")
        self.fire_time_range.addItem("7 days", "7d")
        self.fire_time_range.setCurrentIndex(0)  # Default 24 hours
        self.fire_time_range.setFixedWidth(75)
        fire_layout.addWidget(self.fire_time_range, 0, 1)
        
        self.fire_refresh_btn = QPushButton("🔄")
        self.fire_refresh_btn.setFixedWidth(30)
        self.fire_refresh_btn.clicked.connect(self._fetch_fires)
        self.fire_refresh_btn.setEnabled(False)
        fire_layout.addWidget(self.fire_refresh_btn, 0, 2)
        
        # Satellite source (VIIRS is higher resolution)
        fire_layout.addWidget(QLabel("Source:"), 1, 0)
        self.fire_source = QComboBox()
        self.fire_source.addItem("VIIRS", "VIIRS_SNPP_NRT")
        self.fire_source.addItem("MODIS", "MODIS_NRT")
        self.fire_source.setFixedWidth(65)
        self.fire_source.setToolTip("VIIRS: 375m resolution, MODIS: 1km resolution")
        fire_layout.addWidget(self.fire_source, 1, 1)
        
        self.fire_status = QLabel("⚫")
        self.fire_status.setStyleSheet("color: #888; font-size: 10px;")
        self.fire_status.setMinimumWidth(70)
        fire_layout.addWidget(self.fire_status, 1, 2)
        
        # API Key (required for NASA FIRMS)
        fire_layout.addWidget(QLabel("API Key:"), 2, 0)
        self.fire_api_key = QLineEdit()
        self.fire_api_key.setPlaceholderText("Get free key from firms.modaps.eosdis.nasa.gov")
        self.fire_api_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.fire_api_key.setFixedWidth(140)
        fire_layout.addWidget(self.fire_api_key, 2, 1, 1, 2)
        
        mid_col.addWidget(fire_grp)
        
        # === WEATHER ALERTS ===
        weather_grp = QGroupBox("⚠️ Weather Alerts")
        weather_grp.setStyleSheet(self._group_style())
        weather_layout = QGridLayout(weather_grp)
        weather_layout.setSpacing(3)
        
        self.weather_enabled = QCheckBox("Show")
        self.weather_enabled.setChecked(False)
        self.weather_enabled.stateChanged.connect(self._toggle_weather_layer)
        weather_layout.addWidget(self.weather_enabled, 0, 0)
        
        self.weather_refresh_btn = QPushButton("🔄")
        self.weather_refresh_btn.setFixedWidth(30)
        self.weather_refresh_btn.setToolTip("Refresh weather alerts from NWS")
        self.weather_refresh_btn.clicked.connect(self._fetch_weather_alerts)
        self.weather_refresh_btn.setEnabled(False)
        weather_layout.addWidget(self.weather_refresh_btn, 0, 2)
        
        self.weather_status = QLabel("⚫")
        self.weather_status.setStyleSheet("color: #888; font-size: 10px;")
        self.weather_status.setMinimumWidth(70)
        weather_layout.addWidget(self.weather_status, 1, 0, 1, 3)
        
        mid_col.addWidget(weather_grp)
        
        # === AIR QUALITY (AQI) ===
        aqi_grp = QGroupBox("💨 Air Quality (AQI)")
        aqi_grp.setStyleSheet(self._group_style())
        aqi_layout = QGridLayout(aqi_grp)
        aqi_layout.setSpacing(3)
        
        aqi_layout.addWidget(QLabel("API Key:"), 0, 0)
        self.aqi_api_key = QLineEdit()
        self.aqi_api_key.setPlaceholderText("AirNow API key")
        self.aqi_api_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.aqi_api_key.setStyleSheet("background: #0d2137; color: white; border: 1px solid #1e3a5f; padding: 2px;")
        aqi_layout.addWidget(self.aqi_api_key, 0, 1)
        
        self.aqi_enabled = QCheckBox("Show")
        self.aqi_enabled.setChecked(False)
        self.aqi_enabled.stateChanged.connect(self._toggle_aqi_monitor)
        aqi_layout.addWidget(self.aqi_enabled, 1, 0)
        
        self.aqi_refresh_btn = QPushButton("🔄")
        self.aqi_refresh_btn.setFixedWidth(30)
        self.aqi_refresh_btn.setToolTip("Refresh AQI from AirNow")
        self.aqi_refresh_btn.clicked.connect(self._fetch_aqi_data)
        self.aqi_refresh_btn.setEnabled(False)
        aqi_layout.addWidget(self.aqi_refresh_btn, 1, 1)
        
        self.aqi_status = QLabel("⚫")
        self.aqi_status.setStyleSheet("color: #888; font-size: 10px;")
        self.aqi_status.setMinimumWidth(70)
        aqi_layout.addWidget(self.aqi_status, 2, 0, 1, 2)
        
        mid_col.addWidget(aqi_grp)
        
        # === OFFLINE CACHE ===
        cache_grp = QGroupBox("💾 Offline Cache")
        cache_grp.setStyleSheet(self._group_style())
        cache_layout = QGridLayout(cache_grp)
        cache_layout.setSpacing(3)
        
        # Hospitals with radius slider
        cache_layout.addWidget(QLabel("🏥 Hosp:"), 0, 0)
        self.hospital_radius = QSlider(Qt.Orientation.Horizontal)
        self.hospital_radius.setRange(5, 100)
        self.hospital_radius.setValue(25)
        self.hospital_radius.setFixedWidth(60)
        cache_layout.addWidget(self.hospital_radius, 0, 1)
        self.hospital_radius_label = QLabel("25mi")
        self.hospital_radius_label.setFixedWidth(35)
        self.hospital_radius.valueChanged.connect(lambda v: self.hospital_radius_label.setText(f"{v}mi"))
        cache_layout.addWidget(self.hospital_radius_label, 0, 2)
        self.hospital_status = QLabel("--")
        self.hospital_status.setStyleSheet("color: #888; font-size: 10px;")
        cache_layout.addWidget(self.hospital_status, 0, 3)
        self.hospital_refresh_btn = QPushButton("⬇️")
        self.hospital_refresh_btn.setFixedWidth(30)
        self.hospital_refresh_btn.setToolTip("Download hospitals")
        self.hospital_refresh_btn.clicked.connect(self._fetch_hospitals)
        cache_layout.addWidget(self.hospital_refresh_btn, 0, 4)
        # Hidden checkbox for internal sync (not shown)
        self.hospital_enabled = QCheckBox()
        self.hospital_enabled.setChecked(False)
        self.hospital_enabled.stateChanged.connect(self._toggle_hospital_layer)
        self.hospital_enabled.hide()
        self.hospital_offline_indicator = QLabel("")  # Keep for compatibility
        
        # Map tiles with zoom range selector
        cache_layout.addWidget(QLabel("🗺️ Map:"), 1, 0)
        self.cache_map_zoom_slider = QSlider(Qt.Orientation.Horizontal)
        self.cache_map_zoom_slider.setRange(12, 16)  # Max zoom level to cache
        self.cache_map_zoom_slider.setValue(14)
        self.cache_map_zoom_slider.setFixedWidth(60)
        self.cache_map_zoom_slider.setToolTip("z12=5MB, z13=19MB, z14=72MB, z15=279MB, z16=1GB")
        cache_layout.addWidget(self.cache_map_zoom_slider, 1, 1)
        self.cache_map_zoom_label = QLabel("z8-14")
        self.cache_map_zoom_label.setFixedWidth(40)
        self.cache_map_zoom_slider.valueChanged.connect(lambda v: self.cache_map_zoom_label.setText(f"z8-{v}"))
        cache_layout.addWidget(self.cache_map_zoom_label, 1, 2)
        self.cache_map_status = QLabel("--")
        self.cache_map_status.setStyleSheet("color: #888; font-size: 10px;")
        cache_layout.addWidget(self.cache_map_status, 1, 3)
        self.cache_map_btn = QPushButton("⬇️")
        self.cache_map_btn.setFixedWidth(30)
        self.cache_map_btn.setToolTip("Cache LA map tiles (zoom 8 to selected level)")
        self.cache_map_btn.clicked.connect(self._cache_map_tiles)
        cache_layout.addWidget(self.cache_map_btn, 1, 4)
        
        # Test tile button
        self.test_tile_btn = QPushButton("🔍")
        self.test_tile_btn.setFixedWidth(30)
        self.test_tile_btn.setToolTip("Test tile cache")
        self.test_tile_btn.clicked.connect(self._test_tile_cache)
        cache_layout.addWidget(self.test_tile_btn, 1, 5)
        
        # Digipeaters
        cache_layout.addWidget(QLabel("📡 Digis:"), 2, 0)
        self.cache_digi_status = QLabel("--")
        self.cache_digi_status.setStyleSheet("color: #888; font-size: 10px;")
        cache_layout.addWidget(self.cache_digi_status, 2, 3)
        self.cache_digi_btn = QPushButton("⬇️")
        self.cache_digi_btn.setFixedWidth(30)
        self.cache_digi_btn.setToolTip("Download known digipeaters in area")
        self.cache_digi_btn.clicked.connect(self._cache_digipeaters)
        cache_layout.addWidget(self.cache_digi_btn, 2, 4)
        
        # DARN - import from Excel
        cache_layout.addWidget(QLabel("🔴 DARN:"), 3, 0)
        self.cache_darn_status = QLabel("Built-in 41")
        self.cache_darn_status.setStyleSheet("color: #ff6b6b; font-size: 10px;")
        cache_layout.addWidget(self.cache_darn_status, 3, 2, 1, 2)
        
        mid_col.addWidget(cache_grp)
        
        mid_col.addStretch()
        main_layout.addLayout(mid_col)
        
        # RIGHT COLUMN - Station, VARA FM, Paths, Startup
        right_col = QVBoxLayout()
        right_col.setSpacing(6)
        
        # === STATION INFO (compact) ===
        station_grp = QGroupBox("📻 Station")
        station_grp.setStyleSheet(self._group_style())
        station_layout = QGridLayout(station_grp)
        station_layout.setSpacing(3)
        
        station_layout.addWidget(QLabel("Call:"), 0, 0)
        self.settings_callsign = QLineEdit()
        self.settings_callsign.setPlaceholderText("N0CALL")
        self.settings_callsign.setMaxLength(6)
        self.settings_callsign.setFixedWidth(70)
        self.settings_callsign.textChanged.connect(self._sync_callsign_to_beacon)
        station_layout.addWidget(self.settings_callsign, 0, 1)
        
        station_layout.addWidget(QLabel("-"), 0, 2)
        self.settings_ssid_combo = QComboBox()
        self.settings_ssid_combo.setToolTip("SSID: 0=Primary, 7=HT, 9=Mobile, 15=Digi")
        for i in range(16):
            self.settings_ssid_combo.addItem(str(i), i)
        self.settings_ssid_combo.setCurrentIndex(9)
        self.settings_ssid_combo.setFixedWidth(45)
        self.settings_ssid_combo.currentIndexChanged.connect(self._sync_ssid_to_beacon)
        station_layout.addWidget(self.settings_ssid_combo, 0, 3)
        
        self.ssid_type_label = QLabel("Mobile")
        self.ssid_type_label.setStyleSheet("color: #81c784; font-size: 9px;")
        station_layout.addWidget(self.ssid_type_label, 0, 4)
        
        station_layout.addWidget(QLabel("Comment:"), 1, 0)
        self.settings_comment = QLineEdit("PyTNC Pro")
        station_layout.addWidget(self.settings_comment, 1, 1, 1, 4)
        right_col.addWidget(station_grp)
        
        # === VARA FM (compact) - moved to right column under Station ===
        vara_grp = QGroupBox("📡 VARA FM")
        vara_grp.setStyleSheet(self._group_style())
        vara_layout = QGridLayout(vara_grp)
        vara_layout.setSpacing(3)
        
        vara_layout.addWidget(QLabel("Host:"), 0, 0)
        self.vara_host = QLineEdit("localhost")
        self.vara_host.setFixedWidth(120)
        vara_layout.addWidget(self.vara_host, 0, 1, 1, 3)
        
        vara_layout.addWidget(QLabel("Cmd:"), 1, 0)
        self.vara_cmd_port = QSpinBox()
        self.vara_cmd_port.setRange(1, 65535)
        self.vara_cmd_port.setValue(8300)
        self.vara_cmd_port.setFixedWidth(60)
        vara_layout.addWidget(self.vara_cmd_port, 1, 1)
        
        vara_layout.addWidget(QLabel("Data:"), 1, 2)
        self.vara_data_port = QSpinBox()
        self.vara_data_port.setRange(1, 65535)
        self.vara_data_port.setValue(8301)
        self.vara_data_port.setFixedWidth(60)
        vara_layout.addWidget(self.vara_data_port, 1, 3)
        
        vara_layout.addWidget(QLabel("KISS:"), 2, 0)
        self.vara_port = QSpinBox()
        self.vara_port.setRange(1, 65535)
        self.vara_port.setValue(8100)
        self.vara_port.setFixedWidth(60)
        vara_layout.addWidget(self.vara_port, 2, 1)
        
        self.settings_vara_status = QLabel("⚫")
        self.settings_vara_status.setStyleSheet("color: #607d8b;")
        vara_layout.addWidget(self.settings_vara_status, 2, 2)
        
        vara_btn_row = QHBoxLayout()
        self.vara_open_btn = QPushButton("Open")
        self.vara_open_btn.setMinimumWidth(50)
        self.vara_open_btn.setStyleSheet(self._button_style("#747d10", "#8a9412"))
        self.vara_open_btn.clicked.connect(self._open_vara_fm)
        vara_btn_row.addWidget(self.vara_open_btn)
        
        self.settings_vara_connect_btn = QPushButton("Connect")
        self.settings_vara_connect_btn.setMinimumWidth(60)
        self.settings_vara_connect_btn.clicked.connect(self._toggle_vara_connection)
        vara_btn_row.addWidget(self.settings_vara_connect_btn)
        vara_layout.addLayout(vara_btn_row, 3, 0, 1, 4)
        
        right_col.addWidget(vara_grp)
        
        # === STARTUP OPTIONS ===
        startup_grp = QGroupBox("🚀 Startup")
        startup_grp.setStyleSheet(self._group_style())
        startup_layout = QVBoxLayout(startup_grp)
        startup_layout.setSpacing(2)
        
        self.auto_connect_gps = QCheckBox("Auto-connect GPS")
        self.auto_connect_gps.setToolTip("Automatically connect to GPS on startup")
        startup_layout.addWidget(self.auto_connect_gps)
        
        self.auto_connect_aprs = QCheckBox("Auto-connect APRS-IS")
        self.auto_connect_aprs.setToolTip("Automatically connect to APRS-IS on startup")
        startup_layout.addWidget(self.auto_connect_aprs)
        
        right_col.addWidget(startup_grp)
        
        # === BUTTONS ===
        refresh_btn = QPushButton("↻ Refresh Ports")
        refresh_btn.clicked.connect(self._refresh_settings_ports)
        right_col.addWidget(refresh_btn)
        
        self.settings_save_btn = QPushButton("💾 Save Settings")
        self.settings_save_btn.setStyleSheet(self._button_style("#f57c00", "#ff9800"))
        self.settings_save_btn.clicked.connect(self._save_settings_from_tab)
        right_col.addWidget(self.settings_save_btn)
        
        right_col.addStretch()
        main_layout.addLayout(right_col)
        
        # Add log panel at bottom of Settings tab (spanning full width)
        settings_outer = QWidget()
        settings_outer_layout = QVBoxLayout(settings_outer)
        settings_outer_layout.setContentsMargins(0, 0, 0, 0)
        settings_outer_layout.setSpacing(5)
        
        # The main settings content
        settings_content = QWidget()
        settings_content.setLayout(main_layout)
        settings_outer_layout.addWidget(settings_content)
        
        # Log panel - fixed position and size
        # ============================================
        # ADJUST THESE VALUES FOR POSITION AND SIZE:
        LOG_X = 15       # X position from left
        LOG_Y = 500      # Y position from top  
        LOG_W = 550      # Width
        LOG_H = 250      # Height
        # ============================================
        
        settings_outer_layout.addWidget(self._branding_label())
        
        # Settings tab widget stored for adding in correct order
        self._settings_tab_widget = settings_outer
    
    # =========================================================================
    # IGate Tab
    # =========================================================================

    def _build_vara_tab(self):
        """Build VARA FM tab"""
        # =====================================================================
        # Tab 4: VARA FM Beacon (mirrors Beacon tab layout)
        # =====================================================================
        vara_tab = QWidget()
        vara_layout = QHBoxLayout(vara_tab)
        vara_layout.setContentsMargins(10, 10, 10, 10)
        vara_layout.setSpacing(10)
        
        # Left panel - Connection Status & Beacon Settings
        vara_left_panel = QWidget()
        vara_left_panel.setMaximumWidth(620)  # Constrain left panel
        vara_left_layout = QVBoxLayout(vara_left_panel)
        vara_left_layout.setSpacing(10)
        
        # Connection Status display (same as Beacon tab)
        vara_conn_grp = QGroupBox("🔌 Connection Status")
        vara_conn_grp.setStyleSheet(self._group_style())
        vara_conn_layout = QGridLayout(vara_conn_grp)
        vara_conn_layout.setSpacing(4)
        
        self.vara_ptt_status = QLabel("⚫ PTT: Not connected")
        self.vara_ptt_status.setStyleSheet("color: #607d8b;")
        vara_conn_layout.addWidget(self.vara_ptt_status, 0, 0)
        
        self.vara_gps_status = QLabel("⚫ GPS: Not connected")
        self.vara_gps_status.setStyleSheet("color: #607d8b;")
        vara_conn_layout.addWidget(self.vara_gps_status, 0, 1)
        
        self.vara_tx_status = QLabel("⚫ TX Audio: Not set")
        self.vara_tx_status.setStyleSheet("color: #607d8b;")
        vara_conn_layout.addWidget(self.vara_tx_status, 1, 0)
        
        self.vara_fm_status = QLabel("⚫ VARA: Not connected")
        self.vara_fm_status.setStyleSheet("color: #607d8b;")
        vara_conn_layout.addWidget(self.vara_fm_status, 1, 1)
        
        self.vara_aprs_status = QLabel("⚫ APRS-IS: Not connected")
        self.vara_aprs_status.setStyleSheet("color: #607d8b;")
        vara_conn_layout.addWidget(self.vara_aprs_status, 2, 0)
        
        vara_goto_settings_btn = QPushButton("⚙️ Settings")
        vara_goto_settings_btn.setFixedWidth(100)
        vara_goto_settings_btn.clicked.connect(lambda: self.tabs.setCurrentIndex(4))
        vara_conn_layout.addWidget(vara_goto_settings_btn, 3, 0)
        
        vara_refresh_btn = QPushButton("🔄 Refresh")
        vara_refresh_btn.setFixedWidth(100)
        vara_refresh_btn.clicked.connect(self._sync_vara_fm_connection_status)
        vara_conn_layout.addWidget(vara_refresh_btn, 3, 1)
        
        vara_left_layout.addWidget(vara_conn_grp)
        
        # VARA FM Control buttons
        vara_ctrl_grp = QGroupBox("📻 VARA FM Control")
        vara_ctrl_grp.setStyleSheet(self._group_style())
        vara_ctrl_layout = QHBoxLayout(vara_ctrl_grp)
        
        self.vara_open_btn = QPushButton("📂 Open VARA FM")
        self.vara_open_btn.setMinimumHeight(35)
        self.vara_open_btn.setToolTip("Launch VARA FM application")
        self.vara_open_btn.clicked.connect(self._open_vara_fm)
        self.vara_open_btn.setStyleSheet(self._button_style("#1565c0", "#1976d2"))
        vara_ctrl_layout.addWidget(self.vara_open_btn)
        
        self.vara_connect_btn = QPushButton("🔌 Connect")
        self.vara_connect_btn.setMinimumHeight(35)
        self.vara_connect_btn.setToolTip("Connect to VARA FM")
        self.vara_connect_btn.clicked.connect(self._toggle_vara_connection)
        self.vara_connect_btn.setStyleSheet(self._button_style("#2e7d32", "#388e3c"))
        vara_ctrl_layout.addWidget(self.vara_connect_btn)
        
        vara_left_layout.addWidget(vara_ctrl_grp)
        
        # Beacon Settings Group (mirrors Beacon tab)
        vara_beacon_grp = QGroupBox("📍 Beacon Settings")
        vara_beacon_grp.setStyleSheet(self._group_style())
        vara_beacon_layout = QGridLayout(vara_beacon_grp)
        vara_beacon_layout.setSpacing(8)
        
        # Row 0: Callsign and SSID
        vara_beacon_layout.addWidget(QLabel("Callsign:"), 0, 0)
        vara_call_layout = QHBoxLayout()
        self.vara_callsign_edit = QLineEdit("N0CALL")
        self.vara_callsign_edit.setMaxLength(6)
        self.vara_callsign_edit.setFixedWidth(80)
        self.vara_callsign_edit.setPlaceholderText("N0CALL")
        vara_call_layout.addWidget(self.vara_callsign_edit)
        vara_call_layout.addWidget(QLabel("-"))
        self.vara_ssid_combo = QComboBox()
        self.vara_ssid_combo.addItem("0  Primary/Main", 0)
        self.vara_ssid_combo.addItem("1  Secondary", 1)
        self.vara_ssid_combo.addItem("2  Secondary radio", 2)
        self.vara_ssid_combo.addItem("3  Additional", 3)
        self.vara_ssid_combo.addItem("4  Additional", 4)
        self.vara_ssid_combo.addItem("5  IGate/Gateway", 5)
        self.vara_ssid_combo.addItem("6  Satellite/Special", 6)
        self.vara_ssid_combo.addItem("7  Handheld (HT)", 7)
        self.vara_ssid_combo.addItem("8  Boat/Maritime", 8)
        self.vara_ssid_combo.addItem("9  Mobile/Vehicle", 9)
        self.vara_ssid_combo.addItem("10 Internet/APRS-IS", 10)
        self.vara_ssid_combo.addItem("11 Balloon/Aircraft", 11)
        self.vara_ssid_combo.addItem("12 Portable/Field", 12)
        self.vara_ssid_combo.addItem("13 Weather station", 13)
        self.vara_ssid_combo.addItem("14 Truck/Large veh", 14)
        self.vara_ssid_combo.addItem("15 Digipeater", 15)
        self.vara_ssid_combo.setCurrentIndex(9)
        vara_call_layout.addWidget(self.vara_ssid_combo)
        vara_call_layout.addStretch()
        vara_beacon_layout.addLayout(vara_call_layout, 0, 1)
        
        # Row 1: Location (read-only, synced from Settings/GPS)
        vara_beacon_layout.addWidget(QLabel("Location:"), 1, 0)
        vara_loc_layout = QHBoxLayout()
        self.vara_lat_edit = QDoubleSpinBox()
        self.vara_lat_edit.setRange(-90, 90)
        self.vara_lat_edit.setDecimals(6)
        self.vara_lat_edit.setValue(34.0522)
        self.vara_lat_edit.setPrefix("Lat ")
        self.vara_lat_edit.setFixedWidth(130)
        self.vara_lat_edit.setEnabled(False)  # Read-only - controlled by Settings
        vara_loc_layout.addWidget(self.vara_lat_edit)
        self.vara_lon_edit = QDoubleSpinBox()
        self.vara_lon_edit.setRange(-180, 180)
        self.vara_lon_edit.setDecimals(6)
        self.vara_lon_edit.setValue(-118.2437)
        self.vara_lon_edit.setPrefix("Lon ")
        self.vara_lon_edit.setFixedWidth(140)
        self.vara_lon_edit.setEnabled(False)  # Read-only - controlled by Settings
        vara_loc_layout.addWidget(self.vara_lon_edit)
        
        # GPS source indicator
        self.vara_gps_source = QLabel("📍 Manual")
        self.vara_gps_source.setStyleSheet("color: #ffb74d; font-weight: bold;")
        self.vara_gps_source.setToolTip("GPS source controlled in Settings tab")
        vara_loc_layout.addWidget(self.vara_gps_source)
        vara_loc_layout.addStretch()
        vara_beacon_layout.addLayout(vara_loc_layout, 1, 1)
        
        # Row 2: Symbol and Path (VARA uses digipeater, not WIDE path)
        vara_beacon_layout.addWidget(QLabel("Symbol:"), 2, 0)
        vara_sym_layout = QHBoxLayout()
        self.vara_symbol_preview = QLabel()
        self.vara_symbol_preview.setFixedSize(28, 28)
        self.vara_symbol_preview.setStyleSheet("background: #1a3a5c; border: 1px solid #42a5f5; border-radius: 4px;")
        self.vara_symbol_preview.setToolTip("Click a symbol in the picker")
        vara_sym_layout.addWidget(self.vara_symbol_preview)
        
        # Mode selector: Connected (to one station) vs KISS (broadcast to all)
        vara_sym_layout.addWidget(QLabel("Mode:"))
        self.vara_aprs_mode = QComboBox()
        self.vara_aprs_mode.addItems(["KISS (Broadcast)", "Connected (P2P)"])
        self.vara_aprs_mode.setToolTip("KISS = broadcast to all on frequency\nConnected = direct session to one station")
        self.vara_aprs_mode.setFixedWidth(140)
        self.vara_aprs_mode.currentIndexChanged.connect(self._vara_mode_changed)
        vara_sym_layout.addWidget(self.vara_aprs_mode)
        
        vara_sym_layout.addWidget(QLabel("Path:"))
        self.vara_digi_edit = QLineEdit("")
        self.vara_digi_edit.setPlaceholderText("WIDE1-1 or blank")
        self.vara_digi_edit.setToolTip("Digipeater path (e.g. WIDE1-1) or leave blank for direct")
        self.vara_digi_edit.setMinimumWidth(120)
        vara_sym_layout.addWidget(self.vara_digi_edit)
        vara_sym_layout.addStretch()
        vara_beacon_layout.addLayout(vara_sym_layout, 2, 1)
        
        # Row 3: Radio
        vara_beacon_layout.addWidget(QLabel("Radio:"), 3, 0)
        self.vara_radio_combo = QComboBox()
        self.vara_radio_combo.setEditable(True)
        self.vara_radio_combo.addItems([
            "", "Yaesu FT-991A", "Yaesu FT-991", "Yaesu FT-891", "Yaesu FT-710",
            "Yaesu FTM-500D", "Yaesu FTM-400XD", "Yaesu FT-5D", "Yaesu FT-3D",
            "Icom IC-705", "Icom IC-7100", "Icom IC-7300",
            "Kenwood TM-D710", "Kenwood TH-D74",
        ])
        vara_beacon_layout.addWidget(self.vara_radio_combo, 3, 1)
        
        # Row 4: Comment
        vara_beacon_layout.addWidget(QLabel("Comment:"), 4, 0)
        self.vara_comment_edit = QLineEdit("PyTNC Pro")
        vara_beacon_layout.addWidget(self.vara_comment_edit, 4, 1)
        
        vara_left_layout.addWidget(vara_beacon_grp)
        
        # Save and Send buttons
        vara_btn_layout = QHBoxLayout()
        
        self.vara_save_btn = QPushButton("💾 Save")
        self.vara_save_btn.setFixedWidth(80)
        self.vara_save_btn.setMinimumHeight(35)
        self.vara_save_btn.clicked.connect(self._vara_save_settings)
        self.vara_save_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ff9800, stop:1 #f57c00);
                color: white; font-weight: bold; border: 1px solid #ffb74d; border-radius: 4px;
            }
            QPushButton:hover { background: #ffb74d; }
        """)
        vara_btn_layout.addWidget(self.vara_save_btn)
        vara_btn_layout.addStretch()
        vara_left_layout.addLayout(vara_btn_layout)
        
        # Send Beacon button
        vara_send_grp = QGroupBox("📤 Send Beacon")
        vara_send_grp.setStyleSheet(self._group_style())
        vara_send_layout = QVBoxLayout(vara_send_grp)
        
        self.vara_beacon_btn = QPushButton("📡 Send APRS Beacon via VARA FM")
        self.vara_beacon_btn.setMinimumHeight(50)
        self.vara_beacon_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #8a9412, stop:1 #747d10);
                color: white; font-weight: bold; font-size: 14px;
                border: 2px solid #9ca31a; border-radius: 6px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #9ca31a, stop:1 #8a9412);
            }
        """)
        self.vara_beacon_btn.clicked.connect(self._vara_send_beacon)
        vara_send_layout.addWidget(self.vara_beacon_btn)
        
        vara_left_layout.addWidget(vara_send_grp)
        vara_left_layout.addStretch()
        
        vara_layout.addWidget(vara_left_panel, 1)
        
        # Right panel - Symbol Picker and TX Log
        vara_right_panel = QWidget()
        vara_right_panel.setMinimumWidth(480)  # Match Beacon width
        vara_right_layout = QVBoxLayout(vara_right_panel)
        vara_right_layout.setSpacing(5)
        
        # Table selector
        vara_table_layout = QHBoxLayout()
        vara_table_layout.addWidget(QLabel("Table:"))
        self.vara_symbol_table_picker = QComboBox()
        self.vara_symbol_table_picker.addItems(["/ Primary", "\\ Secondary"])
        self.vara_symbol_table_picker.currentIndexChanged.connect(self._vara_update_symbol_grid)
        vara_table_layout.addWidget(self.vara_symbol_table_picker)
        vara_table_layout.addStretch()
        vara_right_layout.addLayout(vara_table_layout)
        
        # Symbol Picker Group
        vara_symbol_grp = QGroupBox("🎨 Symbol Picker")
        vara_symbol_grp.setStyleSheet(self._group_style())
        vara_symbol_layout = QVBoxLayout(vara_symbol_grp)
        
        vara_scroll = QScrollArea()
        vara_scroll.setWidgetResizable(True)
        vara_scroll.setStyleSheet("""
            QScrollArea { border: 1px solid #1e3a5f; border-radius: 4px; background: #050d18; }
        """)
        
        self.vara_symbol_grid_widget = QWidget()
        self.vara_symbol_grid_widget.setStyleSheet("background: #050d18;")
        self.vara_symbol_grid_layout = QGridLayout(self.vara_symbol_grid_widget)
        self.vara_symbol_grid_layout.setSpacing(1)
        self.vara_symbol_grid_layout.setContentsMargins(2, 2, 2, 2)
        
        vara_scroll.setWidget(self.vara_symbol_grid_widget)
        vara_symbol_layout.addWidget(vara_scroll)
        
        vara_right_layout.addWidget(vara_symbol_grp, 1)
        
        # TX Log
        vara_log_grp = QGroupBox("📝 VARA FM Log")
        vara_log_grp.setStyleSheet(self._group_style())
        vara_log_layout = QVBoxLayout(vara_log_grp)
        
        self.vara_log = QTextEdit()
        self.vara_log.setReadOnly(True)
        self.vara_log.setStyleSheet("""
            QTextEdit {
                background: #000000; color: #00ff00;
                font-family: 'Consolas', 'Courier New', monospace; font-size: 12px;
                border: 2px solid #1e3a5f; border-radius: 4px; padding: 8px;
            }
        """)
        vara_log_layout.addWidget(self.vara_log)
        
        vara_clear_btn = QPushButton("🗑️ Clear Log")
        vara_clear_btn.setFixedWidth(100)
        vara_clear_btn.clicked.connect(lambda: self.vara_log.clear())
        vara_log_layout.addWidget(vara_clear_btn)
        
        vara_right_layout.addWidget(vara_log_grp, 1)
        vara_right_layout.addWidget(self._branding_label())
        
        vara_layout.addWidget(vara_right_panel, 1)
        
        # Initialize VARA state
        self.vara_chat_connected = False
        self.vara_remote_call = None
        self.vara_bytes_sent = 0
        self.vara_bytes_recv = 0
        
        # Build VARA symbol grid after UI is ready
        QTimer.singleShot(200, self._vara_build_symbol_grid)
        
        self.tabs.addTab(vara_tab, "📡 VARA FM")
        
        # Initialize VARA state
        self.vara_cmd_socket = None
        self.vara_data_socket = None
        self.vara_connected = False
    
    def _populate_serial_combo(self, combo):
        """Populate a combo with available serial ports"""
        combo.clear()
        combo.addItem("-- Select --", None)
        if HAS_SERIAL:
            for port in serial.tools.list_ports.comports():
                combo.addItem(f"{port.device} - {port.description}", port.device)
    
    def _populate_audio_inputs(self, combo):
        """Populate combo with audio input devices - show index and host API"""
        combo.clear()
        if not HAS_SOUNDDEVICE:
            combo.addItem("(sounddevice not installed)", -1)
            return
        try:
            devices = sd.query_devices()
            host_apis = sd.query_hostapis()
            for i, dev in enumerate(devices):
                if dev['max_input_channels'] > 0:
                    # Get host API name (MME, WASAPI, etc)
                    api_name = host_apis[dev['hostapi']]['name'] if dev['hostapi'] < len(host_apis) else "?"
                    # Shorten common API names
                    api_short = api_name.replace("Windows ", "").replace("DirectSound", "DS").replace("WASAPI", "WAS")
                    sr = int(dev['default_samplerate'])
                    ch = dev['max_input_channels']
                    combo.addItem(f"{i}: {dev['name']} [{api_short}] {sr}Hz {ch}ch", i)
        except Exception as e:
            combo.addItem(f"(error: {e})", -1)
    
    def _populate_audio_outputs(self, combo):
        """Populate combo with audio output devices - show index and host API"""
        combo.clear()
        if not HAS_SOUNDDEVICE:
            combo.addItem("(sounddevice not installed)", -1)
            return
        try:
            devices = sd.query_devices()
            host_apis = sd.query_hostapis()
            for i, dev in enumerate(devices):
                if dev['max_output_channels'] > 0:
                    # Get host API name (MME, WASAPI, etc)
                    api_name = host_apis[dev['hostapi']]['name'] if dev['hostapi'] < len(host_apis) else "?"
                    # Shorten common API names
                    api_short = api_name.replace("Windows ", "").replace("DirectSound", "DS").replace("WASAPI", "WAS")
                    sr = int(dev['default_samplerate'])
                    ch = dev['max_output_channels']
                    combo.addItem(f"{i}: {dev['name']} [{api_short}] {sr}Hz {ch}ch", i)
        except Exception as e:
            combo.addItem(f"(error: {e})", -1)
    
    def _on_path_changed(self, path_text):
        """Handle PATH combo change - informational only now"""
        # Note: VARA FM has its own path field, so we don't disable VARA buttons
        # based on the Beacon tab's path setting anymore
        pass
    
    def _refresh_settings_ports(self):
        """Refresh all serial port combos in settings"""
        self._populate_serial_combo(self.settings_ptt_combo)
        self._populate_serial_combo(self.settings_gps_combo)
        self._populate_audio_inputs(self.settings_rx_audio_combo)
        self._populate_audio_outputs(self.settings_tx_audio_combo)
    
    def _on_settings_rx_gain(self, value):
        """Update RX gain from settings slider"""
        gain_factor = value / 10.0  # 1-100 -> 0.1x to 10x
        self.settings_rx_gain_label.setText(f"{gain_factor:.1f}x")
        # Sync with hidden main gain slider
        self.gain.setValue(value)
    
    def _on_settings_tx_level(self, value):
        """Update TX level from settings slider"""
        self.settings_tx_level_label.setText(f"{value}%")
        # Sync with hidden TX level slider in Transmit tab
        if hasattr(self, 'tx_level_slider'):
            self.tx_level_slider.setValue(value)
        if hasattr(self, 'tx_level_label'):
            self.tx_level_label.setText(f"{value}%")
    
    def _on_tx_audio_changed(self):
        """Update Beacon tab TX audio status when combo changes"""
        self._sync_beacon_connection_status()
    
    def _toggle_gps(self):
        """Toggle GPS connection"""
        if hasattr(self, 'gps_serial') and self.gps_serial and self.gps_serial.is_open:
            self.gps_running = False
            time.sleep(0.1)
            self.gps_serial.close()
            self.gps_serial = None
            self.settings_gps_btn.setText("Connect")
            self.settings_gps_status.setText("⚫")
            self.settings_gps_status.setStyleSheet("color: #607d8b;")
            # Sync beacon GPS button
            if hasattr(self, 'beacon_gps_btn'):
                self.beacon_gps_btn.setChecked(False)
                self.beacon_gps_btn.setText("🛰️ GPS")
            self._gps_disconnected()
        else:
            port = self.settings_gps_combo.currentData()
            if port:
                # Get baud rate from combo
                baud = 4800  # Default for most GPS
                if hasattr(self, 'gps_baud_combo'):
                    baud = self.gps_baud_combo.currentData() or 4800
                try:
                    self.gps_serial = serial.Serial(port, baud, timeout=0.5)
                    self.gps_running = True
                    self.gps_thread = threading.Thread(target=self._gps_reader, daemon=True)
                    self.gps_thread.start()
                    self.settings_gps_btn.setText("Disconnect")
                    self.settings_gps_status.setText("🟢")
                    self.settings_gps_status.setStyleSheet("color: #69f0ae;")
                    # Sync beacon GPS button
                    if hasattr(self, 'beacon_gps_btn'):
                        self.beacon_gps_btn.setChecked(True)
                        self.beacon_gps_btn.setText("🛰️ Live")
                        self.beacon_gps_btn.setStyleSheet("""
                            QPushButton {
                                background: #2e7d32;
                                color: white;
                                border: none;
                                border-radius: 4px;
                                font-weight: bold;
                                padding: 4px 8px;
                            }
                            QPushButton:hover { background: #388e3c; }
                        """)
                    self._log(f"✓ GPS connected on {port} @ {baud} baud")
                except Exception as e:
                    self.settings_gps_status.setText("🔴")
                    self.settings_gps_status.setStyleSheet("color: #ef5350;")
                    self._log(f"❌ GPS error: {e}")
    
    def _gps_reader(self):
        """Background thread to read NMEA sentences from GPS"""
        buffer = ""
        last_fix_time = 0
        
        while self.gps_running:
            try:
                if not self.gps_serial or not self.gps_serial.is_open:
                    break
                
                # Read available data
                waiting = self.gps_serial.in_waiting
                if waiting > 0:
                    data = self.gps_serial.read(waiting)
                    
                    try:
                        buffer += data.decode('ascii', errors='ignore')
                    except (UnicodeDecodeError, AttributeError):
                        continue  # Bad data - skip this chunk
                    
                    # Process complete sentences
                    while '\n' in buffer:
                        line, buffer = buffer.split('\n', 1)
                        line = line.strip()
                        
                        if line.startswith('$'):
                            result = parse_nmea(line)
                            if result and result.get('valid') and result.get('lat') is not None:
                                lat = result['lat']
                                lon = result['lon']
                                
                                # Emit signal (thread-safe)
                                self.gps_position_signal.emit(lat, lon)
                                
                                # Throttle status updates
                                current_time = time.time()
                                if current_time - last_fix_time > 1.0:
                                    last_fix_time = current_time
                                    speed = result.get('speed_mph', 0)
                                    self.gps_status_signal.emit(True, speed)
                            elif result and not result.get('valid'):
                                # No fix yet
                                self.gps_status_signal.emit(False, 0.0)
                else:
                    time.sleep(0.1)
                    
            except serial.SerialException:
                break
            except Exception:
                time.sleep(0.1)
    
    def _update_gps_position(self, lat: float, lon: float):
        """Update position from GPS (called from main thread via signal)"""
        first_fix = not getattr(self, 'gps_has_fix', False)

        # Check if we've moved enough to warrant a beacon update
        if self.last_beacon_lat is not None and self.last_beacon_lon is not None:
            dist = haversine_meters(self.last_beacon_lat, self.last_beacon_lon, lat, lon)
            self.gps_moved_enough = dist >= self.MIN_BEACON_DISTANCE_M
        else:
            self.gps_moved_enough = True  # first beacon always goes

        # Store GPS coordinates in instance variables
        self.gps_lat = lat
        self.gps_lon = lon
        self.gps_has_fix = True
        
        self.lat_edit.setValue(lat)
        self.lon_edit.setValue(lon)
        
        # Center map only on the first GPS fix — don't hijack the map on every update
        if first_fix and hasattr(self, 'map'):
            self.map.page().runJavaScript(
                f"if(typeof setCenter === 'function') setCenter({lat}, {lon}, 13);"
            )
        
        # Also sync to VARA FM tab
        if hasattr(self, 'vara_lat_edit'):
            self.vara_lat_edit.setValue(lat)
            self.vara_lon_edit.setValue(lon)
        
        # Style beacon tab coordinates (green = GPS active)
        self.lat_edit.setStyleSheet("background: #1a5a3c; color: #69f0ae;")
        self.lon_edit.setStyleSheet("background: #1a5a3c; color: #69f0ae;")
        
        # Style VARA FM tab coordinates
        if hasattr(self, 'vara_lat_edit'):
            self.vara_lat_edit.setStyleSheet("background: #1a5a3c; color: #69f0ae;")
            self.vara_lon_edit.setStyleSheet("background: #1a5a3c; color: #69f0ae;")
        
        # Update Settings tab GPS lat/lon display
        if hasattr(self, 'settings_gps_latlon'):
            self.settings_gps_latlon.setText(f"{lat:.6f}, {lon:.6f}")
            self.settings_gps_latlon.setStyleSheet("color: #69f0ae; font-family: Consolas; font-size: 11px; font-weight: bold;")
        
        # Update source indicator on Beacon tab
        if hasattr(self, 'gps_source_label'):
            self.gps_source_label.setText("🛰️ GPS LIVE")
            self.gps_source_label.setStyleSheet("""
                color: #69f0ae; 
                font-weight: bold; 
                font-size: 14px;
                padding: 2px 8px;
                background: #1a3a2a;
                border: 1px solid #69f0ae;
                border-radius: 4px;
            """)
        
        # Update source indicator on VARA FM tab
        if hasattr(self, 'vara_gps_source'):
            self.vara_gps_source.setText("🛰️ GPS LIVE")
            self.vara_gps_source.setStyleSheet("color: #69f0ae; font-weight: bold;")
    
    def _update_gps_status(self, has_fix: bool, speed_mph: float = 0):
        """Update GPS status display"""
        if has_fix:
            if hasattr(self, 'tx_gps_status'):
                if speed_mph > 0:
                    self.tx_gps_status.setText(f"🟢 GPS: {speed_mph:.0f} mph")
                else:
                    self.tx_gps_status.setText("🟢 GPS: Fixed")
                self.tx_gps_status.setStyleSheet("color: #69f0ae;")
            
            # Update the source label with speed if moving
            if hasattr(self, 'gps_source_label'):
                if speed_mph > 1:
                    self.gps_source_label.setText(f"🛰️ GPS {speed_mph:.0f}mph")
                else:
                    self.gps_source_label.setText("🛰️ GPS LIVE")
                self.gps_source_label.setStyleSheet("""
                    color: #69f0ae; 
                    font-weight: bold; 
                    font-size: 14px;
                    padding: 2px 8px;
                    background: #1a3a2a;
                    border: 1px solid #69f0ae;
                    border-radius: 4px;
                """)
        else:
            if hasattr(self, 'tx_gps_status'):
                self.tx_gps_status.setText("🟡 GPS: Searching...")
                self.tx_gps_status.setStyleSheet("color: #ffb74d;")
            
            # Show "No Fix" in settings GPS display
            if hasattr(self, 'settings_gps_latlon'):
                self.settings_gps_latlon.setText("No Fix")
                self.settings_gps_latlon.setStyleSheet("color: #ffb74d; font-family: Consolas; font-size: 11px;")
            
            # Show searching state
            if hasattr(self, 'gps_source_label'):
                self.gps_source_label.setText("🛰️ Searching...")
                self.gps_source_label.setStyleSheet("""
                    color: #ffb74d; 
                    font-weight: bold; 
                    font-size: 14px;
                    padding: 2px 8px;
                    background: #2a2a1a;
                    border: 1px dashed #ffb74d;
                    border-radius: 4px;
                """)
    
    def _gps_disconnected(self):
        """Handle GPS disconnect (called from main thread)"""
        # Reset GPS fix status
        self.gps_has_fix = False
        
        if hasattr(self, 'tx_gps_status'):
            self.tx_gps_status.setText("⚫ GPS: Not connected")
            self.tx_gps_status.setStyleSheet("color: #607d8b;")
        
        # Reset Settings tab GPS display
        if hasattr(self, 'settings_gps_latlon'):
            self.settings_gps_latlon.setText("---, ---")
            self.settings_gps_latlon.setStyleSheet("color: #607d8b; font-family: Consolas; font-size: 11px;")
        
        # Reset Location label
        if hasattr(self, 'location_label'):
            self.location_label.setText("Location:")
            self.location_label.setStyleSheet("")
        
        # Reset source indicator
        if hasattr(self, 'gps_source_label'):
            self.gps_source_label.setText("📍 Manual")
            self.gps_source_label.setStyleSheet("""
                color: #ffb74d; 
                font-weight: bold; 
                font-size: 14px;
                padding: 2px 8px;
                background: #2a2a1a;
                border-radius: 4px;
            """)
        
        if hasattr(self, 'settings_gps_btn'):
            self.settings_gps_btn.setText("Connect")
        if hasattr(self, 'settings_gps_status'):
            self.settings_gps_status.setText("⚫")
            self.settings_gps_status.setStyleSheet("color: #607d8b;")
        
        # Re-enable manual location editing
        if hasattr(self, 'lat_edit'):
            self.lat_edit.setReadOnly(False)
            self.lat_edit.setStyleSheet("")
        if hasattr(self, 'lon_edit'):
            self.lon_edit.setReadOnly(False)
            self.lon_edit.setStyleSheet("")

    def _apply_manual_location(self):
        """Apply manual lat/long from Google Maps format"""
        text = self.manual_location.text().strip()
        if not text:
            return
        
        try:
            # Parse Google Maps format: "34.0522, -118.2437" or "34.0522,-118.2437"
            parts = text.replace(" ", "").split(",")
            if len(parts) == 2:
                lat = float(parts[0])
                lon = float(parts[1])
                
                # Validate ranges
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    # Update Beacon tab (even though disabled, setValue still works)
                    self.lat_edit.setValue(lat)
                    self.lon_edit.setValue(lon)
                    
                    # Also sync to VARA FM tab
                    if hasattr(self, 'vara_lat_edit'):
                        self.vara_lat_edit.setValue(lat)
                        self.vara_lon_edit.setValue(lon)
                    
                    self._log(f"📍 Location set: {lat:.6f}, {lon:.6f}")
                    self.manual_location.setStyleSheet("background: #1a5a3c;")  # Green tint
                    QTimer.singleShot(1000, lambda: self.manual_location.setStyleSheet(""))

                    # Center map on manually set position
                    if hasattr(self, 'map'):
                        self.map.page().runJavaScript(
                            f"if(typeof setCenter === 'function') setCenter({lat}, {lon}, 13);"
                        )
                    
                    # Update source indicator on Beacon tab
                    if hasattr(self, 'gps_source_label'):
                        self.gps_source_label.setText("📍 Manual")
                        self.gps_source_label.setStyleSheet("""
                            color: #ffb74d; 
                            font-weight: bold; 
                            font-size: 14px;
                            padding: 2px 8px;
                            background: #2a2a1a;
                            border-radius: 4px;
                        """)
                    
                    # Update source indicator on VARA FM tab
                    if hasattr(self, 'vara_gps_source'):
                        self.vara_gps_source.setText("📍 Manual")
                        self.vara_gps_source.setStyleSheet("color: #ffb74d; font-weight: bold;")
                else:
                    self._log("❌ Invalid coordinates range")
            else:
                self._log("❌ Use format: lat, lon (e.g., 34.0522, -118.2437)")
        except ValueError:
            self._log("❌ Invalid number format")
    
    def _vara_log(self, msg):
        """Log message to VARA FM log display AND Beacon tab TX Log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        # Log to VARA FM log display
        if hasattr(self, 'vara_log'):
            self.vara_log.append(f"[{timestamp}] {msg}")
        # Also log to Beacon tab TX Log (preset_log)
        if hasattr(self, 'preset_log'):
            self.preset_log.append(f"[{timestamp}] VARA: {msg}")
        from PyQt6.QtWidgets import QApplication
        QApplication.processEvents()
    
    def _vara_update_tx_indicator(self, is_tx):
        """Update TX/RX indicator"""
        if is_tx:
            self.vara_tx_indicator.setText("● TX")
            self.vara_tx_indicator.setStyleSheet("color: #ef5350; font-weight: bold;")
        else:
            self.vara_tx_indicator.setText("● RX")
            self.vara_tx_indicator.setStyleSheet("color: #69f0ae; font-weight: bold;")
    
    def _vara_send_beacon(self):
        """Send APRS beacon via VARA FM - supports KISS broadcast or Connected mode"""
        # Check which mode we're in
        is_kiss_mode = hasattr(self, 'vara_aprs_mode') and self.vara_aprs_mode.currentIndex() == 0
        
        # Get path/digi from VARA FM tab
        path = self.vara_digi_edit.text().strip().upper()
        
        # Get callsign from Beacon tab
        callsign = self.callsign_edit.text().strip().upper()
        ssid = self.ssid_combo.currentData()
        comment = self.vara_comment_edit.text() if hasattr(self, 'vara_comment_edit') else self.comment_edit.text()
        sym_table = getattr(self, '_vara_symbol_table', '/')
        sym_code = getattr(self, '_vara_symbol_code', '>')
        
        if not callsign or callsign == "N0CALL":
            self._vara_log("❌ Set your callsign in APRS tab first!")
            return
        
        # Get coordinates: GPS > Manual > Fallback
        lat, lon = None, None
        if hasattr(self, 'gps_has_fix') and self.gps_has_fix and hasattr(self, 'gps_lat'):
            lat, lon = self.gps_lat, self.gps_lon
            self._vara_log("🛰️ Using GPS coordinates")
        else:
            manual_text = self.manual_location.text().strip() if hasattr(self, 'manual_location') else ""
            if manual_text:
                try:
                    parts = manual_text.replace(" ", "").split(",")
                    if len(parts) == 2:
                        lat, lon = float(parts[0]), float(parts[1])
                except ValueError:
                    pass
            if lat is None:
                lat, lon = self.lat_edit.value(), self.lon_edit.value()
        
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        
        # Format APRS position packet
        lat_deg, lat_min = int(abs(lat)), (abs(lat) - int(abs(lat))) * 60
        lon_deg, lon_min = int(abs(lon)), (abs(lon) - int(abs(lon))) * 60
        lat_dir, lon_dir = 'N' if lat >= 0 else 'S', 'E' if lon >= 0 else 'W'
        
        position = f"!{lat_deg:02d}{lat_min:05.2f}{lat_dir}{sym_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{sym_code}"
        
        if is_kiss_mode:
            # KISS broadcast mode - send to all on frequency
            self._vara_send_beacon_kiss(full_call, path, position, comment, lat, lon, sym_table, sym_code)
        else:
            # Connected mode - direct session to digipeater
            if not path:
                self._vara_log("❌ Enter digipeater callsign for Connected mode!")
                return
            self._vara_send_beacon_connected(full_call, path, position, comment, lat, lon, sym_table, sym_code)
    
    def _vara_send_beacon_kiss(self, full_call, path, position, comment, lat, lon, sym_table, sym_code):
        """Send APRS beacon via VARA FM KISS port (broadcast to all)"""
        # Build APRS packet with path
        if path:
            aprs_info = f"{full_call}>APPR01,{path}:{position}{comment}"
        else:
            aprs_info = f"{full_call}>APPR01:{position}{comment}"
        
        self._vara_log(f"📡 KISS Broadcast: {full_call}")
        self._vara_log(f"   {aprs_info}")
        
        # Check KISS connection
        if not self.vara_kiss_connected or not self.vara_kiss_socket:
            self._vara_log("❌ KISS not connected. Click 'Connect' first!")
            self._vara_log("   (Make sure VARA FM is running)")
            return
        
        try:
            # Build AX.25 UI frame and wrap in KISS
            kiss_frame = self._build_kiss_frame(full_call, path, position + comment)
            
            self.vara_kiss_socket.send(kiss_frame)
            self._vara_log(f"📤 Sent {len(kiss_frame)} bytes via KISS")
            
            # Add to map
            if self.map_ready:
                import json
                icon_file, overlay = icon_path(sym_table, sym_code)
                try:
                    rel_path = icon_file.relative_to(BASE_DIR)
                    icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                except ValueError:
                    icon_url = f"http://127.0.0.1:{self.http_port}/aprs_symbols_48/primary/29.png"
                tooltip = f"<b>{full_call}</b><br>{comment}<br><i>TX via VARA FM KISS</i>"
                js = f"queueStation({json.dumps(full_call)},{lat},{lon},'{icon_url}',{json.dumps(tooltip)},false,\"\")"
                self.map.page().runJavaScript(js)
                self._vara_log(f"📍 Added to map")
            
            self._vara_log("✅ Beacon sent!")
            
        except Exception as e:
            self._vara_log(f"❌ KISS send error: {e}")
    
    def _vara_send_beacon_connected(self, full_call, digi, position, comment, lat, lon, sym_table, sym_code):
        """Send APRS beacon via VARA FM connected session (original mode)"""
        aprs_packet = f"{full_call}>APPR01,{digi}:{position}{comment}\r"
        
        self._vara_log(f"📍 Connected: {full_call} → {digi}")
        self._vara_log(f"   {aprs_packet.strip()}")
        
        if not (hasattr(self, 'vara_cmd_socket') and self.vara_cmd_socket):
            self._vara_log("❌ Not connected to VARA FM. Click 'Connect' first!")
            return
        
        try:
            # Reset flags
            self._vara_is_connected_to_remote = False
            self._vara_ptt_active = True
            
            # Setup and connect
            self.vara_cmd_socket.send(f"MYCALL {full_call}\r".encode())
            time.sleep(0.3)
            self.vara_cmd_socket.send(b"LISTEN ON\r")
            time.sleep(0.3)
            
            self._vara_log(f"📡 Calling {digi}...")
            self.vara_cmd_socket.send(f"CONNECT {full_call} {digi}\r".encode())
            
            # Wait for connection
            start = time.time()
            while time.time() - start < 30.0:
                if self._vara_is_connected_to_remote:
                    break
                QApplication.processEvents()
                time.sleep(0.1)
            
            if not self._vara_is_connected_to_remote:
                self._vara_log("❌ Connection timeout")
                return
            
            self._vara_log("🔗 Connected!")
            
            # Wait for handshake to complete (PTT OFF)
            ptt_wait_start = time.time()
            while time.time() - ptt_wait_start < 15.0:
                if not self._vara_is_connected_to_remote:
                    self._vara_log("❌ Link dropped")
                    return
                if not self._vara_ptt_active:
                    break
                QApplication.processEvents()
                time.sleep(0.1)
            
            if not self._vara_is_connected_to_remote:
                self._vara_log("❌ Link dropped")
                return
            
            # Send data
            if hasattr(self, 'vara_data_socket') and self.vara_data_socket:
                bytes_sent = self.vara_data_socket.send(aprs_packet.encode())
                self._vara_log(f"📤 Sent {bytes_sent} bytes")
            else:
                self._vara_log("❌ Data socket not available!")
                return
            
            # Wait for TX
            tx_start, saw_ptt = time.time(), False
            while time.time() - tx_start < 10.0:
                QApplication.processEvents()
                if self._vara_ptt_active:
                    saw_ptt = True
                elif saw_ptt:
                    self._vara_log("✅ Transmitted!")
                    break
                time.sleep(0.1)
            
            # Add to map
            if self.map_ready:
                import json
                icon_file, overlay = icon_path(sym_table, sym_code)
                try:
                    rel_path = icon_file.relative_to(BASE_DIR)
                    icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                except ValueError:
                    icon_url = f"http://127.0.0.1:{self.http_port}/aprs_symbols_48/primary/29.png"
                tooltip = f"<b>{full_call}</b><br>{comment}<br><i>TX via VARA FM</i>"
                js = f"queueStation({json.dumps(full_call)},{lat},{lon},'{icon_url}',{json.dumps(tooltip)},false,\"\")"
                self.map.page().runJavaScript(js)
                self._vara_log(f"📍 Added to map")
            
            # Disconnect
            self.vara_cmd_socket.send(b"DISCONNECT\r")
            self._vara_log("📡 Disconnected")
            
        except Exception as e:
            self._vara_log(f"❌ Error: {e}")
        finally:
            self._vara_is_connected_to_remote = False
    
    def _vara_data_rx_loop(self):
        """Background thread to receive data from VARA FM
        
        APRS packets end with CR (\r) or CRLF (\r\n).
        TCP may deliver multiple packets in one recv() or split packets across recv()s.
        We accumulate data in a buffer and split on line endings.
        """
        from PyQt6.QtCore import QMetaObject, Qt, Q_ARG
        
        rx_buffer = ""
        
        while self.vara_rx_running and self.vara_data_socket:
            try:
                self.vara_data_socket.settimeout(0.5)
                data = self.vara_data_socket.recv(4096)
                if data:
                    text = data.decode(errors='replace')
                    if not text:
                        continue
                    
                    rx_buffer += text
                    
                    # Log raw chunk for debugging
                    QMetaObject.invokeMethod(self, "_vara_log_rx",
                                            Qt.ConnectionType.QueuedConnection,
                                            Q_ARG(str, f"[{len(text)}b] {text.strip()[:60]}"))
                    
                    # Process complete lines (APRS packets end with CR or CRLF)
                    # Split on \r or \n, handling both \r\n and bare \r
                    while '\r' in rx_buffer or '\n' in rx_buffer:
                        # Find first line ending
                        cr_pos = rx_buffer.find('\r')
                        lf_pos = rx_buffer.find('\n')
                        
                        if cr_pos >= 0 and (lf_pos < 0 or cr_pos < lf_pos):
                            # CR comes first
                            line = rx_buffer[:cr_pos]
                            # Skip CRLF if present
                            if cr_pos + 1 < len(rx_buffer) and rx_buffer[cr_pos + 1] == '\n':
                                rx_buffer = rx_buffer[cr_pos + 2:]
                            else:
                                rx_buffer = rx_buffer[cr_pos + 1:]
                        else:
                            # LF comes first (bare LF)
                            line = rx_buffer[:lf_pos]
                            rx_buffer = rx_buffer[lf_pos + 1:]
                        
                        line = line.strip()
                        if not line:
                            continue
                        
                        # Check if this looks like an APRS packet (CALL>DEST:info)
                        if '>' in line and ':' in line:
                            gt_pos = line.find('>')
                            colon_pos = line.find(':')
                            if colon_pos > gt_pos:
                                try:
                                    header, info = line.split(':', 1)
                                    src_call = header.split('>')[0].strip()
                                    dest_call = header.split('>')[1].split(',')[0].strip()
                                    
                                    QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                            Qt.ConnectionType.QueuedConnection,
                                                            Q_ARG(str, f"📦 {line}"))
                                    
                                    QMetaObject.invokeMethod(self, "_process_vara_aprs",
                                                            Qt.ConnectionType.QueuedConnection,
                                                            Q_ARG(str, src_call),
                                                            Q_ARG(str, dest_call),
                                                            Q_ARG(str, info))
                                except (ValueError, IndexError) as e:
                                    QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                            Qt.ConnectionType.QueuedConnection,
                                                            Q_ARG(str, f"⚠️ Parse error: {e} in '{line[:40]}'"))
                        else:
                            # Not an APRS packet - log it anyway (might be status info)
                            QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                    Qt.ConnectionType.QueuedConnection,
                                                    Q_ARG(str, f"💬 {line}"))
                    
                    # Prevent buffer from growing indefinitely if no line endings
                    if len(rx_buffer) > 4096:
                        QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                Qt.ConnectionType.QueuedConnection,
                                                Q_ARG(str, f"⚠️ Buffer overflow, clearing"))
                        rx_buffer = ""
                            
            except socket.timeout:
                continue
            except Exception as e:
                if self.vara_rx_running:
                    QMetaObject.invokeMethod(self, "_vara_log_rx",
                                            Qt.ConnectionType.QueuedConnection,
                                            Q_ARG(str, f"❌ RX Error: {e}"))
                break
    
    def _vara_cmd_rx_loop(self):
        """Background thread to listen for VARA FM command events"""
        from PyQt6.QtCore import QMetaObject, Qt, Q_ARG
        
        while self.vara_rx_running and self.vara_cmd_socket:
            try:
                self.vara_cmd_socket.settimeout(0.5)
                data = self.vara_cmd_socket.recv(4096)
                if data:
                    text = data.decode(errors='replace').strip()
                    if not text:
                        continue
                    
                    for line in text.split('\r'):
                        line = line.strip()
                        if not line or line in ("BUSY ON", "BUSY OFF", "IAMALIVE"):
                            continue
                        
                        # Log non-noisy events
                        QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                Qt.ConnectionType.QueuedConnection,
                                                Q_ARG(str, f"CMD: {line}"))
                        
                        # Handle connection states
                        if line.startswith("CONNECTED "):
                            parts = line.split()
                            if len(parts) >= 3:
                                self._vara_is_connected_to_remote = True
                                QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                        Qt.ConnectionType.QueuedConnection,
                                                        Q_ARG(str, f"🔗 Connected to {parts[2]}"))
                        
                        elif line.startswith("REGISTERED ") or line == "LINK REGISTERED":
                            self._vara_is_connected_to_remote = True
                        
                        elif line == "DISCONNECTED":
                            self._vara_is_connected_to_remote = False
                            QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                    Qt.ConnectionType.QueuedConnection,
                                                    Q_ARG(str, "📴 Disconnected"))
                        
                        elif line.startswith("RING"):
                            caller = line.replace("RING", "").strip()
                            QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                    Qt.ConnectionType.QueuedConnection,
                                                    Q_ARG(str, f"📞 Incoming call: {caller}"))
                        
                        # Track PTT
                        elif line == "PTT ON":
                            self._vara_ptt_active = True
                        elif line == "PTT OFF":
                            self._vara_ptt_active = False
                            
            except socket.timeout:
                continue
            except Exception as e:
                if self.vara_rx_running:
                    QMetaObject.invokeMethod(self, "_vara_log_rx",
                                            Qt.ConnectionType.QueuedConnection,
                                            Q_ARG(str, f"❌ Error: {e}"))
                break
    
    from PyQt6.QtCore import pyqtSlot
    
    @pyqtSlot(str)
    def _vara_log_rx(self, msg):
        """Log received data (called from RX thread)"""
        self._vara_log(f"📥 {msg}")
        # Also log to main TX Log so it's visible
        self._log(f"📥 VARA RX: {msg}")
    
    @pyqtSlot(str)
    @pyqtSlot(str, str, str)
    def _process_vara_aprs(self, src_call: str, dest_call: str, info: str):
        """Parse APRS packet from VARA FM and add to map"""
        try:
            parsed = aprs_classify(dest_call, info)
            fields = parsed.get("fields", {})
            
            lat = fields.get("lat")
            lon = fields.get("lon")
            
            if lat is not None and lon is not None:
                sym_table = fields.get("table", "/")
                sym_code = fields.get("sym", ">")
                
                # Get icon
                ic, ov = icon_path(sym_table, sym_code)
                if ov:
                    ic = make_overlay(ic, ov)
                
                try:
                    rel_path = ic.relative_to(BASE_DIR)
                    icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                except ValueError:
                    try:
                        rel_path = ic.relative_to(BUNDLE_DIR)
                        icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                    except ValueError:
                        icon_url = f"http://127.0.0.1:{self.http_port}/aprs_symbols_48/primary/29.png"
                
                # Build tooltip
                tooltip_parts = [f"<b>{src_call}</b>", "📻 VARA FM"]
                if fields.get("comment"):
                    tooltip_parts.append(f"💬 {clean_aprs_comment(fields['comment'], 60)}")
                if fields.get("speed_mph"):
                    tooltip_parts.append(f"🚗 {fields['speed_mph']:.0f} mph")
                if fields.get("altitude_ft"):
                    tooltip_parts.append(f"📈 {fields['altitude_ft']} ft")
                tooltip_parts.append(f"🕐 {datetime.now().strftime('%H:%M:%S')}")
                tooltip = "<br>".join(tooltip_parts)
                
                # Add to map
                import json
                js = f"queueStation({json.dumps(src_call)},{lat},{lon},'{icon_url}',{json.dumps(tooltip)},false,\"\")"
                
                if self.map_ready:
                    self.map.page().runJavaScript(js)
                    self._log(f"📍 VARA RX: {src_call} at {lat:.4f}, {lon:.4f}")
        except Exception as e:
            self._log(f"⚠️ VARA parse error: {e}")
    
    def _vara_mode_changed(self, index):
        """Handle APRS mode change between KISS and Connected"""
        if index == 0:  # KISS broadcast
            self.vara_digi_edit.setPlaceholderText("WIDE1-1 or blank")
            self.vara_digi_edit.setToolTip("Digipeater path (e.g. WIDE1-1) or leave blank for direct")
            self._vara_log("📻 Mode: KISS (broadcast to all on frequency)")
        else:  # Connected P2P
            self.vara_digi_edit.setPlaceholderText("Digipeater callsign")
            self.vara_digi_edit.setToolTip("Direct station callsign (e.g. K6OLI-10)")
            self._vara_log("📻 Mode: Connected (direct session to one station)")
    
    def _build_kiss_frame(self, src_call, path, info):
        """Build a KISS-wrapped AX.25 UI frame for APRS
        
        KISS frame format:
        - FEND (0xC0) - frame start
        - Command (0x00) - data frame on port 0
        - AX.25 frame
        - FEND (0xC0) - frame end
        
        AX.25 UI frame format:
        - Destination address (7 bytes)
        - Source address (7 bytes)
        - Digipeater addresses (7 bytes each, optional)
        - Control (0x03 for UI)
        - PID (0xF0 for no layer 3)
        - Information field
        """
        FEND = 0xC0
        FESC = 0xDB
        TFEND = 0xDC
        TFESC = 0xDD
        
        def encode_callsign(call, last=False):
            """Encode callsign to AX.25 format (7 bytes)"""
            # Split SSID
            if '-' in call:
                cs, ssid = call.split('-', 1)
                ssid = int(ssid)
            else:
                cs, ssid = call, 0
            
            # Pad to 6 chars
            cs = cs.upper().ljust(6)[:6]
            
            # Shift left by 1 (AX.25 encoding)
            encoded = bytearray()
            for c in cs:
                encoded.append(ord(c) << 1)
            
            # SSID byte: 0b011SSSS0 for intermediate, 0b011SSSS1 for last
            ssid_byte = 0b01100000 | ((ssid & 0x0F) << 1)
            if last:
                ssid_byte |= 0x01  # Set last address bit
            encoded.append(ssid_byte)
            
            return bytes(encoded)
        
        def kiss_escape(data):
            """Escape special KISS bytes"""
            result = bytearray()
            for b in data:
                if b == FEND:
                    result.append(FESC)
                    result.append(TFEND)
                elif b == FESC:
                    result.append(FESC)
                    result.append(TFESC)
                else:
                    result.append(b)
            return bytes(result)
        
        # Build AX.25 frame
        ax25 = bytearray()
        
        # Destination: APPR01 (PyTNC Pro v0.1.0-beta - official TOCALL)
        ax25.extend(encode_callsign("APPR01", last=False))
        
        # Source
        if path:
            ax25.extend(encode_callsign(src_call, last=False))
            # Digipeater path
            digis = [d.strip() for d in path.split(',') if d.strip()]
            for i, digi in enumerate(digis):
                is_last = (i == len(digis) - 1)
                ax25.extend(encode_callsign(digi, last=is_last))
        else:
            ax25.extend(encode_callsign(src_call, last=True))
        
        # Control: UI frame (0x03)
        ax25.append(0x03)
        
        # PID: No layer 3 (0xF0)
        ax25.append(0xF0)
        
        # Information field (APRS data)
        ax25.extend(info.encode('ascii', errors='replace'))
        
        # Build KISS frame
        kiss = bytearray()
        kiss.append(FEND)
        kiss.append(0x00)  # Data frame, port 0
        kiss.extend(kiss_escape(bytes(ax25)))
        kiss.append(FEND)
        
        return bytes(kiss)
    
    def _parse_kiss_frame(self, data):
        """Parse a KISS frame and extract AX.25 data
        
        Returns tuple: (src_call, dest_call, digis, info) or None if invalid
        """
        FEND = 0xC0
        FESC = 0xDB
        TFEND = 0xDC
        TFESC = 0xDD
        
        def kiss_unescape(data):
            """Unescape KISS special bytes"""
            result = bytearray()
            i = 0
            while i < len(data):
                if data[i] == FESC and i + 1 < len(data):
                    if data[i + 1] == TFEND:
                        result.append(FEND)
                    elif data[i + 1] == TFESC:
                        result.append(FESC)
                    else:
                        result.append(data[i + 1])
                    i += 2
                else:
                    result.append(data[i])
                    i += 1
            return bytes(result)
        
        def decode_callsign(data):
            """Decode AX.25 callsign (7 bytes)"""
            if len(data) < 7:
                return None, 0, False
            
            cs = ''.join(chr(b >> 1) for b in data[:6]).strip()
            ssid = (data[6] >> 1) & 0x0F
            last = bool(data[6] & 0x01)
            
            if ssid > 0:
                return f"{cs}-{ssid}", ssid, last
            return cs, ssid, last
        
        try:
            # Strip FEND delimiters
            if data and data[0] == FEND:
                data = data[1:]
            if data and data[-1] == FEND:
                data = data[:-1]
            
            if len(data) < 2:
                return None
            
            # First byte is command (should be 0x00 for data)
            cmd = data[0]
            if cmd != 0x00:
                return None  # Not a data frame
            
            # Unescape the rest
            ax25 = kiss_unescape(data[1:])
            
            if len(ax25) < 16:  # Minimum: 7+7+1+1 = 16 bytes
                return None
            
            # Parse addresses
            dest_call, _, _ = decode_callsign(ax25[0:7])
            src_call, _, last = decode_callsign(ax25[7:14])
            
            digis = []
            offset = 14
            while not last and offset + 7 <= len(ax25):
                digi, _, last = decode_callsign(ax25[offset:offset+7])
                if digi:
                    digis.append(digi)
                offset += 7
            
            # Control and PID
            if offset + 2 > len(ax25):
                return None
            control = ax25[offset]
            pid = ax25[offset + 1]
            
            # Info field
            info = ax25[offset + 2:].decode('ascii', errors='replace')
            
            return (src_call, dest_call, digis, info)
            
        except Exception as e:
            return None
    
    def _vara_kiss_rx_loop(self):
        """Background thread to receive KISS frames from VARA FM
        
        Receives AX.25 frames wrapped in KISS, parses APRS packets,
        and adds stations to the map.
        """
        from PyQt6.QtCore import QMetaObject, Qt, Q_ARG
        
        FEND = 0xC0
        buffer = bytearray()
        
        while self.vara_rx_running and self.vara_kiss_socket:
            try:
                self.vara_kiss_socket.settimeout(0.5)
                data = self.vara_kiss_socket.recv(4096)
                
                if not data:
                    continue
                
                buffer.extend(data)
                
                # Extract complete KISS frames (between FEND markers)
                while FEND in buffer:
                    # Find start of frame
                    start = buffer.find(FEND)
                    if start == -1:
                        break
                    
                    # Find end of frame
                    end = buffer.find(FEND, start + 1)
                    if end == -1:
                        break
                    
                    # Extract frame (including delimiters)
                    frame = bytes(buffer[start:end + 1])
                    buffer = buffer[end + 1:]
                    
                    # Skip empty frames (just FEND FEND)
                    if len(frame) <= 2:
                        continue
                    
                    # Parse the frame
                    result = self._parse_kiss_frame(frame)
                    if result:
                        src_call, dest_call, digis, info = result
                        
                        # Log it
                        path_str = ",".join(digis) if digis else "DIRECT"
                        QMetaObject.invokeMethod(self, "_vara_log_rx",
                                                Qt.ConnectionType.QueuedConnection,
                                                Q_ARG(str, f"KISS: {src_call}>{dest_call},{path_str}:{info[:50]}"))
                        
                        # Process as APRS
                        QMetaObject.invokeMethod(self, "_process_vara_aprs",
                                                Qt.ConnectionType.QueuedConnection,
                                                Q_ARG(str, src_call),
                                                Q_ARG(str, dest_call),
                                                Q_ARG(str, info))
                        
            except socket.timeout:
                continue
            except Exception as e:
                if self.vara_rx_running:
                    QMetaObject.invokeMethod(self, "_vara_log_rx",
                                            Qt.ConnectionType.QueuedConnection,
                                            Q_ARG(str, f"❌ KISS RX error: {e}"))
                break

    def _send_beacon_aprs_is(self):
        """Send beacon via APRS-IS"""
        if not self.aprs_is_running or not self.aprs_is_socket:
            if hasattr(self, 'preset_log'):
                self.preset_log.append("⚠️ Not connected to APRS-IS — go to the MAP tab and click START IS")
            return

        self._log("🌐 Attempting APRS-IS beacon...")
        
        # Get beacon data
        callsign = self.callsign_edit.text().strip().upper()
        ssid = self.ssid_combo.currentData()
        symbol_code = self.symbol_code_edit.text() or ">"
        comment = self.comment_edit.text().strip()

        # Overlay char replaces table byte if set
        overlay = ""
        if hasattr(self, 'symbol_overlay_edit'):
            overlay = self.symbol_overlay_edit.text().strip().upper()
        if overlay and (overlay.isalpha() or overlay.isdigit()):
            symbol_table = overlay  # e.g. "I" for IGate, "D" for Digi
        else:
            symbol_table = self.symbol_table_combo.currentText()
        
        # GET COORDINATES: GPS if has fix, otherwise manual from Settings
        lat = None
        lon = None
        
        if hasattr(self, 'gps_has_fix') and self.gps_has_fix and hasattr(self, 'gps_lat'):
            lat = self.gps_lat
            lon = self.gps_lon
            self._log(f"🛰️ Using GPS: {lat:.6f}, {lon:.6f}")
        else:
            manual_text = self.manual_location.text().strip() if hasattr(self, 'manual_location') else ""
            if manual_text:
                try:
                    parts = manual_text.replace(" ", "").split(",")
                    if len(parts) == 2:
                        lat = float(parts[0])
                        lon = float(parts[1])
                        self._log(f"📍 Using manual: {lat:.6f}, {lon:.6f}")
                except ValueError:
                    pass
            if lat is None:
                lat = self.lat_edit.value()
                lon = self.lon_edit.value()
                self._log(f"⚠️ Using fallback: {lat:.6f}, {lon:.6f}")
        
        if not callsign or callsign == "N0CALL":
            self._log("❌ Set your callsign first!")
            return
        
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        
        # Build APRS position
        lat_deg = int(abs(lat))
        lat_min = (abs(lat) - lat_deg) * 60
        lat_dir = "N" if lat >= 0 else "S"
        lon_deg = int(abs(lon))
        lon_min = (abs(lon) - lon_deg) * 60
        lon_dir = "E" if lon >= 0 else "W"
        
        pos = f"!{lat_deg:02d}{lat_min:05.2f}{lat_dir}{symbol_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{symbol_code}"
        if comment:
            # APRS-IS can handle longer comments than RF
            pos += comment[:80]
        
        packet = f"{full_call}>APPR01,TCPIP*:{pos}\r\n"
        self._log(f"📤 Sending: {packet.strip()}")
        
        try:
            self.aprs_is_socket.send(packet.encode())
            self._log(f"✅ Beacon sent via APRS-IS!")
            # Record position so distance filter works correctly
            if self.gps_has_fix and self.gps_lat is not None:
                self.last_beacon_lat = self.gps_lat
                self.last_beacon_lon = self.gps_lon
            
            # Log to APRS tab TX Log (cyan color for APRS-IS)
            ts = datetime.now().strftime("%H:%M:%S")
            self.preset_log.append(f"<br><span style='color:#888'>[{ts}]</span> <span style='color:#00d4ff'>🌐 Transmitting APRS-IS beacon...</span>")
            self.preset_log.append(f"   From: <span style='color:#ffd54f'>{full_call}</span>")
            self.preset_log.append(f"   To: APPR01-0 via TCPIP*")
            self.preset_log.append(f"   Position: <span style='color:#80deea'>{pos}</span>")
            
            # Log to APRS live feed (MAP tab)
            self._log(f"🌐 TX Beacon: {full_call} via APRS-IS")
            
            # Plot our own position on the map
            try:
                ic, ov = icon_path(symbol_table, symbol_code)
                if ov:
                    ic = make_overlay(ic, ov)
                try:
                    rel_path = ic.relative_to(BASE_DIR)
                    icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                except ValueError:
                    icon_url = f"http://127.0.0.1:{self.http_port}/aprs_symbols_48/primary/29.png"
                
                # Build tooltip
                tooltip_parts = []
                tooltip_parts.append("📻 PyTNC Pro")
                if comment:
                    tooltip_parts.append(f"💬 {clean_aprs_comment(comment, 60)}")
                tooltip_parts.append(f"🌐 APRS-IS")
                tooltip_parts.append(f"🕐 {datetime.now().strftime('%H:%M:%S')}")
                tooltip = "<br>".join(tooltip_parts)
                
                # Use JSON encoding for proper escaping
                import json
                call_js = json.dumps(full_call)
                tooltip_js = json.dumps(tooltip)
                
                js = f"queueStation({call_js},{lat},{lon},'{icon_url}',{tooltip_js},false,\"\")"
                if self.map_ready:
                    self.map.page().runJavaScript(js)
            except Exception as e:
                self._log(f"  (Could not plot on map: {e})")
                
        except Exception as e:
            self._log(f"❌ APRS-IS send failed: {e}")
    
    def _toggle_auto_beacon(self, state):
        """Enable or disable auto-beacon"""
        if state == Qt.CheckState.Checked.value:
            # Start auto-beacon
            interval_mins = max(1, self.auto_beacon_interval.value())
            self.auto_beacon_countdown = interval_mins * 60  # Convert to seconds
            self.auto_beacon_timer.start(1000)  # Tick every second
            self.auto_beacon_status.setText(f"Next beacon in: {interval_mins}:00")
            self.auto_beacon_status.setStyleSheet("color: #69f0ae;")
            self._log(f"⏱️ Auto-beacon enabled: every {interval_mins} minutes")
            
            # Log to TX log
            mode = self.auto_beacon_mode.currentData()
            mode_str = {"is": "APRS-IS", "rf": "RF", "both": "RF + APRS-IS"}[mode]
            self.preset_log.append(f"<br><span style='color:#ffd54f'>⏱️ Auto-beacon started: {mode_str} every {interval_mins} min</span>")
        else:
            # Stop auto-beacon
            self.auto_beacon_timer.stop()
            self.auto_beacon_status.setText("Auto-beacon: Off")
            self.auto_beacon_status.setStyleSheet("color: #607d8b;")
            self._log("⏱️ Auto-beacon disabled")
            self.preset_log.append("<span style='color:#ff6b6b'>⏱️ Auto-beacon stopped</span>")
    
    def _update_auto_beacon_interval(self, value):
        """Update auto-beacon interval"""
        if self.auto_beacon_enabled.isChecked():
            # Reset countdown to new interval
            self.auto_beacon_countdown = value * 60
            self._log(f"⏱️ Auto-beacon interval changed to {value} minutes")
    
    def _auto_beacon_tick(self):
        """Called every second when auto-beacon is enabled"""
        self.auto_beacon_countdown -= 1
        
        # Update countdown display
        mins = self.auto_beacon_countdown // 60
        secs = self.auto_beacon_countdown % 60
        self.auto_beacon_status.setText(f"Next beacon in: {mins}:{secs:02d}")
        
        if self.auto_beacon_countdown <= 0:
            # Time to send beacon
            self._send_auto_beacon()
            
            # Reset countdown — enforce minimum 1 minute to prevent runaway
            interval_mins = max(1, self.auto_beacon_interval.value())
            self.auto_beacon_countdown = interval_mins * 60
    
    def _send_auto_beacon(self):
        """Send beacon based on auto-beacon mode"""
        mode = self.auto_beacon_mode.currentData()

        # If GPS active, skip beacon if we haven't moved MIN_BEACON_DISTANCE_M
        if self.gps_has_fix and not getattr(self, 'gps_moved_enough', True):
            self._log(f"⏱️ Auto-beacon skipped — moved less than {self.MIN_BEACON_DISTANCE_M}m")
            return

        self._log("⏱️ Auto-beacon triggered")
        
        if mode == "is" or mode == "both":
            # Send via APRS-IS
            if self.aprs_is_running and self.aprs_is_socket:
                self._send_beacon_aprs_is()
            else:
                self._log("⚠️ APRS-IS not connected, skipping")
        
        if mode == "rf" or mode == "both":
            # Send via RF
            self.send_beacon()
    
    def _save_settings_from_tab(self):
        """Save settings from the settings tab"""
        self.save_settings()
        self._log("✓ Settings saved", "#69f0ae")
    
    def _sync_callsign_to_beacon(self, text):
        """Sync callsign from Settings to Beacon Settings"""
        if hasattr(self, 'callsign_edit'):
            self.callsign_edit.setText(text.upper())
    
    def _sync_ssid_to_beacon(self, index):
        """Sync SSID from Settings to Beacon Settings and update type label"""
        if hasattr(self, 'ssid_combo'):
            self.ssid_combo.setCurrentIndex(index)
        
        # Update SSID type description
        ssid_types = {
            0: "Primary", 1: "Secondary", 2: "Secondary", 3: "Additional", 4: "Additional",
            5: "IGate", 6: "Satellite", 7: "Handheld", 8: "Boat", 9: "Mobile",
            10: "Internet", 11: "Balloon", 12: "Portable", 13: "Weather", 14: "Truck", 15: "Digipeater"
        }
        ssid = self.settings_ssid_combo.currentData()
        if hasattr(self, 'ssid_type_label'):
            self.ssid_type_label.setText(ssid_types.get(ssid, ""))
    
    def _sync_beacon_ssid_to_settings(self, index):
        """Sync SSID from Beacon Settings to Settings tab"""
        if hasattr(self, 'settings_ssid_combo'):
            self.settings_ssid_combo.setCurrentIndex(index)

    def _group_style(self):
        return """
            QGroupBox {
                color: #a0c4ff;
                font-weight: bold;
                border: 1px solid #1e3a5f;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 10px;
                background: #0d2137;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                background: #0d2137;
            }
            QLabel { color: #b0bec5; }
            QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
                background: #1a3a5c;
                color: #e0e0e0;
                border: 1px solid #2a5a8a;
                border-radius: 4px;
                padding: 4px;
            }
        """
    
    def _button_style(self, color1, color2):
        return f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {color1}, stop:1 {color2});
                color: white;
                font-weight: bold;
                border: 1px solid {color2};
                border-radius: 6px;
                padding: 8px 16px;
            }}
            QPushButton:hover {{
                background: {color2};
            }}
            QPushButton:disabled {{
                background: #1a3a5c;
                border-color: #2a5a8a;
                color: #607d8b;
            }}
        """
    
    def _branding_label(self):
        """Create small branding label for bottom-right of tabs"""
        label = QLabel("PyTNC Pro by KO6IKR © 2026")
        label.setStyleSheet("color: #607d8b; font-size: 11px;")
        label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignBottom)
        return label
    
    # =========================================================================
    # MESSAGING FUNCTIONS
    # =========================================================================
    
    def _refresh_message_history(self):
        """Refresh the message history display"""
        self.msg_history.clear()
        if not self.current_conv or self.current_conv not in self.conversations:
            return
        
        my_call = f"{self.callsign_edit.text().strip().upper()}-{self.ssid_combo.currentData()}"
        
        html = ""
        for msg in self.conversations[self.current_conv]:
            time_str = msg.get("time", "")
            text = msg.get("text", "")
            
            if msg.get("from") == my_call:
                # Outgoing message (right aligned, blue)
                ack_icon = "✓" if msg.get("acked") else "⏳"
                html += f'''<div style="text-align: right; margin: 5px;">
                    <span style="background: #1565c0; color: white; padding: 5px 10px; border-radius: 10px; display: inline-block;">
                    {text}</span><br>
                    <span style="color: #607d8b; font-size: 10px;">{time_str} {ack_icon}</span>
                </div>'''
            else:
                # Incoming message (left aligned, gray)
                html += f'''<div style="text-align: left; margin: 5px;">
                    <span style="background: #37474f; color: white; padding: 5px 10px; border-radius: 10px; display: inline-block;">
                    {text}</span><br>
                    <span style="color: #607d8b; font-size: 10px;">{time_str}</span>
                </div>'''
        
        self.msg_history.setHtml(html)
        # Scroll to bottom
        scrollbar = self.msg_history.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def _update_conversation_list(self):
        """Update the conversations list widget"""
        self.conv_list.clear()
        my_call = f"{self.callsign_edit.text().strip().upper()}-{self.ssid_combo.currentData()}"
        
        for callsign, messages in self.conversations.items():
            if messages:
                last_msg = messages[-1]
                preview = last_msg.get("text", "")[:20]
                if len(last_msg.get("text", "")) > 20:
                    preview += "..."
                
                item = QListWidgetItem(f"{callsign}\n{preview}")
                item.setData(Qt.ItemDataRole.UserRole, callsign)
                
                # Check for unacked outgoing messages
                has_unacked = any(not m.get("acked") and m.get("from") == my_call for m in messages)
                if has_unacked:
                    item.setForeground(QColor("#ffb74d"))  # Orange for pending
                
                self.conv_list.addItem(item)
    
    def _save_conversations(self):
        """Save conversations to JSON file"""
        try:
            with open(self.conversations_file, 'w') as f:
                json.dump({
                    "conversations": self.conversations,
                    "msg_seq": self.msg_seq
                }, f, indent=2)
        except Exception as e:
            self._log(f"⚠️ Failed to save conversations: {e}")
    
    def _send_message_rf(self, from_call, info):
        """Send message via RF"""
        self.msg_status.setText("📡 Sending via RF...")
        self.msg_status.setStyleSheet("color: #64b5f6;")
        
        # Build and send the packet (similar to beacon)
        path_str = self.path_combo.currentText().strip()
        if path_str.upper() == "DIRECT":
            path = []
        else:
            path = []
            for p in path_str.split(","):
                p = p.strip()
                if "-" in p:
                    call, ssid = p.split("-")
                    path.append((call, int(ssid)))
                else:
                    path.append((p, 0))
        
        my_call = self.callsign_edit.text().strip().upper()
        my_ssid = self.ssid_combo.currentData()
        
        # Log the actual packet that will be sent
        path_display = ",".join([f"{c}-{s}" if s else c for c, s in path]) if path else "DIRECT"
        self._log(f"📤 MSG TX: {from_call}>APPR01,{path_display}:{info}")
        
        packet = APRSPacketBuilder.build_ui_packet(
            my_call, my_ssid, "APPR01", 0, path, info
        )
        fcs = APRSPacketBuilder.compute_fcs(packet)
        frame = packet + bytes([fcs & 0xFF, (fcs >> 8) & 0xFF])
        
        mod = AFSKModulator(TX_SAMPLE_RATE)
        audio = mod.generate_packet_audio(frame, preamble_flags=40, postamble_flags=8)
        
        tx_level = self.tx_level_slider.value() / 100.0
        audio = audio * tx_level
        audio = apply_cosine_ramp(audio, TX_SAMPLE_RATE, ramp_ms=10.0)
        
        # Get TX device
        tx_idx = self.tx_audio_combo.currentData()
        if tx_idx is None:
            self.msg_status.setText("❌ No TX audio device")
            self.msg_status.setStyleSheet("color: #ef5350;")
            return
        
        # PTT on, play, PTT off
        try:
            self._set_ptt(True)
            time.sleep(0.05)
            sd.play(audio, TX_SAMPLE_RATE, device=tx_idx)
            sd.wait()
            time.sleep(0.05)
            self._set_ptt(False)
            
            self.msg_status.setText("✓ Sent via RF")
            self.msg_status.setStyleSheet("color: #69f0ae;")
        except Exception as e:
            self._set_ptt(False)
            self.msg_status.setText(f"❌ TX failed: {e}")
            self.msg_status.setStyleSheet("color: #ef5350;")
    
    def _send_message_is(self, from_call, to_call, info):
        """Send message via APRS-IS"""
        try:
            packet = f"{from_call}>APPR01,TCPIP*:{info}\r\n"
            self.aprs_is_socket.send(packet.encode())
            self.msg_status.setText("✓ Sent via APRS-IS")
            self.msg_status.setStyleSheet("color: #69f0ae;")
            self._log(f"📤 MSG IS: {packet.strip()}")
        except Exception as e:
            self.msg_status.setText(f"❌ Send failed: {e}")
            self.msg_status.setStyleSheet("color: #ef5350;")
    
    def _handle_incoming_message(self, from_call, to_call, message, seq=None):
        """Handle an incoming APRS message"""
        import datetime
        now = datetime.datetime.now().strftime("%H:%M:%S")
        
        # Store the message
        if from_call not in self.conversations:
            self.conversations[from_call] = []
        
        self.conversations[from_call].append({
            "from": from_call,
            "to": to_call,
            "text": message,
            "time": now,
            "acked": True,  # Incoming messages don't need ack tracking
            "seq": seq
        })
        
        self._update_conversation_list()
        self._save_conversations()  # Auto-save
        
        # If this conversation is currently selected, refresh it
        if self.current_conv == from_call:
            self._refresh_message_history()
        
        # Send ack if sequence number provided
        if seq:
            self._send_ack(from_call, seq)
        
        # Log it
        self._log(f"📨 MSG from {from_call}: {message}")
        
        # Flash the tab or show notification
        self.tabs.setTabText(2, "💬 Messages *")
    
    def _handle_ack(self, from_call, seq):
        """Handle an incoming message acknowledgment"""
        my_call = f"{self.callsign_edit.text().strip().upper()}-{self.ssid_combo.currentData()}"
        
        # Find and mark the message as acked
        for callsign, messages in self.conversations.items():
            for msg in messages:
                if msg.get("seq") == seq and msg.get("from") == my_call and not msg.get("acked"):
                    msg["acked"] = True
                    self._log(f"✓ ACK received from {from_call} for msg #{seq}")
                    self._update_conversation_list()
                    self._save_conversations()  # Auto-save
                    if self.current_conv == callsign:
                        self._refresh_message_history()
                    return
    
    def _send_ack(self, to_call, seq):
        """Send an acknowledgment for a received message"""
        my_call = self.callsign_edit.text().strip().upper()
        my_ssid = self.ssid_combo.currentData()
        full_call = f"{my_call}-{my_ssid}" if my_ssid > 0 else my_call
        
        dest_padded = f"{to_call:9s}"
        info = f":{dest_padded}:ack{seq}"
        
        # Send via APRS-IS if connected
        if hasattr(self, 'aprs_is_socket') and self.aprs_is_socket:
            try:
                packet = f"{full_call}>APPR01,TCPIP*:{info}\r\n"
                self.aprs_is_socket.send(packet.encode())
                self._log(f"📤 ACK: {packet.strip()}")
            except (OSError, BrokenPipeError) as e:
                self._log(f"⚠️ ACK send failed: {e}")
    
    def _build_symbol_grid(self):
        """Build the clickable symbol grid from cached icons"""
        self._update_symbol_grid()
    
    def _update_symbol_grid(self):
        """Update symbol grid based on selected table"""
        # Clear existing grid
        while self.symbol_grid_layout.count():
            item = self.symbol_grid_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        # Determine which table
        table_idx = self.symbol_table_picker.currentIndex()
        table_char = "/" if table_idx == 0 else "\\"
        prefix = "primary" if table_idx == 0 else "secondary"
        
        # Symbol descriptions for tooltips (from APRS spec)
        # PRIMARY table (/)
        primary_names = {
            "!": "Police", "\"": "Reserved", "#": "Digi", "$": "Phone",
            "%": "DX Cluster", "&": "HF Gateway", "'": "Small Aircraft", "(": "Mobile Satellite",
            ")": "Wheelchair", "*": "Snowmobile", "+": "Red Cross", ",": "Boy Scouts",
            "-": "House QTH", ".": "X", "/": "Red Dot", "0": "Circle 0",
            "1": "Circle 1", "2": "Circle 2", "3": "Circle 3", "4": "Circle 4",
            "5": "Circle 5", "6": "Circle 6", "7": "Circle 7", "8": "Circle 8",
            "9": "Circle 9", ":": "Fire", ";": "Campground", "<": "Motorcycle",
            "=": "Railroad", ">": "Car", "?": "File Server", "@": "Hurricane",
            "A": "Aid Station", "B": "BBS", "C": "Canoe", "D": "TBD",
            "E": "Eyeball", "F": "Tractor", "G": "Grid Square", "H": "Hotel",
            "I": "TCP/IP", "J": "TBD", "K": "School", "L": "PC User",
            "M": "Mac", "N": "NTS Station", "O": "Balloon", "P": "Police",
            "Q": "TBD", "R": "RV", "S": "Shuttle", "T": "SSTV",
            "U": "Bus", "V": "ATV", "W": "Weather Svc", "X": "Helicopter",
            "Y": "Yacht", "Z": "Windows", "[": "Jogger", "\\": "Triangle DF",
            "]": "Mailbox", "^": "Large Aircraft", "_": "Weather Stn", "`": "Dish Antenna",
            "a": "Ambulance", "b": "Bike", "c": "Incident Cmd", "d": "Fire Dept",
            "e": "Horse", "f": "Fire Truck", "g": "Glider", "h": "Hospital",
            "i": "IOTA", "j": "Jeep", "k": "Truck", "l": "Laptop",
            "m": "Mic-E Repeater", "n": "Node", "o": "EOC", "p": "Rover (Dog)",
            "q": "Grid Square", "r": "Repeater", "s": "Ship", "t": "Truck Stop",
            "u": "Truck 18-Wheeler", "v": "Van", "w": "Water Station", "x": "xAPRS",
            "y": "Yagi", "z": "Shelter", "{": "TBD", "|": "Reserved",
            "}": "Reserved", "~": "Reserved"
        }
        
        # SECONDARY/ALTERNATE table (\) - different meanings!
        secondary_names = {
            "!": "Emergency", "\"": "Reserved", "#": "Digi (alt)", "$": "Bank/ATM",
            "%": "TBD", "&": "HF Gateway (alt)", "'": "Crash Site", "(": "Cloudy",
            ")": "Firenet", "*": "Snow", "+": "Church", ",": "Girl Scouts",
            "-": "House HF", ".": "Ambiguous", "/": "TBD", "0": "Circle (E)",
            "1": "TBD", "2": "TBD", "3": "TBD", "4": "TBD",
            "5": "TBD", "6": "TBD", "7": "TBD", "8": "802.11",
            "9": "Gas Station", ":": "Hail", ";": "Park", "<": "Advisory",
            "=": "TBD", ">": "Car (alt)", "?": "Info Kiosk", "@": "Hurricane",
            "A": "Box", "B": "Blowing Snow", "C": "Coast Guard", "D": "Drizzle",
            "E": "Smoke", "F": "Freezing Rain", "G": "Snow Shower", "H": "Haze",
            "I": "Rain Shower", "J": "Lightning", "K": "Kenwood", "L": "Lighthouse",
            "M": "TBD", "N": "Nav Buoy", "O": "Rocket", "P": "Parking",
            "Q": "Quake", "R": "Restaurant", "S": "Satellite", "T": "Thunderstorm",
            "U": "Sunny", "V": "VORTAC", "W": "NWS Site", "X": "Pharmacy",
            "Y": "TBD", "Z": "TBD", "[": "Wall Cloud", "\\": "TBD",
            "]": "TBD", "^": "Aircraft (alt)", "_": "WX Stn (Blue)", "`": "Rain",
            "a": "ARRL/ARES", "b": "Blowing Dust", "c": "Civil Defense", "d": "DX Spot",
            "e": "Sleet", "f": "Funnel Cloud", "g": "Gale Flags", "h": "Store",
            "i": "Indoor Box", "j": "Work Zone", "k": "SUV", "l": "TBD",
            "m": "Value Sign", "n": "Triangle", "o": "Small Circle", "p": "Part Cloudy",
            "q": "TBD", "r": "Restrooms", "s": "Ship (alt)", "t": "Tornado",
            "u": "Truck (alt)", "v": "Van (alt)", "w": "Flooding", "x": "TBD",
            "y": "Sky Warn", "z": "Shelter (alt)", "{": "Fog", "|": "Reserved",
            "}": "Reserved", "~": "Reserved"
        }
        
        symbol_names = primary_names if table_idx == 0 else secondary_names
        
        # Build grid - ASCII 33 (!) to 126 (~)
        row, col = 0, 0
        cols = 12  # 12 columns
        
        for code in range(33, 127):
            char = chr(code)
            # Hessu icons are numbered 00-94, corresponding to ASCII 33-126
            hessu_num = code - 33
            icon_path = HESSU_SYMBOLS_DIR / prefix / f"{hessu_num:02d}.png"
            
            btn = QPushButton()
            btn.setFixedSize(34, 34)
            btn.setToolTip(f"{table_char}{char} - {symbol_names.get(char, 'Symbol')}")
            
            if icon_path.exists():
                # Use QIcon for reliable cross-platform icon display
                btn.setIcon(QIcon(str(icon_path)))
                btn.setIconSize(QSize(28, 28))
                btn.setStyleSheet("""
                    QPushButton {
                        background: transparent;
                        border: none;
                    }
                    QPushButton:hover {
                        background: rgba(66, 165, 245, 0.3);
                    }
                    QPushButton:pressed {
                        background: rgba(105, 240, 174, 0.3);
                    }
                """)
            else:
                btn.setText(char)
                btn.setStyleSheet("""
                    QPushButton {
                        background: transparent;
                        border: none;
                        color: #a0c4ff;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: rgba(66, 165, 245, 0.3);
                    }
                """)
            
            btn.clicked.connect(partial(self._symbol_clicked, table_char, char, symbol_names.get(char, "Symbol")))
            self.symbol_grid_layout.addWidget(btn, row, col)
            
            col += 1
            if col >= cols:
                col = 0
                row += 1
    
    def _symbol_clicked(self, table: str, code: str, name: str):
        """Handle symbol button click"""
        # Check if overlay is set - overlay char replaces table byte
        overlay = ""
        if hasattr(self, 'symbol_overlay_edit'):
            overlay = self.symbol_overlay_edit.text().strip().upper()
            if overlay and (overlay.isalpha() or overlay.isdigit()):
                # Overlay mode: table byte = overlay char, use secondary icons
                self.symbol_table_combo.setCurrentIndex(1)  # secondary
                self.symbol_code_edit.setText(code)
                # Store overlay char as table value
                self.symbol_table_combo.setCurrentText(overlay)
            else:
                table_idx = 0 if table == "/" else 1
                self.symbol_table_combo.setCurrentIndex(table_idx)
                self.symbol_code_edit.setText(code)
        else:
            table_idx = 0 if table == "/" else 1
            self.symbol_table_combo.setCurrentIndex(table_idx)
            self.symbol_code_edit.setText(code)

        # Update the symbol preview
        prefix = "secondary" if (table == "\\" or overlay) else "primary"
        hessu_num = ord(code) - 33
        icon_p = HESSU_SYMBOLS_DIR / prefix / f"{hessu_num:02d}.png"

        if icon_p.exists():
            pix = QPixmap(str(icon_p)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            if overlay:
                # Render overlay char on preview
                overlaid = make_overlay(icon_p, overlay)
                pix = QPixmap(str(overlaid)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.symbol_preview.setPixmap(pix)

        overlay_str = f" [{overlay}]" if overlay else ""
        self.preset_log.append(f"🎨 Symbol: {table}{code}{overlay_str} - {name}")

        self.symbol_table_combo.repaint()
        self.symbol_code_edit.repaint()

    def _on_overlay_changed(self, text):
        """Update symbol preview when overlay char changes"""
        overlay = text.strip().upper()
        code = self.symbol_code_edit.text()
        if not code:
            return
        # Determine current table
        table_idx = self.symbol_table_picker.currentIndex()
        prefix = "secondary" if table_idx == 1 or overlay else "primary"
        hessu_num = ord(code) - 33
        icon_p = HESSU_SYMBOLS_DIR / prefix / f"{hessu_num:02d}.png"
        if icon_p.exists():
            if overlay and (overlay.isalpha() or overlay.isdigit()):
                overlaid = make_overlay(icon_p, overlay)
                pix = QPixmap(str(overlaid)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            else:
                pix = QPixmap(str(icon_p)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.symbol_preview.setPixmap(pix)
    
    def load_com_ports(self):
        """Load available COM ports into PTT combo boxes"""
        self.ptt_port_combo.clear()
        if HAS_SERIAL:
            ports = serial.tools.list_ports.comports()
            for port in ports:
                self.ptt_port_combo.addItem(f"{port.device} - {port.description}", port.device)
            
            # Try to select COM4 (Standard) for PTT by default
            for i in range(self.ptt_port_combo.count()):
                if "Standard" in self.ptt_port_combo.itemText(i) or "COM4" in self.ptt_port_combo.itemText(i):
                    self.ptt_port_combo.setCurrentIndex(i)
                    break
        else:
            self.ptt_port_combo.addItem("pyserial not installed")
    
    def load_tx_devices(self):
        """Load available audio OUTPUT devices for TX"""
        # sounddevice is optional - if not available, disable TX device selection
        if not HAS_SOUNDDEVICE or sd is None:
            try:
                self.tx_audio_combo.clear()
                self.tx_audio_combo.addItem("Audio TX disabled (sounddevice unavailable)")
                self.tx_audio_combo.setEnabled(False)
            except Exception:
                pass
            return
        
        try:
            devices = list(sd.query_devices())
        except Exception as e:
            # PortAudio / driver issues should not crash the app
            try:
                self.tx_audio_combo.clear()
                self.tx_audio_combo.addItem(f"Audio TX disabled ({type(e).__name__})")
                self.tx_audio_combo.setEnabled(False)
            except Exception:
                pass
            return
        
        try:
            self.tx_audio_combo.clear()
            self.tx_audio_combo.setEnabled(True)
            for i, d in enumerate(devices):
                # Only show devices with output channels
                if d.get("max_output_channels", 0) > 0:
                    self.tx_audio_combo.addItem(f"{i}: {d['name']}", i)
            
            # Try to select USB audio codec (FT-991A) by default
            for i in range(self.tx_audio_combo.count()):
                text = self.tx_audio_combo.itemText(i).lower()
                if "usb" in text and ("codec" in text or "audio" in text):
                    self.tx_audio_combo.setCurrentIndex(i)
                    break
        except Exception:
            pass
    
    def _sync_beacon_connection_status(self):
        """Sync all connection status from Settings to Beacon tab"""
        # PTT Status
        if self._ptt_is_connected():
            self.tx_ptt_status.setText(f"🟢 PTT: {self._ptt_port_label()}")
            self.tx_ptt_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_ptt_status.setText("⚫ PTT: Not connected")
            self.tx_ptt_status.setStyleSheet("color: #607d8b;")
        
        # GPS Status
        if self.gps_serial and self.gps_serial.is_open:
            self.tx_gps_status.setText(f"🟢 GPS: {self.gps_serial.port}")
            self.tx_gps_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_gps_status.setText("⚫ GPS: Not connected")
            self.tx_gps_status.setStyleSheet("color: #607d8b;")
        
        # TX Audio Status
        if hasattr(self, 'settings_tx_audio_combo'):
            tx_device = self.settings_tx_audio_combo.currentData()
            tx_name = self.settings_tx_audio_combo.currentText()
            if tx_device is not None and tx_name:
                # Truncate long names
                short_name = tx_name[:20] + "..." if len(tx_name) > 20 else tx_name
                self.tx_audio_status.setText(f"🟢 TX Audio: {short_name}")
                self.tx_audio_status.setStyleSheet("color: #69f0ae;")
            else:
                self.tx_audio_status.setText("⚫ TX Audio: Not set")
                self.tx_audio_status.setStyleSheet("color: #607d8b;")
        
        # VARA FM Status
        if hasattr(self, 'vara_connected') and self.vara_connected:
            self.tx_vara_status.setText("🟢 VARA: Connected")
            self.tx_vara_status.setStyleSheet("color: #69f0ae;")
            # Also update Settings tab indicator
            if hasattr(self, 'settings_vara_status'):
                self.settings_vara_status.setText("🟢")
                self.settings_vara_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_vara_status.setText("⚫ VARA: Not connected")
            self.tx_vara_status.setStyleSheet("color: #607d8b;")
            # Also update Settings tab indicator
            if hasattr(self, 'settings_vara_status'):
                self.settings_vara_status.setText("⚫")
                self.settings_vara_status.setStyleSheet("color: #607d8b;")
        
        # RF RX Status (receiver running)
        if hasattr(self, 'tx_rf_status'):
            if hasattr(self, 'receiver') and self.receiver and self.receiver.isRunning():
                # Get RX device from Settings tab
                if hasattr(self, 'settings_rx_audio_combo'):
                    rx_name = self.settings_rx_audio_combo.currentText()
                    short_name = rx_name[:20] + "..." if len(rx_name) > 20 else rx_name
                    self.tx_rf_status.setText(f"🟢 RF: {short_name}")
                else:
                    self.tx_rf_status.setText("🟢 RF: Connected")
                self.tx_rf_status.setStyleSheet("color: #69f0ae;")
            else:
                self.tx_rf_status.setText("⚫ RF: Not connected")
                self.tx_rf_status.setStyleSheet("color: #607d8b;")
        
        # APRS-IS Status
        if hasattr(self, 'aprs_is_connected') and self.aprs_is_connected:
            self.tx_aprs_is_status.setText("🟢 APRS-IS: Connected")
            self.tx_aprs_is_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_aprs_is_status.setText("⚫ APRS-IS: Not connected")
            self.tx_aprs_is_status.setStyleSheet("color: #607d8b;")

    def _sync_vara_fm_connection_status(self):
        """Sync all connection statuses on VARA FM tab"""
        # PTT Status
        if self._ptt_is_connected():
            self.vara_ptt_status.setText(f"🟢 PTT: {self._ptt_port_label()}")
            self.vara_ptt_status.setStyleSheet("color: #69f0ae;")
        else:
            self.vara_ptt_status.setText("⚫ PTT: Not connected")
            self.vara_ptt_status.setStyleSheet("color: #607d8b;")
        
        # GPS Status
        if self.gps_serial and self.gps_serial.is_open:
            self.vara_gps_status.setText(f"🟢 GPS: {self.gps_serial.port}")
            self.vara_gps_status.setStyleSheet("color: #69f0ae;")
        else:
            self.vara_gps_status.setText("⚫ GPS: Not connected")
            self.vara_gps_status.setStyleSheet("color: #607d8b;")
        
        # TX Audio Status
        if hasattr(self, 'settings_tx_audio_combo'):
            tx_device = self.settings_tx_audio_combo.currentData()
            tx_name = self.settings_tx_audio_combo.currentText()
            if tx_device is not None and tx_name:
                short_name = tx_name[:20] + "..." if len(tx_name) > 20 else tx_name
                self.vara_tx_status.setText(f"🟢 TX: {short_name}")
                self.vara_tx_status.setStyleSheet("color: #69f0ae;")
            else:
                self.vara_tx_status.setText("⚫ TX Audio: Not set")
                self.vara_tx_status.setStyleSheet("color: #607d8b;")
        
        # VARA FM Status
        if hasattr(self, 'vara_connected') and self.vara_connected:
            self.vara_fm_status.setText("🟢 VARA: Connected")
            self.vara_fm_status.setStyleSheet("color: #69f0ae;")
            self.vara_connect_btn.setText("🔌 Disconnect")
            # Also update Settings tab indicator
            if hasattr(self, 'settings_vara_status'):
                self.settings_vara_status.setText("🟢")
                self.settings_vara_status.setStyleSheet("color: #69f0ae;")
        else:
            self.vara_fm_status.setText("⚫ VARA: Not connected")
            self.vara_fm_status.setStyleSheet("color: #607d8b;")
            self.vara_connect_btn.setText("🔌 Connect")
            # Also update Settings tab indicator
            if hasattr(self, 'settings_vara_status'):
                self.settings_vara_status.setText("⚫")
                self.settings_vara_status.setStyleSheet("color: #607d8b;")
        
        # APRS-IS Status
        if hasattr(self, 'aprs_is_connected') and self.aprs_is_connected:
            self.vara_aprs_status.setText("🟢 APRS-IS: Connected")
            self.vara_aprs_status.setStyleSheet("color: #69f0ae;")
        else:
            self.vara_aprs_status.setText("⚫ APRS-IS: Not connected")
            self.vara_aprs_status.setStyleSheet("color: #607d8b;")

    def _open_vara_fm(self):
        """Open VARA FM application"""
        import subprocess
        import os
        
        # Common VARA FM installation paths on Windows
        vara_paths = [
            r"C:\VARA FM\VARAFM.exe",
            r"C:\VARA FM\VARA FM.exe",
            r"C:\VARA FM\VARA.exe",
            r"C:\Program Files\VARA FM\VARAFM.exe",
            r"C:\Program Files\VARA FM\VARA FM.exe",
            r"C:\Program Files (x86)\VARA FM\VARAFM.exe",
            r"C:\Program Files (x86)\VARA FM\VARA FM.exe",
            os.path.expanduser(r"~\Desktop\VARA FM\VARAFM.exe"),
            os.path.expanduser(r"~\Desktop\VARA FM\VARA FM.exe"),
        ]
        
        for path in vara_paths:
            if os.path.exists(path):
                try:
                    subprocess.Popen([path], shell=True)
                    self._vara_log("📂 Launched VARA FM")
                    return
                except Exception as e:
                    self._vara_log(f"❌ Failed to launch: {e}")
        
        self._vara_log("⚠️ VARA FM.exe not found in common locations")
        self._vara_log("   Please launch VARA FM manually")

    def _toggle_vara_connection(self):
        """Toggle connection to VARA FM"""
        if hasattr(self, 'vara_connected') and self.vara_connected:
            self._disconnect_vara()
        else:
            self._connect_vara_from_tab()

    def _connect_vara_from_tab(self):
        """Connect to VARA FM from VARA FM tab"""
        host = self.vara_host.text() if hasattr(self, 'vara_host') else "localhost"
        cmd_port = self.vara_cmd_port.value() if hasattr(self, 'vara_cmd_port') else 8300
        data_port = self.vara_data_port.value() if hasattr(self, 'vara_data_port') else 8301
        kiss_port = self.vara_port.value() if hasattr(self, 'vara_port') else 8100
        
        self._vara_log(f"🔌 Connecting to VARA FM at {host}...")
        
        try:
            import socket
            # Connect command port
            self.vara_cmd_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.vara_cmd_socket.settimeout(5.0)
            self.vara_cmd_socket.connect((host, cmd_port))
            self._vara_log(f"   ✓ Command port {cmd_port} connected")
            
            # Connect data port
            self.vara_data_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.vara_data_socket.settimeout(5.0)
            self.vara_data_socket.connect((host, data_port))
            self._vara_log(f"   ✓ Data port {data_port} connected")
            
            # Connect KISS port (for APRS broadcast mode)
            try:
                self.vara_kiss_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                self.vara_kiss_socket.settimeout(5.0)
                self.vara_kiss_socket.connect((host, kiss_port))
                self.vara_kiss_connected = True
                self._vara_log(f"   ✓ KISS port {kiss_port} connected (APRS broadcast ready)")
            except Exception as e:
                self._vara_log(f"   ⚠️ KISS port {kiss_port} not available: {e}")
                self._vara_log(f"      (Connected mode still works, KISS broadcast disabled)")
                self.vara_kiss_connected = False
                self.vara_kiss_socket = None
            
            self.vara_connected = True
            self._vara_log("✅ VARA FM connected!")
            self._sync_vara_fm_connection_status()
            
            # Set our callsign in VARA FM
            try:
                callsign = self.callsign_edit.text().strip().upper() if hasattr(self, 'callsign_edit') else ""
                ssid = self.ssid_combo.currentData() if hasattr(self, 'ssid_combo') else 0
                if callsign and callsign != "N0CALL":
                    full_call = f"{callsign}-{ssid}" if ssid else callsign
                    self.vara_cmd_socket.send(f"MYCALL {full_call}\r".encode())
                    self._vara_log(f"📡 Set MYCALL: {full_call}")
                    time.sleep(0.2)
            except Exception as e:
                self._vara_log(f"⚠️ Could not set MYCALL: {e}")
            
            # Tell VARA FM to listen for incoming connections
            try:
                self.vara_cmd_socket.send(b"LISTEN ON\r")
                self._vara_log("📡 LISTEN mode enabled - accepting incoming calls")
            except Exception as e:
                self._vara_log(f"⚠️ Could not enable LISTEN mode: {e}")
            
            # Start RX thread to receive data
            self.vara_rx_running = True
            self.vara_rx_thread = threading.Thread(target=self._vara_data_rx_loop, daemon=True)
            self.vara_rx_thread.start()
            self._vara_log("📡 Data RX thread started - listening for APRS packets...")
            
            # Start command socket listener for incoming connection notifications
            self.vara_cmd_thread = threading.Thread(target=self._vara_cmd_rx_loop, daemon=True)
            self.vara_cmd_thread.start()
            self._vara_log("📡 Cmd RX thread started - listening for incoming calls...")
            
            # Start KISS RX thread if KISS connected
            if self.vara_kiss_connected:
                self.vara_kiss_rx_thread = threading.Thread(target=self._vara_kiss_rx_loop, daemon=True)
                self.vara_kiss_rx_thread.start()
                self._vara_log("📡 KISS RX thread started - listening for APRS broadcasts...")
            
        except ConnectionRefusedError:
            self._vara_log("❌ Connection refused - is VARA FM running?")
            self.vara_connected = False
        except socket.timeout:
            self._vara_log("❌ Connection timeout")
            self.vara_connected = False
        except Exception as e:
            self._vara_log(f"❌ Connection failed: {e}")
            self.vara_connected = False
        
        self._sync_vara_fm_connection_status()

    def _disconnect_vara(self):
        """Disconnect from VARA FM"""
        self._vara_log("🔌 Disconnecting from VARA FM...")
        
        # Stop RX thread
        self.vara_rx_running = False
        
        if hasattr(self, 'vara_cmd_socket') and self.vara_cmd_socket:
            try:
                self.vara_cmd_socket.close()
            except OSError:
                pass  # Socket already closed
            self.vara_cmd_socket = None
        
        if hasattr(self, 'vara_data_socket') and self.vara_data_socket:
            try:
                self.vara_data_socket.close()
            except OSError:
                pass  # Socket already closed
            self.vara_data_socket = None
        
        # Close KISS socket
        if hasattr(self, 'vara_kiss_socket') and self.vara_kiss_socket:
            try:
                self.vara_kiss_socket.close()
            except OSError:
                pass
            self.vara_kiss_socket = None
        self.vara_kiss_connected = False
        
        self.vara_connected = False
        self._vara_log("✅ Disconnected from VARA FM")
        self._sync_vara_fm_connection_status()

    def _vara_save_settings(self):
        """Save all settings (VARA FM tab uses same global save)"""
        self.save_settings()

    def _vara_build_symbol_grid(self):
        """Build the symbol picker grid for VARA FM tab"""
        # Clear existing
        while self.vara_symbol_grid_layout.count():
            item = self.vara_symbol_grid_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        table_idx = self.vara_symbol_table_picker.currentIndex()
        prefix = "primary" if table_idx == 0 else "secondary"
        table_char = "/" if table_idx == 0 else "\\"
        
        # Symbol name dictionaries (same as Beacon tab)
        primary_names = {
            "!": "Police", "\"": "Reserved", "#": "Digi", "$": "Phone",
            "%": "DX Cluster", "&": "HF Gateway", "'": "Small Aircraft", "(": "Mobile Satellite",
            ")": "Wheelchair", "*": "Snowmobile", "+": "Red Cross", ",": "Boy Scouts",
            "-": "House QTH", ".": "X", "/": "Red Dot", "0": "Circle 0",
            "1": "Circle 1", "2": "Circle 2", "3": "Circle 3", "4": "Circle 4",
            "5": "Circle 5", "6": "Circle 6", "7": "Circle 7", "8": "Circle 8",
            "9": "Circle 9", ":": "Fire", ";": "Campground", "<": "Motorcycle",
            "=": "Railroad", ">": "Car", "?": "File Server", "@": "Hurricane",
            "A": "Aid Station", "B": "BBS", "C": "Canoe", "D": "TBD",
            "E": "Eyeball", "F": "Tractor", "G": "Grid Square", "H": "Hotel",
            "I": "TCP/IP", "J": "TBD", "K": "School", "L": "PC User",
            "M": "Mac", "N": "NTS Station", "O": "Balloon", "P": "Police",
            "Q": "TBD", "R": "RV", "S": "Shuttle", "T": "SSTV",
            "U": "Bus", "V": "ATV", "W": "Weather Svc", "X": "Helicopter",
            "Y": "Yacht", "Z": "Windows", "[": "Jogger", "\\": "Triangle DF",
            "]": "Mailbox", "^": "Large Aircraft", "_": "Weather Stn", "`": "Dish Antenna",
            "a": "Ambulance", "b": "Bike", "c": "Incident Cmd", "d": "Fire Dept",
            "e": "Horse", "f": "Fire Truck", "g": "Glider", "h": "Hospital",
            "i": "IOTA", "j": "Jeep", "k": "Truck", "l": "Laptop",
            "m": "Mic-E Repeater", "n": "Node", "o": "EOC", "p": "Rover (Dog)",
            "q": "Grid Square", "r": "Repeater", "s": "Ship", "t": "Truck Stop",
            "u": "Truck 18-Wheeler", "v": "Van", "w": "Water Station", "x": "xAPRS",
            "y": "Yagi", "z": "Shelter", "{": "TBD", "|": "Reserved",
            "}": "Reserved", "~": "Reserved"
        }
        secondary_names = {
            "!": "Emergency", "\"": "Reserved", "#": "Digi (alt)", "$": "Bank/ATM",
            "%": "TBD", "&": "HF Gateway (alt)", "'": "Crash Site", "(": "Cloudy",
            ")": "Firenet", "*": "Snow", "+": "Church", ",": "Girl Scouts",
            "-": "House HF", ".": "Ambiguous", "/": "TBD", "0": "Circle (E)",
            "1": "TBD", "2": "TBD", "3": "TBD", "4": "TBD",
            "5": "TBD", "6": "TBD", "7": "TBD", "8": "802.11",
            "9": "Gas Station", ":": "Hail", ";": "Park", "<": "Advisory",
            "=": "TBD", ">": "Car (alt)", "?": "Info Kiosk", "@": "Hurricane",
            "A": "Box", "B": "Blowing Snow", "C": "Coast Guard", "D": "Drizzle",
            "E": "Smoke", "F": "Freezing Rain", "G": "Snow Shower", "H": "Haze",
            "I": "Rain Shower", "J": "Lightning", "K": "Kenwood", "L": "Lighthouse",
            "M": "TBD", "N": "Nav Buoy", "O": "Rocket", "P": "Parking",
            "Q": "Quake", "R": "Restaurant", "S": "Satellite", "T": "Thunderstorm",
            "U": "Sunny", "V": "VORTAC", "W": "NWS Site", "X": "Pharmacy",
            "Y": "TBD", "Z": "TBD", "[": "Wall Cloud", "\\": "TBD",
            "]": "TBD", "^": "Aircraft (alt)", "_": "WX Stn (Blue)", "`": "Rain",
            "a": "ARRL/ARES", "b": "Blowing Dust", "c": "Civil Defense", "d": "DX Spot",
            "e": "Sleet", "f": "Funnel Cloud", "g": "Gale Flags", "h": "Store",
            "i": "Indoor Box", "j": "Work Zone", "k": "SUV", "l": "TBD",
            "m": "Value Sign", "n": "Triangle", "o": "Small Circle", "p": "Part Cloudy",
            "q": "TBD", "r": "Restrooms", "s": "Ship (alt)", "t": "Tornado",
            "u": "Truck (alt)", "v": "Van (alt)", "w": "Flooding", "x": "TBD",
            "y": "Sky Warn", "z": "Shelter (alt)", "{": "Fog", "|": "Reserved",
            "}": "Reserved", "~": "Reserved"
        }
        symbol_names = primary_names if table_idx == 0 else secondary_names
        
        # Build 16x6 grid of symbols (96 symbols)
        row, col = 0, 0
        for i in range(96):
            sym_char = chr(33 + i)  # ASCII 33-128
            icon_path = HESSU_SYMBOLS_DIR / prefix / f"{i:02d}.png"
            sym_name = symbol_names.get(sym_char, "Symbol")
            
            btn = QPushButton()
            btn.setFixedSize(34, 34)
            btn.setToolTip(f"{table_char}{sym_char} - {sym_name}")
            
            if icon_path.exists():
                # Use QIcon for reliable cross-platform icon display
                btn.setIcon(QIcon(str(icon_path)))
                btn.setIconSize(QSize(28, 28))
                btn.setStyleSheet("""
                    QPushButton {
                        background: transparent;
                        border: none;
                    }
                    QPushButton:hover {
                        background: rgba(66, 165, 245, 0.3);
                    }
                    QPushButton:pressed {
                        background: rgba(105, 240, 174, 0.3);
                    }
                """)
            else:
                btn.setText(sym_char)
                btn.setStyleSheet("""
                    QPushButton {
                        background: transparent;
                        border: none;
                        color: #a0c4ff;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: rgba(66, 165, 245, 0.3);
                    }
                """)
            
            # Connect click to select this symbol
            btn.clicked.connect(lambda checked, t=table_char, c=sym_char, p=icon_path, n=sym_name: self._vara_select_symbol(t, c, p, n))
            
            self.vara_symbol_grid_layout.addWidget(btn, row, col)
            col += 1
            if col >= 12:  # 12 columns
                col = 0
                row += 1

    def _vara_update_symbol_grid(self):
        """Update VARA FM symbol grid when table changes"""
        self._vara_build_symbol_grid()

    def _vara_select_symbol(self, table, code, icon_path, name="Symbol"):
        """Select a symbol for VARA FM beacon"""
        from pathlib import Path
        if Path(icon_path).exists():
            self.vara_symbol_preview.setPixmap(
                QPixmap(str(icon_path)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            )
            self.vara_symbol_preview.setToolTip(f"{table}{code} - {name}")
        # Store selection
        self._vara_symbol_table = table
        self._vara_symbol_code = code
        self._vara_log(f"🎨 Selected: {table}{code} - {name}")

    # =========================================================================
    # Earthquake Monitor
    # =========================================================================
    
    def send_beacon(self):
        """Send an APRS beacon via radio - auto-connects PTT if needed"""
        # Check if sounddevice is available for RF transmit
        if not HAS_SOUNDDEVICE:
            QMessageBox.warning(self, "RF Disabled", "sounddevice not installed.\nRF AFSK transmit is disabled.\n\nUse VARA FM or APRS-IS instead.")
            return
        
        callsign = self.callsign_edit.text().strip().upper()
        ssid = self.ssid_combo.currentData()
        comment = self.comment_edit.text()
        radio = self.radio_combo.currentText().strip()
        path_str = self.path_combo.currentText()
        sym_table = self.symbol_table_combo.currentText()
        sym_code = self.symbol_code_edit.text()
        
        # GET COORDINATES: GPS if has fix, otherwise manual from Settings
        lat = None
        lon = None
        
        # Check if GPS has actual fix data
        if hasattr(self, 'gps_has_fix') and self.gps_has_fix and hasattr(self, 'gps_lat'):
            lat = self.gps_lat
            lon = self.gps_lon
            self._log(f"🛰️ Using GPS coordinates: {lat:.6f}, {lon:.6f}")
        else:
            # GPS not available or no fix - read from manual_location in Settings tab
            manual_text = self.manual_location.text().strip() if hasattr(self, 'manual_location') else ""
            if manual_text:
                try:
                    parts = manual_text.replace(" ", "").split(",")
                    if len(parts) == 2:
                        lat = float(parts[0])
                        lon = float(parts[1])
                        self._log(f"📍 Using manual coordinates: {lat:.6f}, {lon:.6f}")
                except ValueError:
                    pass
            
            # Fallback to lat_edit if manual parsing failed
            if lat is None:
                lat = self.lat_edit.value()
                lon = self.lon_edit.value()
                self._log(f"⚠️ Using fallback coordinates: {lat:.6f}, {lon:.6f}")
        
        # Append radio to comment if specified
        if radio:
            full_comment = f"{comment} [{radio}]" if comment else f"[{radio}]"
        else:
            full_comment = comment
        
        if not callsign or callsign == "N0CALL":
            QMessageBox.warning(self, "Invalid Callsign", "Please enter your callsign")
            return
        
        # Auto-connect PTT if not connected
        ptt_auto_connected = False
        if not self._ptt_is_connected():
            method = getattr(self, 'civ_ptt_method', 'RTS/DTR')
            if method == "CI-V CAT":
                ptt_port = self.civ_port_combo.currentData() if hasattr(self, 'civ_port_combo') else None
                if ptt_port:
                    try:
                        self._toggle_civ()
                        ptt_auto_connected = True
                        self.preset_log.append(f"✅ Auto-connected CI-V PTT: {ptt_port}")
                        self._update_tx_status()
                    except Exception as e:
                        QMessageBox.warning(self, "CI-V Connection Failed", f"Could not connect CI-V PTT:\n{e}\n\nConfigure in Settings tab.")
                        return
                else:
                    QMessageBox.warning(self, "CI-V Not Configured", "Configure CI-V port in Settings tab first")
                    return
            else:
                ptt_port = self.settings_ptt_combo.currentData() if hasattr(self, 'settings_ptt_combo') else None
                if ptt_port:
                    try:
                        self.ptt_serial = serial.Serial(ptt_port, 9600, timeout=0.1)
                        self._set_ptt(False)
                        ptt_auto_connected = True
                        self.preset_log.append(f"✅ Auto-connected PTT: {ptt_port}")
                        self._update_tx_status()
                    except Exception as e:
                        QMessageBox.warning(self, "PTT Connection Failed", f"Could not connect PTT:\n{e}\n\nConfigure in Settings tab.")
                        return
                else:
                    QMessageBox.warning(self, "PTT Not Configured", "Configure PTT port in Settings tab first")
                    return
        
        # Get TX audio OUTPUT device from Settings
        tx_device = self.settings_tx_audio_combo.currentData() if hasattr(self, 'settings_tx_audio_combo') else None
        if tx_device is None:
            QMessageBox.warning(self, "No TX Audio Device", "Select TX audio output device in Settings tab first")
            return
        
        # Get TX level from Settings
        tx_level_pct = self.settings_tx_level.value() if hasattr(self, 'settings_tx_level') else 10
        
        # Save settings when beacon is attempted
        self.save_settings()
        
        # Format position for APRS
        lat_deg = int(abs(lat))
        lat_min = (abs(lat) - lat_deg) * 60
        lat_dir = "N" if lat >= 0 else "S"
        
        lon_deg = int(abs(lon))
        lon_min = (abs(lon) - lon_deg) * 60
        lon_dir = "E" if lon >= 0 else "W"
        
        # Build APRS position string
        info = f"!{lat_deg:02d}{lat_min:05.2f}{lat_dir}{sym_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{sym_code}{full_comment}"
        
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        
        # Parse path
        path_list = []
        if path_str and path_str != "DIRECT":
            for p in path_str.split(","):
                p = p.strip()
                if "-" in p:
                    pcall, pssid = p.rsplit("-", 1)
                    path_list.append((pcall, int(pssid)))
                else:
                    path_list.append((p, 0))
        
        self.preset_log.append(f"<br><span style='color:#64b5f6'>📡 Transmitting beacon...</span>")
        self.preset_log.append(f"   From: <span style='color:#ffd54f'>{full_call}</span>")
        self.preset_log.append(f"   To: APPR01-0")
        self.preset_log.append(f"   Path: <span style='color:#ce93d8'>{path_str}</span>")
        self.preset_log.append(f"   Info: <span style='color:#80deea'>{info}</span>")
        
        self._log(f"📡 TX Beacon: {full_call} via {path_str}")
        
        # Set TX in progress flag to prevent self-decode
        self.tx_in_progress = True
        
        try:
            # Build AX.25 UI packet
            # Using APPR01 as official TOCALL (APPR0? registered to KO6IKR)
            packet_data = APRSPacketBuilder.build_ui_packet(
                src_call=callsign, src_ssid=ssid,
                dst_call="APPR01", dst_ssid=0,  # Official PyTNC Pro tocall
                path=path_list,
                info=info
            )
            
            # Compute FCS
            fcs = APRSPacketBuilder.compute_fcs(packet_data)
            fcs_bytes = bytes([fcs & 0xFF, (fcs >> 8) & 0xFF])
            full_packet = packet_data + fcs_bytes
            
            # Debug: show the raw packet hex
            hex_dump = ' '.join(f'{b:02X}' for b in full_packet)
            self.preset_log.append(f"   Raw packet ({len(full_packet)} bytes):")
            self.preset_log.append(f"   {hex_dump}")
            self.preset_log.append(f"   FCS: {fcs:04X}")
            
            # Create modulator and generate audio
            modulator = AFSKModulator(TX_SAMPLE_RATE)
            audio = modulator.generate_packet_audio(full_packet, preamble_flags=60, postamble_flags=10)
            
            # Add short silence at start and end
            silence = np.zeros(int(TX_SAMPLE_RATE * 0.03), dtype=np.float32)  # 30ms
            audio = np.concatenate([silence, audio, silence])
            
            # Apply cosine ramp FIRST to reduce clicks (before any scaling)
            audio = apply_cosine_ramp(audio, TX_SAMPLE_RATE, ramp_ms=5.0)
            
            # Apply TX level from settings (scales the audio)
            # Use soft limiting instead of hard clipping to preserve waveform shape
            tx_level = tx_level_pct / 100.0  # Convert to 0.0-1.0
            audio = audio * tx_level
            
            # Soft limiter using tanh (preserves waveform shape, no hard clipping!)
            # This prevents distortion while still limiting peaks
            peak_before = float(np.abs(audio).max())
            if peak_before > 0.9:
                # Apply soft compression only if needed
                audio = np.tanh(audio * 1.5) * 0.9  # Soft limit to ~0.9
            
            # Log peak/RMS
            peak = float(np.abs(audio).max())
            rms = float(np.sqrt(np.mean(audio**2)))
            self.preset_log.append(f"   Audio: {len(audio)} samples ({len(audio)/TX_SAMPLE_RATE:.2f}s)")
            self.preset_log.append(f"   Peak: {peak:.3f}, RMS: {rms:.3f}, Level: {tx_level_pct}%")
            
            # Log which audio device we're using
            tx_device_name = self.settings_tx_audio_combo.currentText() if hasattr(self, 'settings_tx_audio_combo') else "Unknown"
            self.preset_log.append(f"   TX Audio Device: {tx_device_name}")
            
            # Deterministic TX timing
            TX_LEAD_IN_MS = 700   # Time before audio (let radio settle)
            TX_TAIL_MS = 200      # Time after audio (let last bits transmit)
            
            # Validate TX device before keying PTT
            try:
                device_info = sd.query_devices(tx_device)
                max_out_channels = device_info.get('max_output_channels', 0)
                device_samplerate = int(device_info.get('default_samplerate', 48000))
                
                if max_out_channels < 1:
                    raise ValueError(f"Device {tx_device} has no output channels!")
                
                self.preset_log.append(f"   Device: {device_info['name']}")
                self.preset_log.append(f"   Channels: {max_out_channels}, Native SR: {device_samplerate}")
                
                # Resample if device sample rate differs from TX_SAMPLE_RATE
                playback_rate = TX_SAMPLE_RATE
                if device_samplerate != TX_SAMPLE_RATE:
                    from scipy import signal as scipy_signal
                    # Resample to device's native rate
                    num_samples = int(len(audio) * device_samplerate / TX_SAMPLE_RATE)
                    audio = scipy_signal.resample(audio, num_samples).astype(np.float32)
                    playback_rate = device_samplerate
                    self.preset_log.append(f"   Resampled: {TX_SAMPLE_RATE} → {device_samplerate} Hz")
                
                # Reshape audio for stereo if device requires 2 channels
                if max_out_channels >= 2:
                    # Duplicate mono to stereo
                    audio_out = np.column_stack([audio, audio]).astype(np.float32)
                else:
                    audio_out = audio.astype(np.float32)
                    
            except Exception as e:
                raise ValueError(f"TX audio device error: {e}. Try refreshing device list.")
            
            # Key PTT before audio
            ptt_line = "DTR" if self.ptt_dtr_combo.currentIndex() > 0 else "RTS"
            self.preset_log.append(f"   🔴 PTT ON ({ptt_line} on {self.ptt_serial.port})")
            self._set_ptt(True)
            time.sleep(TX_LEAD_IN_MS / 1000.0)
            self.preset_log.append(f"   Lead-in: {TX_LEAD_IN_MS}ms")
            
            # Play audio at correct sample rate (native or resampled)
            self.preset_log.append(f"   🔊 Sending audio at {playback_rate} Hz...")
            sd.play(audio_out, playback_rate, device=tx_device)
            sd.wait()  # Wait for playback to finish
            
            # TX tail - keep PTT on after audio finishes
            time.sleep(TX_TAIL_MS / 1000.0)
            self.preset_log.append(f"   Tail: {TX_TAIL_MS}ms")
            # Unkey PTT
            self._set_ptt(False)
            self.preset_log.append(f"   ⚪ PTT OFF")
            
            self.preset_log.append("✅ Beacon transmitted!")
            self._log("✅ Beacon TX complete")
            
            # Plot our own position on the map
            try:
                ic, ov = icon_path(sym_table, sym_code)
                if ov:
                    ic = make_overlay(ic, ov)
                try:
                    rel_path = ic.relative_to(BASE_DIR)
                    icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                except ValueError:
                    icon_url = f"http://127.0.0.1:{self.http_port}/aprs_symbols_48/primary/29.png"
                
                # Clean tooltip - just radio and comment (no duplication)
                tooltip_parts = []
                if radio:
                    tooltip_parts.append(f"📻 {radio}")
                else:
                    tooltip_parts.append("📻 PyTNC Pro")
                if comment:
                    tooltip_parts.append(f"💬 {clean_aprs_comment(comment, 80)}")
                # Add timestamp
                tooltip_parts.append(f"🕐 {datetime.now().strftime('%H:%M:%S')}")
                tooltip = "<br>".join(tooltip_parts)
                
                # Use JSON encoding for proper escaping
                import json
                call_js = json.dumps(full_call)
                tooltip_js = json.dumps(tooltip)
                
                js = f"queueStation({call_js},{lat},{lon},'{icon_url}',{tooltip_js},false,\"\")"
                if self.map_ready:
                    self.map.page().runJavaScript(js)
            except Exception as e:
                self._log(f"  (Could not plot on map: {e})")
            
            # Auto-disconnect PTT if we auto-connected it
            if ptt_auto_connected:
                try:
                    self._set_ptt(False)
                    self.ptt_serial.close()
                    self.ptt_serial = None
                    self.preset_log.append("🔌 Auto-disconnected PTT")
                    self._update_tx_status()
                except (OSError, AttributeError):
                    pass  # PTT cleanup failed - already closed
            
        except Exception as e:
            # Make sure PTT is off on error
            try:
                self._set_ptt(False)
            except (OSError, AttributeError):
                pass  # PTT may not be connected
            
            # Auto-disconnect on error too
            if ptt_auto_connected:
                try:
                    self.ptt_serial.close()
                    self.ptt_serial = None
                except (OSError, AttributeError):
                    pass  # Already closed
            
            self.preset_log.append(f"❌ TX Error: {e}")
            self._log(f"❌ TX Error: {e}")
            QMessageBox.critical(self, "TX Error", f"Failed to transmit:\n{e}")
        
        finally:
            # Clear TX flag and set end time for holdoff
            self.tx_in_progress = False
            self.tx_end_time = time.time()
            self._update_tx_status()
    
    def _update_tx_status(self):
        """Update connection status on TX tab"""
        if not hasattr(self, 'tx_ptt_status'):
            return
        
        # PTT status
        if self._ptt_is_connected():
            self.tx_ptt_status.setText(f"🟢 PTT: {self._ptt_port_label()}")
            self.tx_ptt_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_ptt_status.setText("⚫ PTT: Not connected")
            self.tx_ptt_status.setStyleSheet("color: #ef5350;")
        
        # GPS status
        if hasattr(self, 'gps_serial') and self.gps_serial and self.gps_serial.is_open:
            self.tx_gps_status.setText("🟢 GPS: Connected")
            self.tx_gps_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_gps_status.setText("⚫ GPS: Not connected")
            self.tx_gps_status.setStyleSheet("color: #607d8b;")
        
        # TX Audio status
        if hasattr(self, 'settings_tx_audio_combo') and self.settings_tx_audio_combo.currentData() is not None:
            self.tx_audio_status.setText(f"🟢 TX: {self.settings_tx_audio_combo.currentText()[:25]}")
            self.tx_audio_status.setStyleSheet("color: #69f0ae;")
        else:
            self.tx_audio_status.setText("⚫ TX Audio: Not set")
            self.tx_audio_status.setStyleSheet("color: #607d8b;")

    def save_settings(self):
        """Save user settings to JSON file"""
        settings = {
            # Beacon settings
            "callsign": self.callsign_edit.text(),
            "ssid": self.ssid_combo.currentData(),
            "latitude": self.lat_edit.value(),
            "longitude": self.lon_edit.value(),
            "comment": self.comment_edit.text(),
            "radio": self.radio_combo.currentText(),
            "path": self.path_combo.currentText(),
            "symbol_table": self.symbol_table_combo.currentText(),
            "symbol_code": self.symbol_code_edit.text(),
            
            # Settings tab - Serial ports
            "settings_ptt_port": self.settings_ptt_combo.currentData() if hasattr(self, 'settings_ptt_combo') else None,
            "settings_gps_port": self.settings_gps_combo.currentData() if hasattr(self, 'settings_gps_combo') else None,
            "settings_gps_baud": self.gps_baud_combo.currentData() if hasattr(self, 'gps_baud_combo') else 4800,
            
            # Settings tab - PTT line settings
            "ptt_rts_mode": self.ptt_rts_combo.currentText() if hasattr(self, 'ptt_rts_combo') else "Off",
            "ptt_dtr_mode": self.ptt_dtr_combo.currentText() if hasattr(self, 'ptt_dtr_combo') else "High=TX",
            
            # Settings tab - Audio
            "settings_rx_audio": self.settings_rx_audio_combo.currentData() if hasattr(self, 'settings_rx_audio_combo') else None,
            "settings_tx_audio": self.settings_tx_audio_combo.currentData() if hasattr(self, 'settings_tx_audio_combo') else None,
            "settings_rx_gain": self.settings_rx_gain.value() if hasattr(self, 'settings_rx_gain') else 10,
            
            # Settings tab - APRS-IS
            "settings_aprs_server": self.settings_aprs_server.text() if hasattr(self, 'settings_aprs_server') else "rotate.aprs2.net",
            "settings_aprs_port": self.settings_aprs_port.value() if hasattr(self, 'settings_aprs_port') else 14580,
            "settings_aprs_radius": self.settings_aprs_radius.value() if hasattr(self, 'settings_aprs_radius') else 100,
            "settings_aprs_passcode": self.settings_aprs_passcode.text() if hasattr(self, 'settings_aprs_passcode') else "",
            
            # Settings tab - Earthquake Monitor
            "quake_enabled": self.quake_enabled.isChecked() if hasattr(self, 'quake_enabled') else False,
            "quake_radius": self.quake_radius.value() if hasattr(self, 'quake_radius') else 100,
            "quake_min_mag": self.quake_min_mag.value() if hasattr(self, 'quake_min_mag') else 2.0,
            "quake_time_range": self.quake_time_range.currentData() if hasattr(self, 'quake_time_range') else "day",
            
            # Settings tab - Fire Monitor (NASA FIRMS)
            "fire_enabled": self.fire_enabled.isChecked() if hasattr(self, 'fire_enabled') else False,
            "fire_api_key": self.fire_api_key.text() if hasattr(self, 'fire_api_key') else "",
            "fire_time_range": self.fire_time_range.currentData() if hasattr(self, 'fire_time_range') else "24h",
            "fire_source": self.fire_source.currentData() if hasattr(self, 'fire_source') else "VIIRS_SNPP_NRT",
            
            # Settings tab - AQI Monitor
            "aqi_enabled": self.aqi_enabled.isChecked() if hasattr(self, 'aqi_enabled') else False,
            "aqi_api_key": self.aqi_api_key.text() if hasattr(self, 'aqi_api_key') else "",
            
            # Settings tab - Hospitals
            "hospital_enabled": self.hospital_enabled.isChecked() if hasattr(self, 'hospital_enabled') else False,
            "hospital_radius": self.hospital_radius.value() if hasattr(self, 'hospital_radius') else 25,
            
            # RX tab layers
            "weather_enabled": self.rx_weather_check.isChecked() if hasattr(self, 'rx_weather_check') else False,
            "darn_enabled": self.rx_darn_check.isChecked() if hasattr(self, 'rx_darn_check') else False,
            
            # Settings tab - Map tile cache zoom
            "cache_map_zoom": self.cache_map_zoom_slider.value() if hasattr(self, 'cache_map_zoom_slider') else 14,
            
            # Settings tab - VARA FM
            "vara_host": self.vara_host.text() if hasattr(self, 'vara_host') else "localhost",
            "vara_cmd_port": self.vara_cmd_port.value() if hasattr(self, 'vara_cmd_port') else 8300,
            "vara_data_port": self.vara_data_port.value() if hasattr(self, 'vara_data_port') else 8301,
            "vara_port": self.vara_port.value() if hasattr(self, 'vara_port') else 8100,
            "vara_digi": self.vara_digi_edit.text() if hasattr(self, 'vara_digi_edit') else "",
            "vara_aprs_mode": self.vara_aprs_mode.currentIndex() if hasattr(self, 'vara_aprs_mode') else 0,
            
            # VARA FM tab - Beacon settings
            "vara_callsign": self.vara_callsign_edit.text() if hasattr(self, 'vara_callsign_edit') else "",
            "vara_ssid": self.vara_ssid_combo.currentData() if hasattr(self, 'vara_ssid_combo') else 9,
            "vara_lat": self.vara_lat_edit.value() if hasattr(self, 'vara_lat_edit') else 34.0522,
            "vara_lon": self.vara_lon_edit.value() if hasattr(self, 'vara_lon_edit') else -118.2437,
            "vara_comment": self.vara_comment_edit.text() if hasattr(self, 'vara_comment_edit') else "PyTNC Pro",
            "vara_radio": self.vara_radio_combo.currentText() if hasattr(self, 'vara_radio_combo') else "",
            "vara_symbol_table": getattr(self, '_vara_symbol_table', '/'),
            "vara_symbol_code": getattr(self, '_vara_symbol_code', '>'),
            
            # Settings tab - Station (use combo for SSID now)
            "settings_callsign": self.settings_callsign.text() if hasattr(self, 'settings_callsign') else "",
            "settings_ssid": self.settings_ssid_combo.currentData() if hasattr(self, 'settings_ssid_combo') else 9,
            "settings_comment": self.settings_comment.text() if hasattr(self, 'settings_comment') else "PyTNC Pro",
            
            # Settings tab - TX Level
            "settings_tx_level": self.settings_tx_level.value() if hasattr(self, 'settings_tx_level') else 100,
            
            # Settings tab - Startup options
            "auto_connect_gps": self.auto_connect_gps.isChecked() if hasattr(self, 'auto_connect_gps') else False,
            "auto_connect_aprs": self.auto_connect_aprs.isChecked() if hasattr(self, 'auto_connect_aprs') else False,
            
            # Settings tab - Manual location
            "manual_location": self.manual_location.text() if hasattr(self, 'manual_location') else "",
            
            # Winlink settings
            "wl_gateway": self.wl_gateway_edit.text() if hasattr(self, 'wl_gateway_edit') else "",
            
            # Legacy fields (for compatibility)
            "ptt_port": self.ptt_port_combo.currentData(),
            "audio_device": self.dev_combo.currentData(),
            "gain": self.gain.value(),
            "tx_device": self.tx_audio_combo.currentData() if hasattr(self, 'tx_audio_combo') else None,
            "tx_level": self.settings_tx_level.value() if hasattr(self, 'settings_tx_level') else 100,
            "aprs_is_server": self.aprs_is_server.text(),
            "aprs_is_port": self.aprs_is_port.value(),
            "aprs_is_filter": self.aprs_is_filter.text(),
            
            # Auto-beacon settings
            "auto_beacon_enabled": self.auto_beacon_enabled.isChecked() if hasattr(self, 'auto_beacon_enabled') else False,
            "auto_beacon_interval": self.auto_beacon_interval.value() if hasattr(self, 'auto_beacon_interval') else 10,
            "auto_beacon_mode": self.auto_beacon_mode.currentData() if hasattr(self, 'auto_beacon_mode') else "is",
        }
        try:
            with open(SETTINGS_FILE, "w") as f:
                json.dump(settings, f, indent=2)
            self._log(f"💾 Settings saved to {SETTINGS_FILE.name}")
            self.preset_log.append(f"💾 Settings saved!")
            if hasattr(self, 'vara_log'):
                self.vara_log.append(f"💾 Settings saved!")
            self._log("✅ Settings saved")
            # Flash save buttons green with confirmation text
            for btn_name in ('save_settings_btn',):
                btn = getattr(self, btn_name, None)
                if btn:
                    btn.setText("✅ Saved!")
                    btn.setStyleSheet("""
                        QPushButton {
                            background: #2e7d32; color: #fff;
                            font-weight: bold; border-radius: 4px;
                            padding: 4px 10px; border: 1px solid #4caf50;
                        }
                    """)
                    QTimer.singleShot(1500, lambda b=btn: (
                        b.setText("💾 Save"),
                        b.setStyleSheet("""
                            QPushButton {
                                background: #1565c0; color: #fff;
                                font-weight: bold; border-radius: 4px;
                                padding: 4px 10px; border: 1px solid #1976d2;
                            }
                            QPushButton:hover { background: #1976d2; }
                        """)
                    ))
            # Also flash the Settings tab save button if it exists
            if hasattr(self, 'settings_save_btn'):
                self.settings_save_btn.setText("✅ Saved!")
                self.settings_save_btn.setStyleSheet("""
                    QPushButton { background: #2e7d32; color: #fff;
                        font-weight: bold; border-radius: 4px;
                        padding: 4px 10px; border: 1px solid #4caf50; }
                """)
                QTimer.singleShot(1500, lambda: (
                    self.settings_save_btn.setText("💾 Save Settings"),
                    self.settings_save_btn.setStyleSheet(self._button_style("#f57c00", "#ff9800"))
                ) if hasattr(self, 'settings_save_btn') else None)
        except Exception as e:
            self._log(f"⚠️ Failed to save settings: {e}")
            self.preset_log.append(f"❌ Failed to save: {e}")
            if hasattr(self, 'vara_log'):
                self.vara_log.append(f"❌ Failed to save: {e}")
            self._log(f"❌ Failed to save settings: {e}")
    
    def load_settings(self):
        """Load user settings from JSON file"""
        # Migrate old settings file if it exists
        old_settings = BASE_DIR / "pytnc_settings.json"
        if old_settings.exists() and not SETTINGS_FILE.exists():
            try:
                import shutil
                SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy(old_settings, SETTINGS_FILE)
                self._log(f"🔄 Migrated settings to {SETTINGS_FILE}")
            except Exception as e:
                self._log(f"⚠️ Settings migration failed: {e}")
        
        if not SETTINGS_FILE.exists():
            return
        
        try:
            with open(SETTINGS_FILE, "r") as f:
                settings = json.load(f)
            
            # Apply settings to UI
            if "callsign" in settings:
                self.callsign_edit.setText(settings["callsign"])
            if "ssid" in settings:
                idx = self.ssid_combo.findData(settings["ssid"])
                if idx >= 0:
                    self.ssid_combo.setCurrentIndex(idx)
            if "latitude" in settings:
                self.lat_edit.setValue(settings["latitude"])
            if "longitude" in settings:
                self.lon_edit.setValue(settings["longitude"])
            # Center map on saved position once the WebView is ready
            if "latitude" in settings and "longitude" in settings:
                _lat = float(settings["latitude"])
                _lon = float(settings["longitude"])
                def _center_map_on_load(lat=_lat, lon=_lon):
                    try:
                        self.map.page().runJavaScript(
                            f"if(typeof setCenter === 'function') setCenter({lat}, {lon}, 12);"
                        )
                    except Exception:
                        pass
                from PyQt6.QtCore import QTimer
                QTimer.singleShot(2000, _center_map_on_load)
            if "comment" in settings:
                self.comment_edit.setText(settings["comment"])
            if "radio" in settings:
                idx = self.radio_combo.findText(settings["radio"])
                if idx >= 0:
                    self.radio_combo.setCurrentIndex(idx)
                else:
                    self.radio_combo.setEditText(settings["radio"])
            if "path" in settings:
                path_text = settings["path"]
                idx = self.path_combo.findText(path_text)
                if idx >= 0:
                    self.path_combo.setCurrentIndex(idx)
                else:
                    # Custom path not in list - add it and select it
                    self.path_combo.addItem(path_text)
                    self.path_combo.setCurrentText(path_text)
            if "symbol_table" in settings:
                idx = self.symbol_table_combo.findText(settings["symbol_table"])
                if idx >= 0:
                    self.symbol_table_combo.setCurrentIndex(idx)
            if "symbol_code" in settings:
                self.symbol_code_edit.setText(settings["symbol_code"])
                # Update symbol preview
                table = settings.get("symbol_table", "/")
                code = settings["symbol_code"]
                if code:
                    # Update symbol preview in beacon settings
                    prefix = "primary" if table == "/" else "secondary"
                    hessu_num = ord(code) - 33
                    sym_icon_path = HESSU_SYMBOLS_DIR / prefix / f"{hessu_num:02d}.png"
                    if sym_icon_path.exists():
                        self.symbol_preview.setPixmap(QPixmap(str(sym_icon_path)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            if "ptt_port" in settings and settings["ptt_port"]:
                for i in range(self.ptt_port_combo.count()):
                    if self.ptt_port_combo.itemData(i) == settings["ptt_port"]:
                        self.ptt_port_combo.setCurrentIndex(i)
                        break
            if "audio_device" in settings and settings["audio_device"] is not None:
                for i in range(self.dev_combo.count()):
                    if self.dev_combo.itemData(i) == settings["audio_device"]:
                        self.dev_combo.setCurrentIndex(i)
                        break
            if "gain" in settings:
                self.gain.setValue(settings["gain"])
            if "tx_device" in settings and settings["tx_device"] is not None:
                for i in range(self.tx_audio_combo.count()):
                    if self.tx_audio_combo.itemData(i) == settings["tx_device"]:
                        self.tx_audio_combo.setCurrentIndex(i)
                        break
            # Load TX level - prefer settings_tx_level, fallback to tx_level
            tx_level_val = settings.get("settings_tx_level", settings.get("tx_level", 100))
            if hasattr(self, 'settings_tx_level'):
                self.settings_tx_level.setValue(tx_level_val)
                self._on_settings_tx_level(tx_level_val)
            if hasattr(self, 'tx_level_slider'):
                self.tx_level_slider.setValue(tx_level_val)
            if hasattr(self, 'tx_level_label'):
                self.tx_level_label.setText(f"{tx_level_val}%")
            
            # APRS-IS settings
            if "aprs_is_server" in settings:
                self.aprs_is_server.setText(settings["aprs_is_server"])
            if "aprs_is_port" in settings:
                self.aprs_is_port.setValue(settings["aprs_is_port"])
            if "aprs_is_filter" in settings:
                self.aprs_is_filter.setText(settings["aprs_is_filter"])
            
            # Settings tab - Serial ports
            if hasattr(self, 'settings_ptt_combo') and "settings_ptt_port" in settings and settings["settings_ptt_port"]:
                for i in range(self.settings_ptt_combo.count()):
                    if self.settings_ptt_combo.itemData(i) == settings["settings_ptt_port"]:
                        self.settings_ptt_combo.setCurrentIndex(i)
                        break
            if hasattr(self, 'settings_gps_combo') and "settings_gps_port" in settings and settings["settings_gps_port"]:
                for i in range(self.settings_gps_combo.count()):
                    if self.settings_gps_combo.itemData(i) == settings["settings_gps_port"]:
                        self.settings_gps_combo.setCurrentIndex(i)
                        break
            
            # Settings tab - PTT line settings
            if hasattr(self, 'ptt_rts_combo') and "ptt_rts_mode" in settings:
                idx = self.ptt_rts_combo.findText(settings["ptt_rts_mode"])
                if idx >= 0:
                    self.ptt_rts_combo.setCurrentIndex(idx)
            if hasattr(self, 'ptt_dtr_combo') and "ptt_dtr_mode" in settings:
                idx = self.ptt_dtr_combo.findText(settings["ptt_dtr_mode"])
                if idx >= 0:
                    self.ptt_dtr_combo.setCurrentIndex(idx)
            
            # Settings tab - Earthquake Monitor
            if hasattr(self, 'quake_radius') and "quake_radius" in settings:
                self.quake_radius.setValue(settings["quake_radius"])
            if hasattr(self, 'quake_min_mag') and "quake_min_mag" in settings:
                self.quake_min_mag.setValue(settings["quake_min_mag"])
            if hasattr(self, 'quake_time_range') and "quake_time_range" in settings:
                for i in range(self.quake_time_range.count()):
                    if self.quake_time_range.itemData(i) == settings["quake_time_range"]:
                        self.quake_time_range.setCurrentIndex(i)
                        break
            if hasattr(self, 'quake_enabled') and "quake_enabled" in settings:
                # Block signals to prevent triggering before map is ready
                self.quake_enabled.blockSignals(True)
                self.quake_enabled.setChecked(settings["quake_enabled"])
                self.quake_enabled.blockSignals(False)
            
            # Settings tab - Fire Monitor (NASA FIRMS)
            if hasattr(self, 'fire_api_key') and "fire_api_key" in settings:
                self.fire_api_key.setText(settings["fire_api_key"])
            if hasattr(self, 'fire_time_range') and "fire_time_range" in settings:
                time_idx = {"24h": 0, "48h": 1, "7d": 2}.get(settings["fire_time_range"], 0)
                self.fire_time_range.setCurrentIndex(time_idx)
            if hasattr(self, 'fire_source') and "fire_source" in settings:
                source_idx = 0 if settings["fire_source"] == "VIIRS_SNPP_NRT" else 1
                self.fire_source.setCurrentIndex(source_idx)
            if hasattr(self, 'fire_enabled') and "fire_enabled" in settings:
                self.fire_enabled.blockSignals(True)
                self.fire_enabled.setChecked(settings["fire_enabled"])
                self.fire_enabled.blockSignals(False)
                # Sync RX tab checkbox
                if hasattr(self, 'rx_fire_check'):
                    self.rx_fire_check.blockSignals(True)
                    self.rx_fire_check.setChecked(settings["fire_enabled"])
                    self.rx_fire_check.blockSignals(False)
            
            # Settings tab - AQI Monitor
            if hasattr(self, 'aqi_api_key') and "aqi_api_key" in settings:
                self.aqi_api_key.setText(settings["aqi_api_key"])
            if hasattr(self, 'aqi_enabled') and "aqi_enabled" in settings:
                self.aqi_enabled.blockSignals(True)
                self.aqi_enabled.setChecked(settings["aqi_enabled"])
                self.aqi_enabled.blockSignals(False)
                # Sync RX tab checkbox
                if hasattr(self, 'rx_aqi_check'):
                    self.rx_aqi_check.blockSignals(True)
                    self.rx_aqi_check.setChecked(settings["aqi_enabled"])
                    self.rx_aqi_check.blockSignals(False)
            
            # Settings tab - Hospitals
            if hasattr(self, 'hospital_radius') and "hospital_radius" in settings:
                self.hospital_radius.setValue(settings["hospital_radius"])
            if hasattr(self, 'hospital_enabled') and "hospital_enabled" in settings:
                # Block signals to prevent triggering before map is ready
                self.hospital_enabled.blockSignals(True)
                self.hospital_enabled.setChecked(settings["hospital_enabled"])
                self.hospital_enabled.blockSignals(False)
                # Sync RX tab checkbox (also block signals)
                if hasattr(self, 'rx_hospital_check'):
                    self.rx_hospital_check.blockSignals(True)
                    self.rx_hospital_check.setChecked(settings["hospital_enabled"])
                    self.rx_hospital_check.blockSignals(False)
            
            # RX tab layers - weather
            if hasattr(self, 'rx_weather_check') and "weather_enabled" in settings:
                self.rx_weather_check.blockSignals(True)
                self.rx_weather_check.setChecked(settings["weather_enabled"])
                self.rx_weather_check.blockSignals(False)
            
            # RX tab layers - DARN
            if hasattr(self, 'rx_darn_check') and "darn_enabled" in settings:
                self.rx_darn_check.blockSignals(True)
                self.rx_darn_check.setChecked(settings["darn_enabled"])
                self.rx_darn_check.blockSignals(False)
            
            # Settings tab - Map tile cache zoom
            if hasattr(self, 'cache_map_zoom_slider') and "cache_map_zoom" in settings:
                self.cache_map_zoom_slider.setValue(settings["cache_map_zoom"])
            
            # Settings tab - Audio
            if hasattr(self, 'settings_rx_audio_combo') and "settings_rx_audio" in settings and settings["settings_rx_audio"] is not None:
                for i in range(self.settings_rx_audio_combo.count()):
                    if self.settings_rx_audio_combo.itemData(i) == settings["settings_rx_audio"]:
                        self.settings_rx_audio_combo.setCurrentIndex(i)
                        break
            if hasattr(self, 'settings_tx_audio_combo') and "settings_tx_audio" in settings and settings["settings_tx_audio"] is not None:
                for i in range(self.settings_tx_audio_combo.count()):
                    if self.settings_tx_audio_combo.itemData(i) == settings["settings_tx_audio"]:
                        self.settings_tx_audio_combo.setCurrentIndex(i)
                        break
            if hasattr(self, 'settings_rx_gain') and "settings_rx_gain" in settings:
                self.settings_rx_gain.setValue(settings["settings_rx_gain"])
                self._on_settings_rx_gain(settings["settings_rx_gain"])
            
            # Settings tab - APRS-IS
            if hasattr(self, 'settings_aprs_server') and "settings_aprs_server" in settings:
                self.settings_aprs_server.setText(settings["settings_aprs_server"])
            if hasattr(self, 'settings_aprs_port') and "settings_aprs_port" in settings:
                self.settings_aprs_port.setValue(settings["settings_aprs_port"])
            if hasattr(self, 'settings_aprs_radius') and "settings_aprs_radius" in settings:
                self.settings_aprs_radius.setValue(settings["settings_aprs_radius"])
            if hasattr(self, 'settings_aprs_passcode') and "settings_aprs_passcode" in settings:
                self.settings_aprs_passcode.setText(settings["settings_aprs_passcode"])
            
            # Settings tab - VARA FM
            if hasattr(self, 'vara_host') and "vara_host" in settings:
                self.vara_host.setText(settings["vara_host"])
            if hasattr(self, 'vara_cmd_port') and "vara_cmd_port" in settings:
                self.vara_cmd_port.setValue(settings["vara_cmd_port"])
            if hasattr(self, 'vara_data_port') and "vara_data_port" in settings:
                self.vara_data_port.setValue(settings["vara_data_port"])
            if hasattr(self, 'vara_port') and "vara_port" in settings:
                self.vara_port.setValue(settings["vara_port"])
            if hasattr(self, 'vara_digi_edit') and "vara_digi" in settings:
                self.vara_digi_edit.setText(settings["vara_digi"])
            if hasattr(self, 'vara_aprs_mode') and "vara_aprs_mode" in settings:
                self.vara_aprs_mode.setCurrentIndex(settings["vara_aprs_mode"])
            
            # VARA FM tab - Beacon settings
            if hasattr(self, 'vara_callsign_edit') and "vara_callsign" in settings:
                self.vara_callsign_edit.setText(settings["vara_callsign"])
            if hasattr(self, 'vara_ssid_combo') and "vara_ssid" in settings:
                idx = self.vara_ssid_combo.findData(settings["vara_ssid"])
                if idx >= 0:
                    self.vara_ssid_combo.setCurrentIndex(idx)
            if hasattr(self, 'vara_lat_edit') and "vara_lat" in settings:
                self.vara_lat_edit.setValue(settings["vara_lat"])
            if hasattr(self, 'vara_lon_edit') and "vara_lon" in settings:
                self.vara_lon_edit.setValue(settings["vara_lon"])
            if hasattr(self, 'vara_comment_edit') and "vara_comment" in settings:
                self.vara_comment_edit.setText(settings["vara_comment"])
            if hasattr(self, 'vara_radio_combo') and "vara_radio" in settings:
                idx = self.vara_radio_combo.findText(settings["vara_radio"])
                if idx >= 0:
                    self.vara_radio_combo.setCurrentIndex(idx)
                else:
                    self.vara_radio_combo.setEditText(settings["vara_radio"])
            # VARA FM symbol
            if "vara_symbol_table" in settings:
                self._vara_symbol_table = settings["vara_symbol_table"]
            if "vara_symbol_code" in settings:
                self._vara_symbol_code = settings["vara_symbol_code"]
                # Update symbol preview
                if hasattr(self, 'vara_symbol_preview'):
                    prefix = "primary" if self._vara_symbol_table == "/" else "secondary"
                    hessu_num = ord(self._vara_symbol_code) - 33
                    sym_icon_path = HESSU_SYMBOLS_DIR / prefix / f"{hessu_num:02d}.png"
                    if sym_icon_path.exists():
                        self.vara_symbol_preview.setPixmap(QPixmap(str(sym_icon_path)).scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            
            # Settings tab - Station (now using combo for SSID)
            if hasattr(self, 'settings_callsign') and "settings_callsign" in settings:
                self.settings_callsign.setText(settings["settings_callsign"])
            if hasattr(self, 'settings_ssid_combo') and "settings_ssid" in settings:
                idx = self.settings_ssid_combo.findData(settings["settings_ssid"])
                if idx >= 0:
                    self.settings_ssid_combo.setCurrentIndex(idx)
            if hasattr(self, 'settings_comment') and "settings_comment" in settings:
                self.settings_comment.setText(settings["settings_comment"])
            
            # Settings tab - GPS baud rate
            if hasattr(self, 'gps_baud_combo') and "settings_gps_baud" in settings:
                idx = self.gps_baud_combo.findData(settings["settings_gps_baud"])
                if idx >= 0:
                    self.gps_baud_combo.setCurrentIndex(idx)
            
            # Settings tab - Startup options
            if hasattr(self, 'auto_connect_gps') and "auto_connect_gps" in settings:
                self.auto_connect_gps.setChecked(settings["auto_connect_gps"])
            if hasattr(self, 'auto_connect_aprs') and "auto_connect_aprs" in settings:
                self.auto_connect_aprs.setChecked(settings["auto_connect_aprs"])
            
            # Settings tab - Manual location
            if hasattr(self, 'manual_location') and "manual_location" in settings:
                self.manual_location.setText(settings["manual_location"])
            
            # Winlink settings
            if hasattr(self, 'wl_gateway_edit') and "wl_gateway" in settings:
                self.wl_gateway_edit.setText(settings["wl_gateway"])
            
            # Auto-beacon settings
            if hasattr(self, 'auto_beacon_interval') and "auto_beacon_interval" in settings:
                self.auto_beacon_interval.setValue(settings["auto_beacon_interval"])
            if hasattr(self, 'auto_beacon_mode') and "auto_beacon_mode" in settings:
                idx = self.auto_beacon_mode.findData(settings["auto_beacon_mode"])
                if idx >= 0:
                    self.auto_beacon_mode.setCurrentIndex(idx)
            # Note: Don't auto-enable auto-beacon on startup - user should manually enable
            # if hasattr(self, 'auto_beacon_enabled') and "auto_beacon_enabled" in settings:
            #     self.auto_beacon_enabled.setChecked(settings["auto_beacon_enabled"])
            
            self._log(f"📂 Loaded settings from {SETTINGS_FILE.name}")
            
            # Auto-connect after settings loaded (use timer to let UI finish)
            QTimer.singleShot(500, self._auto_connect_startup)
            
            # Sync APRS tab status with current connections
            QTimer.singleShot(600, self._sync_aprs_tab_status)
            
        except Exception as e:
            self._log(f"⚠️ Failed to load settings: {e}")
    
    def _sync_aprs_tab_status(self):
        """Sync APRS AX25 tab connection status with actual state"""
        # PTT status
        if hasattr(self, 'tx_ptt_status'):
            if self._ptt_is_connected():
                self.tx_ptt_status.setText(f"🟢 PTT: {self._ptt_port_label()}")
                self.tx_ptt_status.setStyleSheet("color: #69f0ae;")
            else:
                self.tx_ptt_status.setText("⚫ PTT: Not connected")
                self.tx_ptt_status.setStyleSheet("color: #607d8b;")
        
        # GPS status
        if hasattr(self, 'tx_gps_status'):
            if hasattr(self, 'gps_serial') and self.gps_serial and self.gps_serial.is_open:
                self.tx_gps_status.setText("🟢 GPS: Connected")
                self.tx_gps_status.setStyleSheet("color: #69f0ae;")
            else:
                self.tx_gps_status.setText("⚫ GPS: Not connected")
                self.tx_gps_status.setStyleSheet("color: #607d8b;")
        
        # TX Audio status
        self._on_tx_audio_changed()
    
    def _auto_connect_startup(self):
        """Auto-connect GPS and APRS-IS on startup if enabled"""
        # Auto-connect GPS
        if hasattr(self, 'auto_connect_gps') and self.auto_connect_gps.isChecked():
            gps_port = self.settings_gps_combo.currentData()
            if gps_port and not (hasattr(self, 'gps_serial') and self.gps_serial and self.gps_serial.is_open):
                self._log("🚀 Auto-connecting GPS...")
                self._toggle_gps()
        
        # Auto-connect APRS-IS (slight delay to let GPS connect first)
        if hasattr(self, 'auto_connect_aprs') and self.auto_connect_aprs.isChecked():
            callsign = self.callsign_edit.text().strip().upper()
            if callsign and callsign != "N0CALL":
                QTimer.singleShot(1000, self._auto_connect_aprs_is)
            else:
                self._log("⚠️ Set callsign to enable APRS-IS auto-connect")
        
        # Sync connection status to Beacon tab
        QTimer.singleShot(1500, self._sync_beacon_connection_status)
    
    def closeEvent(self, event):
        """Save settings when closing the application"""
        try:
            self.save_settings()
        except Exception as e:
            print(f"Error saving on close: {e}")
        
        # Clean up APRS-IS connection
        try:
            self.aprs_is_running = False
            if self.aprs_is_socket:
                self.aprs_is_socket.close()
        except Exception as e:
            print(f"Error closing APRS-IS: {e}")
        
        # Clean up GPS connection
        try:
            self.gps_running = False
            if self.gps_timer:
                self.gps_timer.stop()
            if hasattr(self, 'gps_serial') and self.gps_serial and self.gps_serial.is_open:
                self.gps_serial.close()
        except Exception as e:
            print(f"Error closing GPS: {e}")
        
        # Clean up PTT connection first (ensure PTT is off!)
        try:
            if self.ptt_serial and self.ptt_serial.is_open:
                self.ptt_serial.rts = False
                self.ptt_serial.dtr = False
                self.ptt_serial.close()
        except Exception as e:
            print(f"Error closing PTT: {e}")
        try:
            if self.civ_serial and self.civ_serial.is_open:
                self._set_ptt(False)
                self.civ_serial.close()
        except Exception as e:
            print(f"Error closing CI-V: {e}")
        try:
            if self.cm108_device is not None:
                self._cm108_set_gpio(False)
                self.cm108_device.close()
                self.cm108_device = None
        except Exception as e:
            print(f"Error closing CM108: {e}")
        
        # Stop receiver if running
        try:
            if self.receiver:
                self.receiver.stop()
        except Exception as e:
            print(f"Error stopping receiver: {e}")
        
        event.accept()

    def _map_loaded(self, ok):
        self._log(f"Map load: {ok}")
        if not ok:
            self.map_lbl.setText("Map: FAILED")
            self.map_lbl.setStyleSheet("color:#f44")
            return
        self.map_checks = 0
        QTimer.singleShot(300, self._check_map)
        # Pre-warm tile cache in background
        QTimer.singleShot(1000, self._prewarm_tile_cache)
    
    def _prewarm_tile_cache(self):
        """Pre-load nearby tiles into memory cache to prevent stalls."""
        import threading
        def prewarm():
            # Load tiles around LA (default center) for common zoom levels
            # This prevents disk I/O stalls when user first interacts with map
            center_lat, center_lon = 34.05, -118.25
            loaded = 0
            for z in [10, 11, 12]:  # Common zoom levels
                # Calculate tile coords for center
                import math
                n = 2 ** z
                x_center = int((center_lon + 180) / 360 * n)
                y_center = int((1 - math.log(math.tan(math.radians(center_lat)) + 1/math.cos(math.radians(center_lat))) / math.pi) / 2 * n)
                
                # Load 5x5 grid around center
                for dx in range(-2, 3):
                    for dy in range(-2, 3):
                        x, y = x_center + dx, y_center + dy
                        tile_key = f"{z}/{x}/{y}"
                        if tile_key in _tile_memory_cache:
                            continue
                        tile_path = TILE_CACHE_DIR / str(z) / str(x) / f"{y}.png"
                        if tile_path.exists():
                            try:
                                with open(tile_path, 'rb') as f:
                                    _tile_memory_cache[tile_key] = f.read()
                                    loaded += 1
                            except:
                                pass
            if loaded > 0:
                print(f"[TILE] Pre-warmed {loaded} tiles into memory cache")
        
        threading.Thread(target=prewarm, daemon=True).start()
    
    def _refresh_map(self):
        """Refresh map tiles"""
        if self.map_ready:
            self.map.page().runJavaScript("refreshMap()")
            self._log("🔄 Map refreshed")

    def _on_map_layer_changed(self, index):
        """Switch map tile layer from Qt combo"""
        if not self.map_ready:
            return
        layer = self.map_layer_combo.currentData()
        if layer:
            self.map.page().runJavaScript(f"setMapLayer({layer!r})")
    
    def _load_locations_menu(self):
        """Show menu to choose file or folder loading"""
        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: #1e3a5f; color: white; border: 1px solid #2d5a87; }
            QMenu::item:selected { background: #2d5a87; }
        """)
        
        file_action = menu.addAction("📄 Load File(s)")
        folder_action = menu.addAction("📁 Load Folder")
        
        action = menu.exec(self.load_locations_btn.mapToGlobal(
            self.load_locations_btn.rect().bottomLeft()))
        
        if action == file_action:
            self._load_locations_files()
        elif action == folder_action:
            self._load_locations_folder()
    
    def _load_locations_files(self):
        """Load locations from multiple CSV/XLS files"""
        filenames, _ = QFileDialog.getOpenFileNames(
            self, "Load Location Files", str(BASE_DIR),
            "CSV/Excel Files (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not filenames:
            return
        
        total_loaded = 0
        for filename in filenames:
            count = self._load_single_location_file(Path(filename))
            total_loaded += count
        
        if total_loaded > 0:
            self._log(f"📍 Loaded {total_loaded} locations from {len(filenames)} files (total: {len(self.custom_locations)})")
            self.location_count_lbl.setText(f"({len(self.custom_locations)})")
            self.beacon_locations_btn.show()
            self.clear_locations_btn.show()
            self._display_locations_on_map()
    
    def _load_locations_folder(self):
        """Load all CSV/XLS files from a folder"""
        folder = QFileDialog.getExistingDirectory(
            self, "Select Locations Folder", str(BASE_DIR)
        )
        if not folder:
            return
        
        folder_path = Path(folder)
        files = list(folder_path.glob("*.csv")) + list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))
        
        if not files:
            QMessageBox.warning(self, "No Files", f"No CSV or Excel files found in:\n{folder}")
            return
        
        total_loaded = 0
        for filepath in sorted(files):
            count = self._load_single_location_file(filepath)
            total_loaded += count
        
        if total_loaded > 0:
            self._log(f"📍 Loaded {total_loaded} locations from {len(files)} files (total: {len(self.custom_locations)})")
            self.location_count_lbl.setText(f"({len(self.custom_locations)})")
            self.beacon_locations_btn.show()
            self.clear_locations_btn.show()
            self._display_locations_on_map()
        
        QMessageBox.information(self, "Folder Loaded",
            f"Loaded {total_loaded} locations from {len(files)} files.")
    
    def _load_single_location_file(self, filepath: Path) -> int:
        """Load locations from a single file. Returns count loaded."""
        try:
            locations = []
            
            if filepath.suffix.lower() == '.csv':
                import csv
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        name = row.get('Name', '').strip()
                        lat_str = row.get('LAT', row.get('Lat', row.get('lat', ''))).strip()
                        lon_str = row.get('Long', row.get('Lon', row.get('lon', ''))).strip()
                        
                        if not name or not lat_str or not lon_str:
                            continue
                        
                        try:
                            lat = float(lat_str)
                            lon = float(lon_str)
                        except ValueError:
                            continue
                        
                        locations.append({
                            'name': name,
                            'address': row.get('Address', '').strip(),
                            'lat': lat,
                            'lon': lon,
                            'symbol': row.get('Symbol', '\\h').strip(),
                            'comment': row.get('Comment', '').strip(),
                            'source': filepath.name
                        })
            
            elif filepath.suffix.lower() in ('.xlsx', '.xls'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(filepath)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        name = str(row_dict.get('Name', '')).strip()
                        lat_str = str(row_dict.get('LAT', row_dict.get('Lat', ''))).strip()
                        lon_str = str(row_dict.get('Long', row_dict.get('Lon', ''))).strip()
                        
                        if not name or not lat_str or not lon_str:
                            continue
                        
                        try:
                            lat = float(lat_str)
                            lon = float(lon_str)
                        except ValueError:
                            continue
                        
                        locations.append({
                            'name': name,
                            'address': str(row_dict.get('Address', '')).strip(),
                            'lat': lat,
                            'lon': lon,
                            'symbol': str(row_dict.get('Symbol', '\\h')).strip(),
                            'comment': str(row_dict.get('Comment', '')).strip(),
                            'source': filepath.name
                        })
                except ImportError:
                    self._log(f"⚠️ Skipped {filepath.name} - openpyxl not installed")
                    return 0
            
            if locations:
                # Deduplicate: don't add if same name+lat+lon already exists
                existing_keys = {(loc['name'], round(loc['lat'], 5), round(loc['lon'], 5)) 
                                 for loc in self.custom_locations}
                new_locations = []
                for loc in locations:
                    key = (loc['name'], round(loc['lat'], 5), round(loc['lon'], 5))
                    if key not in existing_keys:
                        new_locations.append(loc)
                        existing_keys.add(key)
                
                if new_locations:
                    self.custom_locations.extend(new_locations)
                    self._log(f"  📄 {filepath.name}: {len(new_locations)} new locations")
                    return len(new_locations)
                else:
                    self._log(f"  ⚠️ {filepath.name}: all {len(locations)} locations already loaded")
                    return 0
            
            return 0
            
        except Exception as e:
            self._log(f"⚠️ Error loading {filepath.name}: {e}")
            return 0

    def _display_locations_on_map(self):
        """Display custom locations on the map"""
        if not self.map_ready or not self.custom_locations:
            return
        
        # Clear existing custom markers first
        self.map.page().runJavaScript("clearCustomLocations()")
        
        for loc in self.custom_locations:
            name = loc['name']
            lat = loc['lat']
            lon = loc['lon']
            symbol = loc.get('symbol', '\\h')
            comment = loc.get('comment', '')
            address = loc.get('address', '')
            
            # Escape for JavaScript
            name_js = json.dumps(name)
            comment_js = json.dumps(comment)
            address_js = json.dumps(address)
            symbol_js = json.dumps(symbol)
            
            js = f"addCustomLocation({name_js}, {lat}, {lon}, {symbol_js}, {comment_js}, {address_js})"
            self.map.page().runJavaScript(js)
        
        self._log(f"📍 Displayed {len(self.custom_locations)} locations on map")
    
    def _clear_locations(self):
        """Clear all loaded locations"""
        if not self.custom_locations:
            return
        
        count = len(self.custom_locations)
        self.custom_locations = []
        
        # Clear map markers
        if self.map_ready:
            self.map.page().runJavaScript("clearCustomLocations()")
        
        # Hide buttons and clear label
        self.beacon_locations_btn.hide()
        self.clear_locations_btn.hide()
        self.location_count_lbl.setText("")
        
        self._log(f"🗑️ Cleared {count} locations")
    
    def _beacon_locations_menu(self):
        """Show menu to choose RF or APRS-IS beacon"""
        if not self.custom_locations:
            self._log("❌ No locations to beacon")
            return
        
        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: #1e3a5f; color: white; border: 1px solid #2d5a87; }
            QMenu::item:selected { background: #2d5a87; }
        """)
        
        rf_action = menu.addAction("📻 Beacon via RF (Simplex)")
        is_action = menu.addAction("🌐 Beacon via APRS-IS")
        menu.addSeparator()
        cancel_action = menu.addAction("Cancel")
        
        action = menu.exec(self.beacon_locations_btn.mapToGlobal(
            self.beacon_locations_btn.rect().bottomLeft()))
        
        if action == rf_action:
            self._beacon_locations_rf()
        elif action == is_action:
            self._beacon_all_locations()
    
    def _beacon_locations_rf(self):
        """Beacon all locations as APRS objects via RF"""
        if not self.custom_locations:
            self._log("❌ No locations to beacon")
            return
        
        if not HAS_SOUNDDEVICE:
            QMessageBox.warning(self, "RF Disabled", 
                "sounddevice not installed.\nRF AFSK transmit is disabled.")
            return
        
        callsign = self.callsign_edit.text().strip().upper()
        if not callsign or callsign == "N0CALL":
            QMessageBox.warning(self, "No Callsign", "Set your callsign first")
            return
        
        ssid = self.ssid_combo.currentData()
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        path_str = self.path_combo.currentText().strip()
        
        # Parse path
        path_list = []
        if path_str and path_str.upper() != "DIRECT":
            for p in path_str.split(","):
                p = p.strip()
                if "-" in p:
                    pcall, pssid = p.rsplit("-", 1)
                    path_list.append((pcall, int(pssid)))
                else:
                    path_list.append((p, 0))
        
        # Check PTT
        if not self._ptt_is_connected():
            method = getattr(self, 'civ_ptt_method', 'RTS/DTR')
            if method == "CI-V CAT":
                ptt_port = self.civ_port_combo.currentData() if hasattr(self, 'civ_port_combo') else None
                if ptt_port:
                    try:
                        self._toggle_civ()
                        self._log(f"✅ Auto-connected CI-V PTT: {ptt_port}")
                    except Exception as e:
                        QMessageBox.warning(self, "PTT Error", f"Could not connect CI-V PTT:\n{e}")
                        return
                else:
                    QMessageBox.warning(self, "PTT Not Configured", "Configure CI-V PTT in Settings")
                    return
            else:
                ptt_port = self.settings_ptt_combo.currentData() if hasattr(self, 'settings_ptt_combo') else None
                if ptt_port:
                    try:
                        self.ptt_serial = serial.Serial(ptt_port, 9600, timeout=0.1)
                        self._set_ptt(False)
                        self._log(f"✅ Auto-connected PTT: {ptt_port}")
                    except Exception as e:
                        QMessageBox.warning(self, "PTT Error", f"Could not connect PTT:\n{e}")
                        return
                else:
                    QMessageBox.warning(self, "PTT Not Configured", "Configure PTT in Settings")
                    return
        
        # Get TX audio device
        tx_device = self.settings_tx_audio_combo.currentData() if hasattr(self, 'settings_tx_audio_combo') else None
        if tx_device is None:
            QMessageBox.warning(self, "No TX Audio", "Select TX audio device in Settings")
            return
        
        tx_level_pct = self.settings_tx_level.value() if hasattr(self, 'settings_tx_level') else 10
        
        # Timestamp
        from datetime import datetime
        now = datetime.utcnow()
        timestamp = now.strftime("%d%H%M") + "z"
        
        # Confirm
        reply = QMessageBox.question(self, "Beacon Locations via RF",
            f"Beacon {len(self.custom_locations)} locations via RF?\n\n"
            f"Path: {path_str}\n"
            f"This will take ~{len(self.custom_locations) * 3} seconds.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        self._log(f"📻 Starting RF beacon of {len(self.custom_locations)} locations...")
        self.tx_in_progress = True
        sent = 0
        
        try:
            for i, loc in enumerate(self.custom_locations):
                name = loc['name'][:9].ljust(9)
                lat = loc['lat']
                lon = loc['lon']
                symbol = loc.get('symbol', '\\h')
                address = loc.get('address', '')
                comment = loc.get('comment', '')
                # Combine address and comment (address first)
                full_comment = f"{address} {comment}".strip()[:43]
                
                # Parse symbol
                sym_table = symbol[0] if len(symbol) >= 1 else '\\'
                sym_code = symbol[1] if len(symbol) >= 2 else 'h'
                
                # Format position
                lat_deg = int(abs(lat))
                lat_min = (abs(lat) - lat_deg) * 60
                lat_dir = "N" if lat >= 0 else "S"
                lon_deg = int(abs(lon))
                lon_min = (abs(lon) - lon_deg) * 60
                lon_dir = "E" if lon >= 0 else "W"
                
                # APRS Object format
                info = f";{name}*{timestamp}{lat_deg:02d}{lat_min:05.2f}{lat_dir}{sym_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{sym_code}{full_comment}"
                
                self._log(f"📡 TX [{i+1}/{len(self.custom_locations)}]: {name.strip()}")
                
                # Build packet
                packet_data = APRSPacketBuilder.build_ui_packet(
                    src_call=callsign, src_ssid=ssid,
                    dst_call="APPR01", dst_ssid=0,
                    path=path_list,
                    info=info
                )
                fcs = APRSPacketBuilder.compute_fcs(packet_data)
                full_packet = packet_data + bytes([fcs & 0xFF, (fcs >> 8) & 0xFF])
                
                # Generate audio
                modulator = AFSKModulator(TX_SAMPLE_RATE)
                audio = modulator.generate_packet_audio(full_packet, preamble_flags=50, postamble_flags=8)
                
                # Add silence
                silence = np.zeros(int(TX_SAMPLE_RATE * 0.03), dtype=np.float32)
                audio = np.concatenate([silence, audio, silence])
                
                # Apply level
                audio = apply_cosine_ramp(audio, TX_SAMPLE_RATE, ramp_ms=5.0)
                audio = audio * (tx_level_pct / 100.0)
                
                # Soft limit
                if float(np.abs(audio).max()) > 0.9:
                    audio = np.tanh(audio * 1.5) * 0.9
                
                # Get device info for resampling
                device_info = sd.query_devices(tx_device)
                device_sr = int(device_info.get('default_samplerate', 48000))
                max_ch = device_info.get('max_output_channels', 1)
                
                if device_sr != TX_SAMPLE_RATE:
                    from scipy import signal as scipy_signal
                    num_samples = int(len(audio) * device_sr / TX_SAMPLE_RATE)
                    audio = scipy_signal.resample(audio, num_samples).astype(np.float32)
                
                if max_ch >= 2:
                    audio_out = np.column_stack([audio, audio]).astype(np.float32)
                else:
                    audio_out = audio.astype(np.float32)
                
                # Key PTT
                self._set_ptt(True)
                time.sleep(0.5)  # Let radio settle
                
                # Play audio
                sd.play(audio_out, device_sr, device=tx_device)
                sd.wait()
                
                # Unkey PTT
                time.sleep(0.15)
                self._set_ptt(False)
                
                sent += 1
                
                # Pause between packets
                if i < len(self.custom_locations) - 1:
                    time.sleep(1.5)
                
                QApplication.processEvents()
                
        except Exception as e:
            self._log(f"❌ RF beacon error: {e}")
            QMessageBox.warning(self, "Error", f"RF beacon failed:\n{e}")
        finally:
            self.tx_in_progress = False
            self._set_ptt(False)
        
        self._log(f"📻 RF beacon complete: {sent}/{len(self.custom_locations)} sent")
        QMessageBox.information(self, "RF Beacon Complete",
            f"Sent {sent} location objects via RF.")

    def _beacon_all_locations(self):
        """Beacon all locations as APRS objects"""
        if not self.custom_locations:
            self._log("❌ No locations to beacon")
            return
        
        callsign = self.callsign_edit.text().strip().upper()
        if not callsign or callsign == "N0CALL":
            QMessageBox.warning(self, "No Callsign", "Set your callsign first")
            return
        
        # Check if APRS-IS connected
        if not (hasattr(self, 'aprs_is_socket') and self.aprs_is_socket):
            QMessageBox.warning(self, "Not Connected", 
                "Connect to APRS-IS first to beacon objects.\n\n"
                "Go to Settings → APRS-IS and connect.")
            return
        
        ssid = self.ssid_combo.currentData()
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        
        # Timestamp (DHM format for objects)
        from datetime import datetime
        now = datetime.utcnow()
        timestamp = now.strftime("%d%H%M") + "z"
        
        sent = 0
        for loc in self.custom_locations:
            try:
                name = loc['name'][:9].ljust(9)  # Object names are 9 chars
                lat = loc['lat']
                lon = loc['lon']
                symbol = loc.get('symbol', '\\h')
                address = loc.get('address', '')
                comment = loc.get('comment', '')
                # Combine address and comment (address first)
                full_comment = f"{address} {comment}".strip()[:43]
                
                # Parse symbol table/code
                if len(symbol) >= 2:
                    sym_table = symbol[0]
                    sym_code = symbol[1]
                else:
                    sym_table = '\\'
                    sym_code = 'h'
                
                # Format position
                lat_deg = int(abs(lat))
                lat_min = (abs(lat) - lat_deg) * 60
                lat_dir = "N" if lat >= 0 else "S"
                
                lon_deg = int(abs(lon))
                lon_min = (abs(lon) - lon_deg) * 60
                lon_dir = "E" if lon >= 0 else "W"
                
                # APRS Object format: ;NAME*DDHHMMz/DDMM.MMN/DDDMM.MMWsComment
                # ; = object, * = live object, _ = killed object
                info = f";{name}*{timestamp}{lat_deg:02d}{lat_min:05.2f}{lat_dir}{sym_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{sym_code}{full_comment}"
                
                packet = f"{full_call}>APPR01,TCPIP*:{info}\r\n"
                self.aprs_is_socket.send(packet.encode())
                sent += 1
                
                # Small delay between packets
                time.sleep(0.1)
                
            except Exception as e:
                self._log(f"⚠️ Failed to beacon {loc['name']}: {e}")
        
        self._log(f"📡 Beaconed {sent}/{len(self.custom_locations)} location objects")
        QMessageBox.information(self, "Beaconed", 
            f"Sent {sent} location objects via APRS-IS.\n\n"
            f"They should appear on aprs.fi within a minute.")

    def _check_map(self):
        self.map_checks += 1
        
        def handle(r):
            if r is None:
                if self.map_checks < 15:
                    QTimer.singleShot(300, self._check_map)
                else:
                    self._log("Map timeout")
                    self.map_lbl.setText("Map: TIMEOUT")
                    self.map_lbl.setStyleSheet("color:#f44")
                return
            
            self._log(f"Map: {r}")
            
            if r.get("mapError"):
                self.map_lbl.setText("Map: ERROR")
                self.map_lbl.setStyleSheet("color:#f44")
                return
            
            if r.get("mapReady"):
                self.map_ready = True
                tiles = r.get("tilesLoaded", 0)
                errs = r.get("tileErrors", 0)
                ver = r.get("leafletVersion", "?")
                gpu = r.get("gpuInfo", "unknown")
                
                # Log GPU info
                self._log(f"🖥️ GPU: {gpu}")
                
                if tiles > 0:
                    self.map_lbl.setText(f"Map: OK ({tiles} tiles)")
                    self.map_lbl.setStyleSheet("color:#4f4")
                elif errs > 0:
                    self.map_lbl.setText(f"Map: OK (tile errors: {errs})")
                    self.map_lbl.setStyleSheet("color:#fa0")
                else:
                    self.map_lbl.setText(f"Map: OK (v{ver})")
                    self.map_lbl.setStyleSheet("color:#4f4")
                
                for js in self.pending_js:
                    self.map.page().runJavaScript(js)
                self.pending_js.clear()
                
                # Apply saved layer states now that map is ready
                QTimer.singleShot(500, self._apply_saved_layers)
            elif self.map_checks < 15:
                QTimer.singleShot(300, self._check_map)
        
        self.map.page().runJavaScript("typeof getMapDiagnostics==='function'?getMapDiagnostics():null", handle)
    
    def _apply_saved_layers(self):
        """Apply saved layer states after map is ready"""
        layers = []
        
        # Hospitals
        if hasattr(self, 'rx_hospital_check') and self.rx_hospital_check.isChecked():
            layers.append("hospitals")
            self._toggle_hospital_layer(Qt.CheckState.Checked.value)
        
        # Weather/NOAA
        if hasattr(self, 'rx_weather_check') and self.rx_weather_check.isChecked():
            layers.append("weather")
            self._toggle_weather_layer(Qt.CheckState.Checked.value)
        
        # DARN repeaters
        if hasattr(self, 'rx_darn_check') and self.rx_darn_check.isChecked():
            layers.append("DARN")
            self._rx_toggle_darn(Qt.CheckState.Checked.value)
        
        # Earthquakes
        if hasattr(self, 'quake_enabled') and self.quake_enabled.isChecked():
            layers.append("earthquakes")
            self._fetch_earthquakes()
        
        # Fires (NASA FIRMS)
        if hasattr(self, 'fire_enabled') and self.fire_enabled.isChecked():
            api_key = self.fire_api_key.text().strip() if hasattr(self, 'fire_api_key') else ""
            if api_key:
                layers.append("fires")
                self._toggle_fire_monitor(Qt.CheckState.Checked.value)
        
        if layers:
            self._log(f"🗺️ Enabled layers: {', '.join(layers)}")

    def load_devices(self):
        self.dev_combo.clear()
        
        # sounddevice is optional - if not available, disable device selection
        if not HAS_SOUNDDEVICE or sd is None:
            try:
                self.dev_combo.addItem("Audio RX disabled (sounddevice unavailable)")
                self.dev_combo.setEnabled(False)
            except Exception:
                pass
            return
        
        try:
            devices = list(sd.query_devices())
            for i, d in enumerate(devices):
                if d.get("max_input_channels", 0) > 0:
                    self.dev_combo.addItem(f"{i}: {d['name']}", i)
        except Exception as e:
            try:
                self.dev_combo.addItem(f"Audio RX disabled ({type(e).__name__})")
                self.dev_combo.setEnabled(False)
            except Exception:
                pass

    def on_gain(self):
        g = self.gain.value() / 10.0
        self.gain_lbl.setText(f"{g:.1f}x")
        if self.receiver:
            self.receiver.set_gain(g)

    def start(self):
        dev = self.dev_combo.currentData()
        if dev is None:
            return
        
        self.receiver = AudioReceiver(dev, self.gain.value() / 10.0)
        self.receiver.packet_received.connect(self.on_packet)
        self.receiver.audio_level.connect(lambda v: self.meter.setValue(min(max(int(v*500),0),100)))
        self.receiver.status_update.connect(self._log)
        self.receiver.error_occurred.connect(self._log)
        self.receiver.start()
        
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.status_lbl.setText("RECEIVING")
        self.dot.setStyleSheet("font-size:20px;color:#69f0ae")  # Bright green
        
        # Update APRS tab connection status (delayed to ensure thread starts)
        if hasattr(self, '_sync_beacon_connection_status'):
            QTimer.singleShot(100, self._sync_beacon_connection_status)

    def stop(self):
        if self.receiver:
            self.receiver.stop()
            self.receiver = None
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_lbl.setText("STOPPED")
        self.dot.setStyleSheet("font-size:20px;color:#ff6b6b")  # Soft red
        
        # Update APRS tab connection status
        if hasattr(self, '_sync_beacon_connection_status'):
            self._sync_beacon_connection_status()
        self.meter.setValue(0)

    def on_packet(self, pkt, sl):
        # Ignore packets during TX and for 2 seconds after (prevent self-decode)
        TX_HOLDOFF_SECONDS = 2.0
        if self.tx_in_progress:
            return
        if time.time() - self.tx_end_time < TX_HOLDOFF_SECONDS:
            return
            
        ts = datetime.now().strftime("%H:%M:%S")
        src, dst = str(pkt.source), str(pkt.destination)
        via = ",".join(str(r) for r in pkt.digipeaters) if pkt.digipeaters else "-"
        info = pkt.info.decode("latin-1", errors="replace")
        
        key = (src, dst, via, info)
        now = time.time()
        if now - self.dedup.get(key, 0) < 1.5:
            return
        self.dedup[key] = now

        # Track this station as RF-heard (for TX IGate eligibility)
        if hasattr(self, 'igate_rf_heard'):
            self.igate_rf_heard[src.upper()] = now

        # RX IGate: gate to APRS-IS if enabled
        if getattr(self, 'igate_rx_enabled', False):
            self._igate_log_entry(f"📻 RF heard: {src}>{dst} via {via}", "#64b5f6")
            self._gate_packet_to_is(src, dst, via, info, pkt)
        
        # Track digipeater usage from RF
        if via and via != "-":
            for digi in via.split(','):
                digi_clean = digi.strip().rstrip('*').upper()
                if digi_clean and not digi_clean.startswith(('WIDE', 'RELAY', 'TRACE')):
                    if digi_clean not in self.digi_traffic:
                        self.digi_traffic[digi_clean] = []
                    self.digi_traffic[digi_clean].append((src, ts))
                    self.digi_traffic[digi_clean] = self.digi_traffic[digi_clean][-20:]
        
        self.packets += 1
        self.pkt_lbl.setText(f"Packets: {self.packets}")
        
        # Check if it's an APRS UI frame (control=0x03, pid=0xF0)
        is_aprs = (pkt.control == 0x03 and pkt.pid == 0xF0)
        
        if is_aprs:
            aprs = aprs_classify(dst, info)
            
            # Look up device type from destination
            device = get_device_from_tocall(dst)
            
            # Simplified color scheme - header is gold, details are light blue
            # Only special packets get unique colors
            is_my_packet = src.upper().startswith(self.callsign_edit.text().strip().upper())
            
            if is_my_packet:
                header_color = "#69f0ae"  # Green for YOUR packets
            else:
                header_color = "#ff9800"  # Orange for RF packets (distinct from IS gold)
            
            # Detail line color (used for coords, comments, weather)
            detail_color = "#64b5f6"  # Light blue for all details
            
            # Header with device info if known - callsign BOLD, rest normal
            # 📻 = RF packet (no globe)
            header = f"📻 <a href='aprs://pan/{src}' style='color:{header_color};text-decoration:none;font-weight:bold'>{src}</a><span style='color:{header_color}'>&gt;{dst} via {via}</span>"
            if device:
                header += f" <span style='color:{header_color}'>[{device}]</span>"
            self._log(header)
            
            # Handle telemetry definitions - store them
            if aprs["kind"] in ("Telem-PARM", "Telem-UNIT", "Telem-EQNS", "Telem-BITS"):
                f = aprs["fields"]
                target = f.get("for", src)
                if target not in self.telem_defs:
                    self.telem_defs[target] = {}
                
                if aprs["kind"] == "Telem-PARM":
                    self.telem_defs[target]["parm"] = f.get("params", [])
                elif aprs["kind"] == "Telem-UNIT":
                    self.telem_defs[target]["unit"] = f.get("units", [])
                elif aprs["kind"] == "Telem-EQNS":
                    self.telem_defs[target]["eqns"] = f.get("eqns", [])
                elif aprs["kind"] == "Telem-BITS":
                    self.telem_defs[target]["bits"] = f.get("bits", [])
                
                self._log(f"  {aprs['kind']}: {aprs['summary']}", detail_color)
                self._log(f"  (Stored for {target})", detail_color)
                return
            
            # Handle regular telemetry - apply coefficients if available
            if aprs["kind"] == "Telemetry":
                f = aprs["fields"]
                tdef = self.telem_defs.get(src, {})
                
                analogs = f.get("analog", [])
                digitals = f.get("digital", [])
                seq = f.get("sequence", 0)
                
                eqns = tdef.get("eqns", [])
                units = tdef.get("unit", [])
                parms = tdef.get("parm", [])
                
                # Build formatted telemetry string
                telem_parts = [f"Seq={seq}"]
                for i, raw in enumerate(analogs):
                    # Apply equation if available: value = a*x² + b*x + c
                    if i < len(eqns):
                        a, b, c = eqns[i]
                        val = a * raw * raw + b * raw + c
                    else:
                        val = raw
                    
                    # Get parameter name and unit
                    pname = parms[i] if i < len(parms) else f"A{i+1}"
                    unit = units[i] if i < len(units) else ""
                    
                    if val == int(val):
                        telem_parts.append(f"{pname}={int(val)}{unit}")
                    else:
                        telem_parts.append(f"{pname}={val:.2f}{unit}")
                
                # Digital bits
                if digitals:
                    for i, bit in enumerate(digitals):
                        dname = f"D{i+1}"
                        telem_parts.append(f"{dname}={bit}")
                
                self._log(f"  Telemetry: {', '.join(telem_parts)}", detail_color)
                return
            
            # Skip third-party frames in live feed — valid APRS but noisy
            if aprs["kind"] == "Other" and "[Third-party]" in aprs.get("summary", ""):
                return

            self._log(f"  {aprs['kind']}: {aprs['summary']}", detail_color)
            
            # Handle RF messages addressed to us
            if aprs["kind"] in ("Message", "Message-ACK", "Message-REJ"):
                f = aprs["fields"]
                my_call = self.callsign_edit.text().strip().upper()
                my_ssid = self.ssid_combo.currentData()
                my_full = f"{my_call}-{my_ssid}" if my_ssid > 0 else my_call
                
                to_call = f.get("to", "").strip()
                
                # Check if addressed to us (using flexible matching)
                if callsigns_match(to_call, my_full):
                    if aprs["kind"] == "Message-ACK":
                        seq = f.get("ack", "")
                        self._handle_ack(src, seq)
                    elif aprs["kind"] == "Message-REJ":
                        self._log(f"  ❌ Message rejected by {src}")
                    else:
                        # Regular message
                        msg_text = f.get("message", "")
                        # Extract sequence number if present
                        seq = None
                        if '{' in msg_text:
                            msg_text, seq = msg_text.rsplit('{', 1)
                            seq = seq.rstrip('}').strip()
                        self._handle_incoming_message(src, to_call, msg_text.strip(), seq)
                return
            
            # Handle any packet type with position (Position, Position+Time, Mic-E, NMEA, Weather)
            if aprs["kind"] in ("Position", "Position+Time", "Mic-E", "NMEA", "Weather"):
                f = aprs["fields"]
                
                # Check if we have coordinates
                if "lat" not in f or "lon" not in f:
                    self._log(f"  (no position data)")
                    return
                
                try:
                    lat, lon = float(f["lat"]), float(f["lon"])
                except Exception as e:
                    self._log(f"  ERROR parsing coords: {e}")
                    return
                
                # Extra info display based on packet type
                if aprs["kind"] == "Mic-E":
                    # Summary already shows position/speed/course/alt
                    # Just show telemetry if present
                    if f.get("telemetry"):
                        t = f["telemetry"]
                        telem_parts = []
                        if "sequence" in t:
                            telem_parts.append(f"Seq={t['sequence']}")
                        if "analog" in t:
                            for i, v in enumerate(t["analog"]):
                                telem_parts.append(f"A{i+1}={v}")
                        if telem_parts:
                            self._log(f"  Telemetry: {', '.join(telem_parts)}")
                
                elif aprs["kind"] == "NMEA":
                    speed = f.get("speed_mph", 0) or 0
                    course = f.get("course", 0) or 0
                    if speed > 0:
                        self._log(f"  → {lat:.5f}, {lon:.5f} | {speed:.0f} mph, {course:.0f}°", "#64b5f6")
                    else:
                        self._log(f"  → {lat:.5f}, {lon:.5f}", "#64b5f6")
                
                elif aprs["kind"] == "Weather":
                    wx_info = []
                    if "temp_f" in f:
                        wx_info.append(f"{f['temp_f']}°F")
                    if "wind_speed" in f or "wind_dir" in f:
                        wind_str = "wind"
                        if "wind_speed" in f:
                            wind_str += f" {f['wind_speed']}mph"
                        if "wind_dir" in f:
                            wind_str += f" {f['wind_dir']}°"
                        wx_info.append(wind_str)
                    if "humidity" in f:
                        wx_info.append(f"{f['humidity']}% RH")
                    self._log(f"  → {lat:.5f}, {lon:.5f} | {', '.join(wx_info)}", "#64b5f6")
                else:
                    self._log(f"  → {lat:.5f}, {lon:.5f}", "#64b5f6")
                
                ic, ov = icon_path(f.get("table", "/"), f.get("sym", ">"))
                if ov:
                    ic = make_overlay(ic, ov)
                
                # Build tooltip with more info - callsign added separately as QRZ link
                tooltip_parts = []
                
                # Check if this is a digipeater
                is_digi = False
                ssid = src.split('-')[1] if '-' in src else ""
                if ssid in ['10', '11', '12', '15']:
                    is_digi = True
                if f.get('sym') == '#':
                    is_digi = True
                
                if is_digi:
                    tooltip_parts.append("📡 Digipeater")
                    # Show recent traffic through this digi
                    if src in self.digi_traffic and self.digi_traffic[src]:
                        recent = self.digi_traffic[src][-5:]  # Last 5
                        traffic_list = ", ".join([f"{s[0]}" for s in reversed(recent)])
                        tooltip_parts.append(f"📶 Recent: {traffic_list}")
                
                # Device type from tocall
                device = get_device_from_tocall(dst)
                if device:
                    tooltip_parts.append(f"📻 {device}")
                
                # Mic-E radio type (if different from tocall device)
                if aprs["kind"] == "Mic-E":
                    if f.get("radio_type") and f.get("radio_type") != device:
                        tooltip_parts.append(f"📻 {f['radio_type']}")
                
                # Speed/course - check for any packet type
                speed_mph = f.get("speed_mph") or 0
                if speed_mph > 0:
                    speed_str = f"🚗 {speed_mph:.0f} mph"
                    course = f.get("course") or 0
                    if course > 0:
                        speed_str += f" @ {course:.0f}°"
                    tooltip_parts.append(speed_str)
                
                # Altitude - check for any packet type
                altitude_ft = f.get("altitude_ft")
                if altitude_ft:
                    tooltip_parts.append(f"📍 {altitude_ft:,} ft")
                
                # Mic-E status message
                if aprs["kind"] == "Mic-E" and f.get("msg_type"):
                    tooltip_parts.append(f"[{f['msg_type']}]")
                
                # Weather specific info
                if aprs["kind"] == "Weather":
                    if "temp_f" in f:
                        tooltip_parts.append(f"🌡️ {f['temp_f']}°F")
                    if "humidity" in f:
                        tooltip_parts.append(f"💧 {f['humidity']}%")
                    if "wind_speed" in f or "wind_dir" in f:
                        wind_str = "💨"
                        if "wind_dir" in f:
                            wind_str += f" {f['wind_dir']}°"
                        if "wind_speed" in f:
                            wind_str += f" {f['wind_speed']} mph"
                        if f.get("wind_gust", 0) > 0:
                            wind_str += f" (gust {f['wind_gust']})"
                        tooltip_parts.append(wind_str)
                    if "pressure_mb" in f:
                        tooltip_parts.append(f"📊 {f['pressure_mb']:.1f} mb")
                    if "baro_mb" in f:
                        tooltip_parts.append(f"📊 {f['baro_mb']:.1f} mb")
                    # Rain - combine into one line if present
                    rain_parts = []
                    if f.get("rain_1h") is not None:
                        rain_parts.append(f"{f['rain_1h']:.2f}\"/1h")
                    if f.get("rain_24h") is not None:
                        rain_parts.append(f"{f['rain_24h']:.2f}\"/24h")
                    if rain_parts:
                        tooltip_parts.append(f"🌧️ {' '.join(rain_parts)}")
                
                # Add comment/status if present
                if f.get("comment"):
                    comment_text = clean_aprs_comment(f["comment"], 120)
                    if comment_text:
                        tooltip_parts.append(f"💬 {comment_text}")
                        # Log comment to live feed
                        self._log(f"  💬 {clean_aprs_comment(f['comment'], 80)}", "#64b5f6")
                
                # Via path is added by JS updateStation() for the popup
                # Don't duplicate it here in the tooltip
                
                # Add timestamp
                tooltip_parts.append(f"🕐 {ts}")
                
                # Join with <br>
                tooltip = "<br>".join(tooltip_parts)
                
                # Build URL relative to BASE_DIR
                try:
                    rel_path = ic.relative_to(BASE_DIR)
                    icon_url = f"http://127.0.0.1:{self.http_port}/{rel_path.as_posix()}"
                except ValueError:
                    icon_url = f"http://127.0.0.1:{self.http_port}/aprs_icon_cache/{ic.name}"
                
                # Use JSON encoding for proper escaping
                import json
                src_js = json.dumps(src)
                tooltip_js = json.dumps(tooltip)
                via_js = json.dumps(via if via else "")
                is_digi_js = "true" if is_digi else "false"
                
                js = f"queueStation({src_js},{lat},{lon},'{icon_url}',{tooltip_js},{is_digi_js},{via_js})"
                
                
                if self.map_ready:
                    self.map.page().runJavaScript(js)
                else:
                    self.pending_js.append(js)
        else:
            self._log(f"{src}>{dst}: {info[:60]}")

    def _log(self, txt, color=None, no_ts=False):
        """Log text to the display. Optionally with color (HTML color code).
        
        Args:
            txt: Text to log
            color: Optional HTML color code
            no_ts: If True, skip timestamp (for continuation lines)
        """
        # Skip empty messages
        if not txt or not txt.strip():
            return
        
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        # Auto-detect continuation lines (start with spaces, →, or certain emoji)
        is_continuation = no_ts or txt.startswith('  ') or txt.startswith('→') or txt.startswith('    ')
        
        if hasattr(self, "log_txt"):
            if is_continuation:
                # Continuation line - no timestamp
                if color:
                    formatted = f'<span style="color:{color}">{txt}</span>'
                else:
                    formatted = txt
            elif color:
                # Use HTML for colored text - bright white bold timestamp
                formatted = f'<span style="color:#ffffff;font-weight:bold">[{timestamp}]</span> <span style="color:{color}">{txt}</span>'
            else:
                # Bright white bold timestamp, default text
                formatted = f'<span style="color:#ffffff;font-weight:bold">[{timestamp}]</span> {txt}'
            
            # Store in history
            if hasattr(self, "log_history"):
                self.log_history.append(formatted)
                # Limit history to 1000 entries
                if len(self.log_history) > 1000:
                    self.log_history = self.log_history[-1000:]
            
            # Only show if matches filter (or no filter)
            filter_text = ""
            if hasattr(self, "log_filter"):
                filter_text = self.log_filter.text().strip().upper()
            
            if not filter_text or filter_text in txt.upper():
                self.log_txt.append(formatted)
                self.log_txt.verticalScrollBar().setValue(self.log_txt.verticalScrollBar().maximum())
        else:
            self.log_buf.append(txt)
    
    def _log_link_clicked(self, url):
        """Handle clicks on callsign links in the live feed."""
        url_str = url.toString()
        if url_str.startswith("aprs://pan/"):
            # Extract callsign and pan to it on the map
            callsign = url_str.replace("aprs://pan/", "")
            if self.map_ready:
                import json
                js = f"panToStation({json.dumps(callsign)})"
                self.map.page().runJavaScript(js)
                # Switch to RX tab if not there
                if hasattr(self, 'tabs'):
                    self.tabs.setCurrentIndex(0)  # RX tab
        elif url_str.startswith("http://") or url_str.startswith("https://"):
            # External link - open in browser
            from PyQt6.QtGui import QDesktopServices
            QDesktopServices.openUrl(url)
    
    def _filter_log(self, filter_text):
        """Filter the live feed log by callsign or text."""
        if not hasattr(self, "log_history"):
            return
        
        filter_text = filter_text.strip().upper()
        self.log_txt.clear()
        
        for entry in self.log_history:
            # Check if filter matches (case insensitive)
            if not filter_text or filter_text in entry.upper():
                self.log_txt.append(entry)
        
        # Scroll to bottom
        self.log_txt.verticalScrollBar().setValue(self.log_txt.verticalScrollBar().maximum())


def main():
    # Force GPU acceleration for QtWebEngine (Chromium-based)
    # Must be set BEFORE QApplication is created
    import os
    os.environ['QTWEBENGINE_CHROMIUM_FLAGS'] = ' '.join([
        '--enable-gpu-rasterization',
        '--enable-native-gpu-memory-buffers', 
        '--enable-accelerated-2d-canvas',
        '--enable-zero-copy',
        '--ignore-gpu-blocklist',
        '--disable-gpu-driver-bug-workarounds',
        '--num-raster-threads=4',
    ])
    
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    
    # Start local HTTP server - try multiple ports if needed
    http_port = HTTP_PORT
    server = None
    for port_offset in range(10):  # Try up to 10 ports
        try:
            test_port = HTTP_PORT + port_offset
            print(f"Trying HTTP server on port {test_port}...")
            server = start_http_server(test_port)
            http_port = test_port
            print(f"HTTP server running at http://127.0.0.1:{http_port}/")
            break
        except OSError as e:
            # Port in use - try next (Windows: 10048, Linux/Mac: 98)
            if getattr(e, 'winerror', None) == 10048 or getattr(e, 'errno', None) in (98, 48) or "Address already in use" in str(e):
                print(f"Port {test_port} in use, trying next...")
                continue
            else:
                QMessageBox.critical(None, "Error", f"Cannot start HTTP server:\n{e}")
                return
    
    if server is None:
        QMessageBox.critical(None, "Error", 
            f"Cannot start HTTP server - ports {HTTP_PORT}-{HTTP_PORT+9} all in use.\n\n"
            "Another PyTNC Pro instance may be running.\n"
            "Check Task Manager and close any existing instances.")
        return
    
    try:
        QWebEngineView()
    except Exception as e:
        QMessageBox.critical(None, "Error", f"QtWebEngine error:\n{e}")
        return
    
    win = MainWindow(http_port)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()