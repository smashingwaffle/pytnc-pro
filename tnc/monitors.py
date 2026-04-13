"""
monitors.py — Map overlay monitors mixin for PyTNC Pro
Handles: earthquakes, AQI, fires, weather alerts, hospitals
Used as a mixin: class MainWindow(MonitorsMixin, ...)
"""

import csv
import json
import shutil
import time
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

from PyQt6.QtCore import Qt, QTimer, QThreadPool
from PyQt6.QtWidgets import QApplication, QFileDialog, QMenu, QMessageBox, QProgressDialog


def _mg():
    """Lazy accessor for main module globals."""
    import sys
    return sys.modules['__main__'].__dict__


class MonitorsMixin:
    """Mixin providing all map overlay monitor methods."""

    def _get_center_location(self):
        """Return (lat, lon) for the current user location.
        Priority: GPS fix → lat_edit/lon_edit (saved/manual) → manual_location text → LA default
        """
        if hasattr(self, 'gps_has_fix') and self.gps_has_fix and self.gps_lat is not None:
            return self.gps_lat, self.gps_lon
        if hasattr(self, 'lat_edit') and hasattr(self, 'lon_edit'):
            lv, ln = self.lat_edit.value(), self.lon_edit.value()
            if lv != 0.0 and ln != 0.0:
                return lv, ln
        if hasattr(self, 'manual_location'):
            text = self.manual_location.text().strip()
            if text:
                try:
                    parts = text.replace(" ", "").split(",")
                    if len(parts) == 2:
                        return float(parts[0]), float(parts[1])
                except (ValueError, IndexError):
                    pass
        return 34.05, -118.25  # last-resort default

    def _toggle_earthquake_monitor(self, state):
        """Enable/disable earthquake monitoring"""
        enabled = state == Qt.CheckState.Checked.value
        self.quake_refresh_btn.setEnabled(enabled)
        
        # Sync RX tab checkbox
        if hasattr(self, 'rx_quake_check'):
            self.rx_quake_check.blockSignals(True)
            self.rx_quake_check.setChecked(enabled)
            self.rx_quake_check.blockSignals(False)
        
        if enabled:
            self.quake_status.setText("🟢 Enabled")
            self.quake_status.setStyleSheet("color: #69f0ae;")
            self._fetch_earthquakes()
            # Set up auto-refresh timer (every 5 minutes)
            if not hasattr(self, 'quake_timer'):
                self.quake_timer = QTimer()
                self.quake_timer.timeout.connect(self._fetch_earthquakes)
            self.quake_timer.start(5 * 60 * 1000)  # 5 minutes
        else:
            self.quake_status.setText("⚫ Disabled")
            self.quake_status.setStyleSheet("color: #888;")
            if hasattr(self, 'quake_timer'):
                self.quake_timer.stop()
            # Clear earthquakes from map
            if self.map_ready:
                self.map.page().runJavaScript("clearEarthquakes()")
    
    def _fetch_earthquakes(self):
        """Fetch earthquake data from USGS"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        import urllib.request
        import json as json_module
        
        # Get settings
        radius_miles = self.quake_radius.value()
        min_mag = self.quake_min_mag.value()
        time_range = self.quake_time_range.currentData() if hasattr(self, 'quake_time_range') else "day"
        
        # Map time range to USGS starttime parameter
        time_params = {
            "hour": "&starttime=" + (datetime.now() - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S"),
            "day": "&starttime=" + (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%dT%H:%M:%S"),
            "week": "&starttime=" + (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S"),
            "month": "&starttime=" + (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S")
        }
        time_param = time_params.get(time_range, time_params["day"])
        
        # Get center point (use manual location or GPS)
        center_lat, center_lon = self._get_center_location()
        
        # Convert radius to km for USGS API
        radius_km = radius_miles * 1.60934
        
        # USGS API with time range
        url = f"https://earthquake.usgs.gov/fdsnws/event/1/query?format=geojson&latitude={center_lat}&longitude={center_lon}&maxradiuskm={radius_km}&minmagnitude={min_mag}&orderby=time&limit=100{time_param}"
        
        time_labels = {"hour": "1hr", "day": "24hr", "week": "7d", "month": "30d"}
        self._log(f"🌋 Fetching {time_labels.get(time_range, '24hr')} earthquakes within {radius_miles}mi...")
        self._log(f"   URL: {url[:80]}...")
        self.quake_status.setText("🔄")
        self.quake_status.setStyleSheet("color: #ffd54f;")
        
        # Non-blocking fetch using QThreadPool
        worker = NetworkFetchWorker(url, timeout=30)
        worker.signals.finished.connect(self._process_earthquake_data)
        worker.signals.error.connect(lambda e: self._process_earthquake_data({"error": e}))
        QThreadPool.globalInstance().start(worker)
    
    def _process_earthquake_data(self, data):
        """Process earthquake data and update map"""
        if "error" in data:
            self._log(f"❌ Earthquake fetch failed: {data['error']}")
            self.quake_status.setText("🔴 Error")
            self.quake_status.setStyleSheet("color: #ef5350;")
            return
        
        features = data.get("features", [])
        count = len(features)
        
        if count == 0:
            self._log("🌋 No earthquakes found in range")
            self.quake_status.setText("🟢 0 quakes")
            self.quake_status.setStyleSheet("color: #69f0ae;")
            if self.map_ready:
                self.map.page().runJavaScript("clearEarthquakes()")
            return
        
        self._log(f"🌋 Found {count} earthquakes")
        self.quake_status.setText(f"🟢 {count} quakes")
        self.quake_status.setStyleSheet("color: #69f0ae;")
        
        # Track recent quakes for alerts
        recent_count = 0
        now_ms = datetime.now().timestamp() * 1000
        
        # Build list for bulk loading (faster than Python→JS per marker)
        quake_list = []
        import json
        
        for feature in features:
            props = feature.get("properties", {})
            geom = feature.get("geometry", {})
            coords = geom.get("coordinates", [0, 0, 0])
            
            lon, lat, depth = coords[0], coords[1], coords[2] if len(coords) > 2 else 0
            mag = props.get("mag", 0) or 0
            place = props.get("place", "Unknown")
            time_ms = props.get("time", 0)
            
            # Check if recent (within 24 hours)
            age_hours = (now_ms - time_ms) / (1000 * 60 * 60)
            is_recent = age_hours < 24
            if is_recent:
                recent_count += 1
            
            # Format time
            try:
                quake_time = datetime.fromtimestamp(time_ms / 1000).strftime("%m/%d %H:%M")
                if age_hours < 1:
                    quake_time += " (NEW!)"
                elif age_hours < 24:
                    quake_time += f" ({age_hours:.0f}h ago)"
            except (ValueError, OSError, OverflowError):
                quake_time = "Unknown"
            
            # Color based on magnitude AND recency
            if is_recent:
                # Recent quakes - brighter colors
                if mag >= 5.0:
                    color = "#ff0000"  # Bright red
                elif mag >= 4.0:
                    color = "#ff6600"  # Bright orange
                elif mag >= 3.0:
                    color = "#ffcc00"  # Bright yellow
                else:
                    color = "#00ff00"  # Bright green
            else:
                # Older quakes - muted colors
                if mag >= 5.0:
                    color = "#aa4444"  # Muted red
                elif mag >= 4.0:
                    color = "#aa6644"  # Muted orange
                elif mag >= 3.0:
                    color = "#aaaa44"  # Muted yellow
                else:
                    color = "#44aa44"  # Muted green
            
            # Size based on magnitude
            size = max(8, min(30, int(mag * 5)))
            
            tooltip = f"M{mag:.1f} - {place}<br>Depth: {depth:.1f}km<br>{quake_time}"
            
            quake_list.append({
                "lat": lat,
                "lon": lon,
                "mag": mag,
                "color": color,
                "size": size,
                "tooltip": tooltip,
                "isRecent": is_recent
            })
        
        # Send all earthquakes in one JS call
        if self.map_ready and quake_list:
            js = f"setEarthquakesBulk({json.dumps(quake_list)})"
            self.map.page().runJavaScript(js)

        # Alert for M5.0+ within 50km
        ALERT_MAG = 5.0
        ALERT_KM  = 50
        import math

        def _dist_km(lat1, lon1, lat2, lon2):
            R = 6371
            phi1, phi2 = math.radians(lat1), math.radians(lat2)
            dp = math.radians(lat2 - lat1)
            dl = math.radians(lon2 - lon1)
            a = math.sin(dp/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dl/2)**2
            return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

        center_lat, center_lon = self._get_center_location()
        alert_quake = None
        for q in quake_list:
            if q["mag"] >= ALERT_MAG and q.get("isRecent"):
                dist = _dist_km(center_lat, center_lon, q["lat"], q["lon"])
                if dist <= ALERT_KM:
                    if alert_quake is None or q["mag"] > alert_quake[0]:
                        alert_quake = (q["mag"], q["tooltip"].split("<br>")[0].replace("M","M"), dist, q["color"], q["tooltip"])

        if alert_quake and self.map_ready:
            mag, place, dist, color, tooltip = alert_quake
            place_str = place.split(" - ")[-1] if " - " in place else place
            self.map.page().runJavaScript(
                f"showQuakeAlert({mag}, {json.dumps(place_str)}, {dist:.1f}, {json.dumps(color)});"
            )
            self._log(f"🚨 ALERT: M{mag:.1f} earthquake {dist:.0f}km away — {place_str}")
        
        # Alert if there are recent quakes
        if recent_count > 0:
            self.quake_status.setText(f"🔴 {recent_count} new!")
            self.quake_status.setStyleSheet("color: #ff6600; font-weight: bold;")
            self._log(f"⚠️ {recent_count} earthquakes in last 24 hours!")
        else:
            self.quake_status.setText(f"🟢 {count}")
            self.quake_status.setStyleSheet("color: #69f0ae;")

    # =========================================================================
    # Air Quality (AQI) Monitor - AirNow API
    # =========================================================================
    
    def _toggle_aqi_monitor(self, state):
        """Enable/disable AQI monitoring"""
        enabled = state == Qt.CheckState.Checked.value
        
        if hasattr(self, 'aqi_refresh_btn'):
            self.aqi_refresh_btn.setEnabled(enabled)
        
        # Sync RX tab checkbox
        if hasattr(self, 'rx_aqi_check'):
            self.rx_aqi_check.blockSignals(True)
            self.rx_aqi_check.setChecked(enabled)
            self.rx_aqi_check.blockSignals(False)
        
        if enabled:
            if hasattr(self, 'aqi_status'):
                self.aqi_status.setText("🟢 Enabled")
                self.aqi_status.setStyleSheet("color: #69f0ae;")
            self._fetch_aqi_data()
            # Set up auto-refresh timer (every 30 minutes - AQI doesn't change fast)
            if not hasattr(self, 'aqi_timer'):
                self.aqi_timer = QTimer()
                self.aqi_timer.timeout.connect(self._fetch_aqi_data)
            self.aqi_timer.start(30 * 60 * 1000)  # 30 minutes
        else:
            if hasattr(self, 'aqi_status'):
                self.aqi_status.setText("⚫ Disabled")
                self.aqi_status.setStyleSheet("color: #888;")
            if hasattr(self, 'aqi_timer'):
                self.aqi_timer.stop()
            # Clear AQI from map
            if self.map_ready:
                self.map.page().runJavaScript("clearAQI()")
    
    def _fetch_aqi_data(self):
        """Fetch AQI data from AirNow API"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        # Get center point (use GPS first, then manual location)
        center_lat, center_lon = self._get_center_location()
        
        # AirNow API - requires API key
        api_key = self.aqi_api_key.text().strip() if hasattr(self, 'aqi_api_key') else ""
        if not api_key:
            self._log("💨 AQI: API key required!")
            self._log("   Get free key from: https://docs.airnowapi.org/")
            if hasattr(self, 'aqi_status'):
                self.aqi_status.setText("⚠️ Need key")
                self.aqi_status.setStyleSheet("color: #ffd54f;")
            return
        
        url = f"https://www.airnowapi.org/aq/observation/latLong/current/?format=application/json&latitude={center_lat}&longitude={center_lon}&distance=25&API_KEY={api_key}"
        
        self._log(f"💨 Fetching AQI data...")
        if hasattr(self, 'aqi_status'):
            self.aqi_status.setText("🔄")
            self.aqi_status.setStyleSheet("color: #ffd54f;")
        
        # Non-blocking fetch using QThreadPool
        headers = {'User-Agent': 'PyTNC-Pro/1.0'}
        worker = NetworkFetchWorker(url, headers=headers, timeout=15)
        worker.signals.finished.connect(self._process_aqi_data)
        worker.signals.error.connect(lambda e: self._process_aqi_data({"error": e}))
        QThreadPool.globalInstance().start(worker)
    
    def _process_aqi_data(self, data):
        """Process AQI data from AirNow"""
        if "error" in data:
            self._log(f"❌ AQI fetch failed: {data['error']}")
            if hasattr(self, 'aqi_status'):
                self.aqi_status.setText("🔴 Err")
                self.aqi_status.setStyleSheet("color: #ef5350;")
            return
        
        # AirNow returns a list of observations
        if not isinstance(data, list) or len(data) == 0:
            self._log("💨 No AQI data available for this location")
            if hasattr(self, 'aqi_status'):
                self.aqi_status.setText("🟡 No data")
                self.aqi_status.setStyleSheet("color: #ffd54f;")
            return
        
        # Find PM2.5 or O3 reading (most relevant for smoke)
        pm25_data = None
        o3_data = None
        for obs in data:
            param = obs.get("ParameterName", "")
            if "PM2.5" in param:
                pm25_data = obs
            elif "O3" in param or "OZONE" in param.upper():
                o3_data = obs
        
        # Prefer PM2.5 (better for smoke), fall back to O3
        aqi_obs = pm25_data or o3_data or (data[0] if data else None)
        
        if not aqi_obs:
            self._log("💨 No valid AQI readings found")
            return
        
        aqi_value = aqi_obs.get("AQI", 0)
        param = aqi_obs.get("ParameterName", "AQI")
        category = aqi_obs.get("Category", {}).get("Name", "Unknown")
        reporting_area = aqi_obs.get("ReportingArea", "")
        lat = aqi_obs.get("Latitude", 0)
        lon = aqi_obs.get("Longitude", 0)
        
        # AQI color coding per EPA standards
        if aqi_value <= 50:
            color = "#00e400"  # Good - Green
            emoji = "🟢"
        elif aqi_value <= 100:
            color = "#ffff00"  # Moderate - Yellow
            emoji = "🟡"
        elif aqi_value <= 150:
            color = "#ff7e00"  # Unhealthy for Sensitive Groups - Orange
            emoji = "🟠"
        elif aqi_value <= 200:
            color = "#ff0000"  # Unhealthy - Red
            emoji = "🔴"
        elif aqi_value <= 300:
            color = "#8f3f97"  # Very Unhealthy - Purple
            emoji = "🟣"
        else:
            color = "#7e0023"  # Hazardous - Maroon
            emoji = "⛔"
        
        self._log(f"💨 AQI: {aqi_value} ({category}) - {param}")
        self._log(f"   {reporting_area}")
        
        if hasattr(self, 'aqi_status'):
            self.aqi_status.setText(f"{emoji} {aqi_value}")
            self.aqi_status.setStyleSheet(f"color: {color}; font-weight: bold;")
        
        # Add AQI marker to map
        if self.map_ready and lat and lon:
            tooltip = f"<b>💨 AQI: {aqi_value}</b><br>{category}<br>{param}<br>{reporting_area}"
            import json
            js = f"addAQIMarker({lat},{lon},{aqi_value},'{color}',{json.dumps(tooltip)})"
            self.map.page().runJavaScript(js)

    # =========================================================================
    # Fire/Wildfire Monitor (NASA FIRMS)
    # =========================================================================
    
    def _toggle_fire_monitor(self, state):
        """Toggle fire monitoring on/off"""
        enabled = state == Qt.CheckState.Checked.value
        self.fire_refresh_btn.setEnabled(enabled)
        
        # Sync RX tab checkbox
        if hasattr(self, 'rx_fire_check'):
            self.rx_fire_check.blockSignals(True)
            self.rx_fire_check.setChecked(enabled)
            self.rx_fire_check.blockSignals(False)
        
        if enabled:
            api_key = self.fire_api_key.text().strip()
            if not api_key:
                self._log("🔥 Fire monitor: API key required!")
                self._log("   Get free key from: https://firms.modaps.eosdis.nasa.gov/api/area/")
                self.fire_status.setText("⚠️ Need key")
                self.fire_status.setStyleSheet("color: #ffd54f;")
                return
            
            self._fetch_fires()
            # Auto-refresh every 30 minutes
            if not hasattr(self, 'fire_timer'):
                self.fire_timer = QTimer()
                self.fire_timer.timeout.connect(self._fetch_fires)
            self.fire_timer.start(30 * 60 * 1000)  # 30 minutes
        else:
            if hasattr(self, 'fire_timer'):
                self.fire_timer.stop()
            # Clear fire markers
            if self.map_ready:
                self.map.page().runJavaScript("clearFires()")
            self.fire_status.setText("⚫")
            self.fire_status.setStyleSheet("color: #888;")
    
    def _fetch_fires(self):
        """Fetch fire/hotspot data from NASA FIRMS API"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        import urllib.request
        
        api_key = self.fire_api_key.text().strip()
        if not api_key:
            self._log("🔥 Fire fetch: No API key")
            return
        
        # Get settings
        time_range = self.fire_time_range.currentData() if hasattr(self, 'fire_time_range') else "24h"
        source = self.fire_source.currentData() if hasattr(self, 'fire_source') else "VIIRS_SNPP_NRT"
        
        # Get center point
        lat, lon = self._get_center_location()
        
        # NASA FIRMS API - area query (bounding box around center)
        # Format: west,south,east,north
        radius_deg = 2.0  # ~140 miles at LA latitude
        west = lon - radius_deg
        east = lon + radius_deg
        south = lat - radius_deg
        north = lat + radius_deg
        
        # Map time range to FIRMS day_range parameter
        day_map = {"24h": 1, "48h": 2, "7d": 7}
        days = day_map.get(time_range, 1)
        
        # FIRMS CSV API URL
        url = f"https://firms.modaps.eosdis.nasa.gov/api/area/csv/{api_key}/{source}/{west},{south},{east},{north}/{days}"
        
        self._log(f"🔥 Fetching {time_range} fire data ({source})...")
        self.fire_status.setText("🔄")
        self.fire_status.setStyleSheet("color: #ffd54f;")
        
        # Non-blocking fetch
        worker = NetworkFetchWorker(url, timeout=30)
        worker.signals.finished.connect(self._process_fire_data_csv)
        worker.signals.error.connect(lambda e: self._process_fire_error(e))
        QThreadPool.globalInstance().start(worker)
    
    def _process_fire_error(self, error):
        """Handle fire fetch error"""
        self._log(f"❌ Fire fetch failed: {error}")
        self.fire_status.setText("🔴 Error")
        self.fire_status.setStyleSheet("color: #ef5350;")
        
        if "403" in str(error) or "401" in str(error):
            self._log("   Check your NASA FIRMS API key")
    
    def _process_fire_data_csv(self, data):
        """Process fire CSV data from NASA FIRMS"""
        # NetworkFetchWorker returns dict with raw text for non-JSON
        if isinstance(data, dict):
            if "error" in data:
                self._process_fire_error(data["error"])
                return
            # If it parsed as JSON (error response), check for message
            if "message" in data:
                self._log(f"❌ FIRMS API error: {data.get('message', 'Unknown')}")
                self.fire_status.setText("🔴 API Error")
                self.fire_status.setStyleSheet("color: #ef5350;")
                return
        
        # Clear existing fires
        if self.map_ready:
            self.map.page().runJavaScript("clearFires()")
        
        # Parse CSV response
        try:
            lines = data if isinstance(data, str) else str(data)
            if not lines or "latitude" not in lines.lower():
                self._log("🔥 No fire data or invalid response")
                self.fire_status.setText("⚫ No data")
                self.fire_status.setStyleSheet("color: #888;")
                return
            
            rows = lines.strip().split('\n')
            if len(rows) < 2:
                self._log("🔥 No fires detected in area")
                self.fire_status.setText("🟢 0 fires")
                self.fire_status.setStyleSheet("color: #69f0ae;")
                return
            
            # Parse header to find column indices
            header = rows[0].lower().split(',')
            try:
                lat_idx = header.index('latitude')
                lon_idx = header.index('longitude')
                bright_idx = header.index('bright_ti4') if 'bright_ti4' in header else header.index('brightness')
                conf_idx = header.index('confidence') if 'confidence' in header else -1
                sat_idx = header.index('satellite') if 'satellite' in header else -1
                time_idx = header.index('acq_time') if 'acq_time' in header else -1
                date_idx = header.index('acq_date') if 'acq_date' in header else -1
            except ValueError as e:
                self._log(f"🔥 CSV parse error: {e}")
                self.fire_status.setText("🔴 Parse err")
                return
            
            # Build list for bulk loading (faster than Python→JS per marker)
            fire_list = []
            import json
            
            for row in rows[1:]:
                cols = row.split(',')
                if len(cols) <= max(lat_idx, lon_idx, bright_idx):
                    continue
                
                try:
                    lat = float(cols[lat_idx])
                    lon = float(cols[lon_idx])
                    brightness = float(cols[bright_idx])
                    confidence = cols[conf_idx] if conf_idx >= 0 and conf_idx < len(cols) else "N/A"
                    satellite = cols[sat_idx] if sat_idx >= 0 and sat_idx < len(cols) else "Unknown"
                    acq_time = cols[time_idx] if time_idx >= 0 and time_idx < len(cols) else ""
                    acq_date = cols[date_idx] if date_idx >= 0 and date_idx < len(cols) else ""
                    
                    # Build tooltip
                    tooltip = f"Brightness: {brightness:.0f}K<br>"
                    tooltip += f"Confidence: {confidence}<br>"
                    tooltip += f"Satellite: {satellite}<br>"
                    if acq_date and acq_time:
                        tooltip += f"Detected: {acq_date} {acq_time}"
                    
                    fire_list.append({
                        "lat": lat,
                        "lon": lon,
                        "brightness": brightness,
                        "tooltip": tooltip
                    })
                except (ValueError, IndexError):
                    continue
            
            fire_count = len(fire_list)
            
            # Send all fires in one JS call
            if self.map_ready and fire_list:
                js = f"setFiresBulk({json.dumps(fire_list)})"
                self.map.page().runJavaScript(js)
            
            self._log(f"🔥 Found {fire_count} fire hotspots")
            
            if fire_count > 0:
                self.fire_status.setText(f"🔴 {fire_count} fires!")
                self.fire_status.setStyleSheet("color: #ff6600; font-weight: bold;")
            else:
                self.fire_status.setText("🟢 0 fires")
                self.fire_status.setStyleSheet("color: #69f0ae;")
                
        except Exception as e:
            self._log(f"🔥 Fire data processing error: {e}")
            self.fire_status.setText("🔴 Error")
            self.fire_status.setStyleSheet("color: #ef5350;")

    # =========================================================================
    # Hospital Layer (with offline caching)
    # =========================================================================
    
    def _get_hospital_cache_file(self):
        """Get path to hospital cache file"""
        cache_dir = _mg()['CACHE_DIR']
        cache_dir.mkdir(exist_ok=True)
        return cache_dir / "hospitals.json"
    
    def _load_hospital_cache(self):
        """Load hospitals from local cache"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        cache_file = self._get_hospital_cache_file()
        if cache_file.exists():
            try:
                import json as json_module
                with open(cache_file, 'r') as f:
                    data = json_module.load(f)
                    # Update offline indicator
                    cached_time = data.get("cached_at", "Unknown")
                    count = len(data.get("hospitals", []))
                    self.hospital_offline_indicator.setText(f"💾 {count}")
                    self.hospital_offline_indicator.setToolTip(f"Cached: {cached_time}")
                    return data
            except Exception as e:
                self._log(f"⚠️ Hospital cache load failed: {e}")
        return None
    
    def _save_hospital_cache(self, hospitals):
        """Save hospitals to local cache for offline use"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        cache_file = self._get_hospital_cache_file()
        try:
            import json as json_module
            cur_lat, cur_lon = self._get_center_location()
            data = {
                "cached_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "lat": cur_lat,
                "lon": cur_lon,
                "hospitals": hospitals
            }
            with open(cache_file, 'w') as f:
                json_module.dump(data, f)
            self._log(f"💾 Saved {len(hospitals)} hospitals to cache")
            self.hospital_offline_indicator.setText(f"💾 {len(hospitals)}")
            self.hospital_offline_indicator.setToolTip(f"Cached: {data['cached_at']}")
        except Exception as e:
            self._log(f"⚠️ Hospital cache save failed: {e}")
    
    def _rx_toggle_callsigns(self, state):
        """Toggle callsign labels visibility on map"""
        enabled = state == Qt.CheckState.Checked.value
        if self.map_ready:
            js = f"toggleCallsignLabels({str(enabled).lower()})"
            self.map.page().runJavaScript(js)
    
    def _rx_toggle_trails(self, state):
        """Toggle station trails visibility on map"""
        enabled = state == Qt.CheckState.Checked.value
        if self.map_ready:
            js = f"toggleTrails({str(enabled).lower()})"
            self.map.page().runJavaScript(js)
    
    def _rx_toggle_hospitals(self, state):
        """Toggle hospitals from RX page checkbox - syncs with Settings"""
        # Sync to settings checkbox without triggering its signal
        if hasattr(self, 'hospital_enabled'):
            self.hospital_enabled.blockSignals(True)
            self.hospital_enabled.setChecked(state == Qt.CheckState.Checked.value)
            self.hospital_enabled.blockSignals(False)
        # Actually toggle the layer (only once, here)
        self._toggle_hospital_layer(state)
    
    def _toggle_hospital_layer(self, state):
        """Enable/disable hospital layer"""
        enabled = state == Qt.CheckState.Checked.value
        
        # Sync RX page checkbox
        if hasattr(self, 'rx_hospital_check'):
            self.rx_hospital_check.blockSignals(True)
            self.rx_hospital_check.setChecked(enabled)
            self.rx_hospital_check.blockSignals(False)
        
        if enabled:
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setText("🔄")
                self.hospital_status.setStyleSheet("color: #ffd54f;")
            cache = self._load_hospital_cache()
            if cache and cache.get("hospitals"):
                cur_lat, cur_lon = self._get_center_location()
                cached_lat = cache.get("lat", None)
                cached_lon = cache.get("lon", None)
                if cached_lat is not None and cached_lon is not None:
                    dist = ((cur_lat - cached_lat)**2 + (cur_lon - cached_lon)**2) ** 0.5
                    location_changed = dist > 0.5
                else:
                    location_changed = True
                if location_changed:
                    self._log("🏥 Location changed — fetching hospitals for current area...")
                    self._fetch_hospitals()
                else:
                    self._log("🏥 Loading hospitals from offline cache...")
                    self._display_hospitals(cache["hospitals"])
                    return  # don't fall through to fetch
            else:
                self._log("🏥 No cached data - fetching from internet...")
            self._fetch_hospitals()
        else:
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setText("⚫")
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setStyleSheet("color: #888;")
            # Clear hospitals from map
            if self.map_ready:
                self.map.page().runJavaScript("clearHospitals()")
    
    def _fetch_hospitals(self):
        """Fetch hospital data from OpenStreetMap Overpass API"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        import json as json_module
        
        radius_miles = self.hospital_radius.value()
        radius_meters = int(radius_miles * 1609.34)
        
        # Get center point — GPS → lat_edit/lon_edit → manual_location text → default
        if hasattr(self, 'gps_has_fix') and self.gps_has_fix and self.gps_lat is not None:
            center_lat = self.gps_lat
            center_lon = self.gps_lon
        elif hasattr(self, 'lat_edit') and self.lat_edit.value() != 0.0:
            center_lat = self.lat_edit.value()
            center_lon = self.lon_edit.value()
        else:
            manual_text = self.manual_location.text().strip() if hasattr(self, 'manual_location') else ""
        # Get center point — GPS → lat_edit/lon_edit → manual_location text → default
        center_lat, center_lon = self._get_center_location()
        
        # Overpass API query
        query = f'[out:json];(node["amenity"="hospital"](around:{radius_meters},{center_lat},{center_lon});way["amenity"="hospital"](around:{radius_meters},{center_lat},{center_lon}););out center 100;'
        
        # Try multiple Overpass servers (some may be down or slow)
        overpass_servers = [
            "https://overpass-api.de/api/interpreter",
            "https://overpass.kumi.systems/api/interpreter",
        ]
        url = f"{overpass_servers[0]}?data={urllib.parse.quote(query)}"
        
        # Store fallback URL for retry
        self._hospital_fallback_url = f"{overpass_servers[1]}?data={urllib.parse.quote(query)}"
        self._hospital_fallback_tried = False
        
        self._log(f"🏥 Downloading hospitals within {radius_miles}mi...")
        if hasattr(self, 'hospital_status'):
            self.hospital_status.setText("⬇️")
        if hasattr(self, 'hospital_status'):
            self.hospital_status.setStyleSheet("color: #ffd54f;")
        
        # Non-blocking fetch using QThreadPool
        worker = NetworkFetchWorker(url, timeout=15)
        worker.signals.finished.connect(self._process_hospital_data)
        worker.signals.error.connect(lambda e: self._process_hospital_data({"error": e}))
        QThreadPool.globalInstance().start(worker)
    
    def _process_hospital_data(self, data):
        """Process hospital data and update map"""
        NetworkFetchWorker = _mg().get("NetworkFetchWorker", None)
        if "error" in data:
            # Try fallback server if not already tried
            if hasattr(self, '_hospital_fallback_url') and not getattr(self, '_hospital_fallback_tried', True):
                self._hospital_fallback_tried = True
                self._log(f"⚠️ Primary server failed, trying fallback...")
                worker = NetworkFetchWorker(self._hospital_fallback_url, timeout=30)
                worker.signals.finished.connect(self._process_hospital_data)
                worker.signals.error.connect(lambda e: self._process_hospital_data({"error": e}))
                QThreadPool.globalInstance().start(worker)
                return
            
            self._log(f"❌ Hospital fetch failed: {data['error']}")
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setText("🔴 Err")
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setStyleSheet("color: #ef5350;")
            # Try to use cached data
            cache = self._load_hospital_cache()
            if cache and cache.get("hospitals"):
                self._log("📴 Using offline cache instead...")
                self._display_hospitals(cache["hospitals"])
            return
        
        elements = data.get("elements", [])
        
        if len(elements) == 0:
            self._log("🏥 No hospitals found in range")
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setText("0")
            if hasattr(self, 'hospital_status'):
                self.hospital_status.setStyleSheet("color: #69f0ae;")
            return
        
        # Parse and store hospitals
        hospitals = []
        for element in elements:
            # Get coordinates (node has lat/lon directly, way has center)
            if element.get("type") == "node":
                lat = element.get("lat", 0)
                lon = element.get("lon", 0)
            else:
                center = element.get("center", {})
                lat = center.get("lat", 0)
                lon = center.get("lon", 0)
            
            if lat == 0 or lon == 0:
                continue
            
            tags = element.get("tags", {})
            hospital = {
                "lat": lat,
                "lon": lon,
                "name": tags.get("name", "Hospital"),
                "address": tags.get("addr:street", ""),
                "housenumber": tags.get("addr:housenumber", ""),
                "phone": tags.get("phone", ""),
                "emergency": tags.get("emergency", "")
            }
            hospitals.append(hospital)
        
        # Save to cache for offline use
        self._save_hospital_cache(hospitals)
        
        # Display on map
        self._display_hospitals(hospitals)
    
    def _display_hospitals(self, hospitals):
        """Display hospitals on the map"""
        # Guard: only display if hospital layer is enabled
        if hasattr(self, 'rx_hospital_check') and not self.rx_hospital_check.isChecked():
            self._log("🏥 Skipping hospital display - layer disabled")
            return
        
        count = len(hospitals)
        self._log(f"🏥 Displaying {count} hospitals")
        if hasattr(self, 'hospital_status'):
            self.hospital_status.setText(f"🟢 {count}")
        if hasattr(self, 'hospital_status'):
            self.hospital_status.setStyleSheet("color: #69f0ae;")
        
        # Build list for bulk loading (faster than Python→JS per marker)
        hospital_list = []
        import json
        
        for h in hospitals:
            name = h.get("name", "Hospital")
            
            # Build tooltip
            tooltip_parts = [f"<b>{name}</b>"]
            addr = h.get("address", "")
            if addr:
                housenumber = h.get("housenumber", "")
                if housenumber:
                    addr = housenumber + " " + addr
                tooltip_parts.append(addr)
            if h.get("phone"):
                tooltip_parts.append(f"📞 {h['phone']}")
            if h.get("emergency"):
                tooltip_parts.append(f"🚑 Emergency: {h['emergency']}")
            
            tooltip = "<br>".join(tooltip_parts)
            
            hospital_list.append({
                "lat": h['lat'],
                "lon": h['lon'],
                "name": name,
                "tooltip": tooltip
            })
        
        # Send all hospitals in one JS call
        if self.map_ready and hospital_list:
            js = f"setHospitalsBulk({json.dumps(hospital_list)})"
            self.map.page().runJavaScript(js)

    # =========================================================================
    # Weather Alerts Layer (NWS API)
    # =========================================================================
    
    def _rx_toggle_weather(self, state):
        """Toggle weather alerts from RX page checkbox"""
        if hasattr(self, 'weather_enabled'):
            self.weather_enabled.blockSignals(True)
            self.weather_enabled.setChecked(state == Qt.CheckState.Checked.value)
            self.weather_enabled.blockSignals(False)
        self._toggle_weather_layer(state)
    
    def _toggle_weather_layer(self, state):
        """Enable/disable weather alerts layer"""
        enabled = state == Qt.CheckState.Checked.value
        
        # Sync RX page checkbox
        if hasattr(self, 'rx_weather_check'):
            self.rx_weather_check.blockSignals(True)
            self.rx_weather_check.setChecked(enabled)
            self.rx_weather_check.blockSignals(False)
        
        if hasattr(self, 'weather_refresh_btn'):
            self.weather_refresh_btn.setEnabled(enabled)
        
        if enabled:
            self.weather_status.setText("🔄")
            self.weather_status.setStyleSheet("color: #ffd54f;")
            self._fetch_weather_alerts()
            # Start 5-minute auto-refresh timer
            if not hasattr(self, 'weather_timer'):
                self.weather_timer = QTimer()
                self.weather_timer.timeout.connect(self._fetch_weather_alerts)
            self.weather_timer.start(5 * 60 * 1000)  # 5 minutes
            self._log("⚠️ Weather alerts auto-refresh: every 5 min")
        else:
            self.weather_status.setText("⚫")
            self.weather_status.setStyleSheet("color: #888;")
            # Stop timer
            if hasattr(self, 'weather_timer'):
                self.weather_timer.stop()
            if self.map_ready:
                self.map.page().runJavaScript("clearWeatherAlerts()")
    
    def _fetch_weather_alerts(self):
        """Fetch weather alerts from NWS API"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        import json as json_module
        
        # Get center point
        center_lat, center_lon = self._get_center_location()
        
        # NWS API - alerts for point
        url = f"https://api.weather.gov/alerts/active?point={center_lat},{center_lon}"
        
        self._log(f"⚠️ Fetching weather alerts...")
        self.weather_status.setText("🔄")
        self.weather_status.setStyleSheet("color: #ffd54f;")
        
        # Non-blocking fetch using QThreadPool
        headers = {'User-Agent': 'PyTNC-Pro/1.0', 'Accept': 'application/geo+json'}
        worker = NetworkFetchWorker(url, headers=headers, timeout=15)
        worker.signals.finished.connect(self._process_weather_data)
        worker.signals.error.connect(lambda e: self._process_weather_data({"error": e}))
        QThreadPool.globalInstance().start(worker)
    
    def _process_weather_data(self, data):
        """Process weather alert data"""
        if "error" in data:
            self._log(f"❌ Weather alert fetch failed: {data['error']}")
            self.weather_status.setText("🔴 Err")
            self.weather_status.setStyleSheet("color: #ef5350;")
            return
        
        features = data.get("features", [])
        count = len(features)
        
        if count == 0:
            self._log("⚠️ No active weather alerts")
            self.weather_status.setText("🟢 0")
            self.weather_status.setStyleSheet("color: #69f0ae;")
            return
        
        self._log(f"⚠️ Found {count} weather alerts!")
        
        # Color by severity
        has_severe = False
        for f in features:
            props = f.get("properties", {})
            severity = props.get("severity", "").lower()
            if severity in ["extreme", "severe"]:
                has_severe = True
                break
        
        if has_severe:
            self.weather_status.setText(f"🔴 {count}!")
            self.weather_status.setStyleSheet("color: #ff0000; font-weight: bold;")
        else:
            self.weather_status.setText(f"🟡 {count}")
            self.weather_status.setStyleSheet("color: #ff9800;")
        
        # Clear existing alerts
        if self.map_ready:
            self.map.page().runJavaScript("clearWeatherAlerts()")
        
        # Add each alert
        for feature in features:
            props = feature.get("properties", {})
            
            event = props.get("event", "Weather Alert")
            headline = props.get("headline", "")[:100]
            severity = props.get("severity", "Unknown")
            urgency = props.get("urgency", "Unknown")
            areas = props.get("areaDesc", "")[:50]
            
            # Color by severity
            if severity.lower() == "extreme":
                color = "#ff0000"
            elif severity.lower() == "severe":
                color = "#ff6600"
            elif severity.lower() == "moderate":
                color = "#ffcc00"
            else:
                color = "#66ccff"
            
            # Log the alert details - ALWAYS BOLD for alerts
            self._log(f"   <span style='color:{color};font-weight:bold'>⚠️ {event}: {headline}</span>")
            
            # Get coordinates - try polygon first, fall back to user location
            center_lat, center_lon = None, None
            
            geom = feature.get("geometry")
            if geom and geom.get("type") == "Polygon":
                coords = geom.get("coordinates", [[]])[0]
                if coords:
                    # Get centroid of polygon
                    lats = [c[1] for c in coords]
                    lons = [c[0] for c in coords]
                    center_lat = sum(lats) / len(lats)
                    center_lon = sum(lons) / len(lons)
            
            # Fall back to user's location if no polygon
            if center_lat is None:
                if hasattr(self, 'gps_has_fix') and self.gps_has_fix:
                    center_lat, center_lon = self.gps_lat, self.gps_lon
                else:
                    manual_text = self.manual_location.text().strip() if hasattr(self, 'manual_location') else ""
                    if manual_text:
                        try:
                            parts = manual_text.replace(" ", "").split(",")
                            center_lat, center_lon = float(parts[0]), float(parts[1])
                        except (ValueError, IndexError):
                            pass  # center_lat stays None, alert won't be plotted
            
            # Add to map if we have coordinates
            if center_lat is not None and center_lon is not None:
                tooltip = f"<b>⚠️ {event}</b><br>{headline}<br>Severity: {severity}<br>Area: {areas}"
                
                # Use JSON encoding to properly escape all special characters
                import json
                event_js = json.dumps(event)
                tooltip_js = json.dumps(tooltip)
                
                js = f"addWeatherAlert({center_lat},{center_lon},{event_js},'{color}',{tooltip_js})"
                if self.map_ready:
                    self.map.page().runJavaScript(js)
    # Offline Cache Functions
    # =========================================================================
    
    def _ensure_cache_dir(self):
        """Ensure cache directory exists"""
        CACHE_DIR = _mg()['CACHE_DIR']
        CACHE_DIR.mkdir(exist_ok=True)
        return CACHE_DIR
    
    def _update_cache_status(self):
        """Update cache status indicators"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        cache_dir = self._ensure_cache_dir()
        
        # Check for tiles in OLD location (app folder) - auto-migrate
        old_tile_dir = BASE_DIR / "tile_cache"
        if old_tile_dir.exists():
            old_count = sum(1 for _ in old_tile_dir.rglob("*.png"))
            new_count = sum(1 for _ in TILE_CACHE_DIR.rglob("*.png")) if TILE_CACHE_DIR.exists() else 0
            
            if old_count > 0 and new_count == 0:
                self._log(f"🔄 Migrating {old_count:,} cached tiles...")
                try:
                    import shutil
                    if TILE_CACHE_DIR.exists():
                        shutil.rmtree(TILE_CACHE_DIR)
                    shutil.copytree(old_tile_dir, TILE_CACHE_DIR)
                    self._log(f"✅ Tile migration complete!")
                except Exception as e:
                    self._log(f"⚠️ Migration failed: {e}")
        
        # Map tiles - check TILE_CACHE_DIR
        if TILE_CACHE_DIR.exists():
            tile_count = sum(1 for _ in TILE_CACHE_DIR.rglob("*.png"))
            if tile_count > 0:
                zoom_dirs = sorted([int(d.name) for d in TILE_CACHE_DIR.iterdir() if d.is_dir() and d.name.isdigit()])
                zoom_range = f"z{zoom_dirs[0]}-{zoom_dirs[-1]}" if zoom_dirs else "?"
                self.cache_map_status.setText(f"💾 {tile_count:,}")
                self.cache_map_status.setToolTip(f"{tile_count:,} tiles cached ({zoom_range})")
                self._log(f"🗺️ Tile cache: {tile_count:,} tiles ({zoom_range})")
                # Show per-zoom counts
                for z in zoom_dirs:
                    z_path = TILE_CACHE_DIR / str(z)
                    z_count = sum(1 for _ in z_path.rglob("*.png"))
                    self._log(f"   z{z}: {z_count:,} tiles")
            else:
                self.cache_map_status.setText("--")
        else:
            self.cache_map_status.setText("--")
        
        # Digipeaters
        digi_file = cache_dir / "digipeaters.json"
        if digi_file.exists():
            try:
                import json
                with open(digi_file) as f:
                    data = json.load(f)
                    count = len(data.get("digipeaters", []))
                    self.cache_digi_status.setText(f"💾 {count}")
            except (json.JSONDecodeError, OSError, KeyError):
                self.cache_digi_status.setText("💾 ?")
        else:
            self.cache_digi_status.setText("--")
    
    def _cache_map_tiles(self):
        """Download and cache map tiles for the LA area"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        import math
        import urllib.request
        
        # LA bounding box (covers greater LA area)
        NORTH = 34.4
        SOUTH = 33.6
        WEST = -118.8
        EAST = -117.5
        
        def lat_lon_to_tile(lat, lon, zoom):
            n = 2 ** zoom
            x = int((lon + 180) / 360 * n)
            y = int((1 - math.asinh(math.tan(math.radians(lat))) / math.pi) / 2 * n)
            return x, y
        
        # Get zoom range from slider
        min_zoom = 8
        max_zoom = self.cache_map_zoom_slider.value() if hasattr(self, 'cache_map_zoom_slider') else 14
        
        self._log(f"🗺️ Caching tiles z{min_zoom}-{max_zoom}...")
        
        # Calculate total tiles AND log per-zoom
        total_tiles = 0
        zoom_tile_counts = {}
        for zoom in range(min_zoom, max_zoom + 1):
            x1, y1 = lat_lon_to_tile(NORTH, WEST, zoom)
            x2, y2 = lat_lon_to_tile(SOUTH, EAST, zoom)
            x_min, x_max = min(x1, x2), max(x1, x2)
            y_min, y_max = min(y1, y2), max(y1, y2)
            count = (x_max - x_min + 1) * (y_max - y_min + 1)
            zoom_tile_counts[zoom] = count
            total_tiles += count
            self._log(f"   z{zoom}: x={x_min}-{x_max} ({x_max-x_min+1}), y={y_min}-{y_max} ({y_max-y_min+1}) = {count:,} tiles")
        
        # Estimate size (~20KB per tile average)
        est_size_mb = total_tiles * 20 / 1024
        
        # Confirm with user
        reply = QMessageBox.question(self, "Cache LA Map Tiles",
            f"Download {total_tiles:,} map tiles for LA area?\n\n"
            f"Zoom levels: {min_zoom}-{max_zoom}\n"
            f"Estimated size: ~{est_size_mb:.0f} MB\n\n"
            f"Higher zoom = more detail = more tiles\n"
            f"The app will wait until complete.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Create progress dialog
        progress = QProgressDialog("Caching map tiles...", "Cancel", 0, total_tiles, self)
        progress.setWindowTitle("Downloading Map Tiles")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        # Ensure cache directory
        TILE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        
        downloaded = 0
        skipped = 0
        errors = 0
        
        self._log(f"🗺️ Starting tile cache: {total_tiles} tiles, zoom {min_zoom}-{max_zoom}...")
        
        for zoom in range(min_zoom, max_zoom + 1):
            x1, y1 = lat_lon_to_tile(NORTH, WEST, zoom)
            x2, y2 = lat_lon_to_tile(SOUTH, EAST, zoom)
            x_min, x_max = min(x1, x2), max(x1, x2)
            y_min, y_max = min(y1, y2), max(y1, y2)
            
            zoom_dir = TILE_CACHE_DIR / str(zoom)
            zoom_dir.mkdir(exist_ok=True)
            
            for x in range(x_min, x_max + 1):
                x_dir = zoom_dir / str(x)
                x_dir.mkdir(exist_ok=True)
                
                for y in range(y_min, y_max + 1):
                    if progress.wasCanceled():
                        self._log(f"🗺️ Tile caching cancelled. Downloaded: {downloaded}, Skipped: {skipped}")
                        self._update_cache_status()
                        return
                    
                    tile_file = x_dir / f"{y}.png"
                    
                    if tile_file.exists():
                        skipped += 1
                    else:
                        try:
                            url = f"https://tile.openstreetmap.org/{zoom}/{x}/{y}.png"
                            req = urllib.request.Request(url, headers={'User-Agent': 'PyTNC-Pro/1.0'})
                            with urllib.request.urlopen(req, timeout=10) as resp:
                                with open(tile_file, 'wb') as f:
                                    f.write(resp.read())
                            downloaded += 1
                            # Small delay to avoid rate limiting
                            import time
                            time.sleep(0.05)
                        except Exception as e:
                            errors += 1
                    
                    progress.setValue(downloaded + skipped + errors)
                    progress.setLabelText(f"Zoom {zoom}: {downloaded} downloaded, {skipped} cached, {errors} errors")
                    QApplication.processEvents()
        
        progress.close()
        
        self._log(f"🗺️ Tile cache complete: {downloaded} new, {skipped} existing, {errors} errors")
        self._log(f"   Location: {TILE_CACHE_DIR}")
        
        # Verify tiles are accessible
        test_tile = None
        for png in TILE_CACHE_DIR.rglob("*.png"):
            test_tile = png
            break
        if test_tile:
            self._log(f"   ✓ Verified: {test_tile.relative_to(TILE_CACHE_DIR)}")
        
        self.cache_map_status.setText(f"✓ {downloaded + skipped}")
        self._update_cache_status()
        
        QMessageBox.information(self, "Tile Cache Complete",
            f"Map tiles cached!\n\n"
            f"Downloaded: {downloaded}\n"
            f"Already cached: {skipped}\n"
            f"Errors: {errors}\n"
            f"Total tiles: {downloaded + skipped}\n\n"
            f"Location:\n{TILE_CACHE_DIR}")
    
    def _test_tile_cache(self):
        """Test if tile cache is working"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        CACHE_DIR = _g.get("CACHE_DIR", None)
        TILE_CACHE_DIR = _g.get("TILE_CACHE_DIR", None)
        BASE_DIR = _g.get("BASE_DIR", None)
        import urllib.request
        
        self._log("🔍 Testing tile cache...")
        
        # Check if cache exists
        if not TILE_CACHE_DIR.exists():
            QMessageBox.warning(self, "No Tile Cache", f"Tile cache directory doesn't exist:\n{TILE_CACHE_DIR}")
            return
        
        # List zoom levels
        zoom_dirs = sorted([d.name for d in TILE_CACHE_DIR.iterdir() if d.is_dir()])
        
        # Find a cached tile
        test_tile = None
        for png in TILE_CACHE_DIR.rglob("*.png"):
            test_tile = png
            break
        
        if not test_tile:
            QMessageBox.warning(self, "Empty Cache", "No tiles found in cache. Click ⬇️ to download tiles.")
            return
        
        # Get the relative path
        rel_path = test_tile.relative_to(TILE_CACHE_DIR)
        
        # Try to fetch it via HTTP
        tile_url = f"http://127.0.0.1:{self.http_port}/tile_cache/{rel_path.as_posix()}"
        
        try:
            req = urllib.request.Request(tile_url)
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = resp.read()
                source = resp.headers.get('X-Tile-Source', 'unknown')
                self._log(f"   ✅ Cache test OK ({len(data)} bytes)")
                
                QMessageBox.information(self, "Tile Cache Working!", 
                    f"Tile cache is working!\n\n"
                    f"Tiles: {sum(1 for _ in TILE_CACHE_DIR.rglob('*.png')):,}\n"
                    f"Zoom levels: {', '.join(zoom_dirs)}\n\n"
                    f"Your map should work offline.")
        except Exception as e:
            self._log(f"   ❌ Cache test failed: {e}")
            QMessageBox.warning(self, "Tile Cache Error",
                f"Could not fetch tile via HTTP!\n\n"
                f"File exists: {test_tile.exists()}\n"
                f"URL: {tile_url}\n"
                f"Error: {e}\n\n"
                f"Check console for more details.")
    
    def _cache_digipeaters(self):
        """Download digipeaters from Overpass API — non-blocking"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)
        if not NetworkFetchWorker:
            self._log("❌ Digi cache: NetworkFetchWorker not available")
            return

        self._log("📡 Caching digipeaters...")
        if hasattr(self, 'cache_digi_status'):
            self.cache_digi_status.setText("⬇️")

        center_lat, center_lon = self._get_center_location()
        radius_meters = 80000  # ~50 miles
        query = f'[out:json];node["radio:aprs"="yes"](around:{radius_meters},{center_lat},{center_lon});out 50;'

        overpass_servers = [
            "https://overpass-api.de/api/interpreter",
            "https://overpass.kumi.systems/api/interpreter",
        ]

        url = f"{overpass_servers[0]}?data={urllib.parse.quote(query)}"
        self._digi_fallback_url = f"{overpass_servers[1]}?data={urllib.parse.quote(query)}"
        self._digi_fallback_tried = False

        worker = NetworkFetchWorker(url, timeout=30)
        worker.signals.finished.connect(self._process_digi_data)
        worker.signals.error.connect(lambda e: self._process_digi_data({"error": str(e)}))
        QThreadPool.globalInstance().start(worker)

    def _process_digi_data(self, data):
        """Process digipeater data from Overpass"""
        _g = _mg()
        NetworkFetchWorker = _g.get("NetworkFetchWorker", None)

        if isinstance(data, dict) and "error" in data:
            if not getattr(self, '_digi_fallback_tried', True) and hasattr(self, '_digi_fallback_url'):
                self._digi_fallback_tried = True
                self._log("⚠️ Primary Overpass failed, trying fallback...")
                worker = NetworkFetchWorker(self._digi_fallback_url, timeout=30)
                worker.signals.finished.connect(self._process_digi_data)
                worker.signals.error.connect(lambda e: self._process_digi_data({"error": str(e)}))
                QThreadPool.globalInstance().start(worker)
                return
            # Both servers failed — try existing cache
            cache_file = self._ensure_cache_dir() / "digipeaters.json"
            if cache_file.exists():
                self._log("📡 Overpass unavailable — using existing digi cache")
            else:
                self._log(f"❌ Digi cache failed: {data['error']}")
            self._update_cache_status()
            return

        try:
            import json as _json
            elements = data.get("elements", []) if isinstance(data, dict) else []
            digipeaters = []
            for elem in elements:
                tags = elem.get("tags", {})
                digipeaters.append({
                    "lat": elem.get("lat"),
                    "lon": elem.get("lon"),
                    "call": tags.get("callsign", tags.get("name", "Unknown")),
                    "freq": tags.get("frequency", "144.390")
                })
            cache_file = self._ensure_cache_dir() / "digipeaters.json"
            with open(cache_file, 'w') as f:
                _json.dump({
                    "cached_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "digipeaters": digipeaters
                }, f)
            self._log(f"📡 Cached {len(digipeaters)} digipeaters")
        except Exception as e:
            self._log(f"❌ Digi cache processing failed: {e}")
        self._update_cache_status()
    

    # =========================================================================
    
    def _rx_toggle_fires(self, state):
        """Toggle fire layer from RX page checkbox"""
        enabled = state == Qt.CheckState.Checked.value
        
        # Sync with Settings tab checkbox
        if hasattr(self, 'fire_enabled'):
            self.fire_enabled.blockSignals(True)
            self.fire_enabled.setChecked(enabled)
            self.fire_enabled.blockSignals(False)
        
        if enabled:
            # Check for API key
            api_key = self.fire_api_key.text().strip() if hasattr(self, 'fire_api_key') else ""
            if not api_key:
                self._log("🔥 Fire layer requires API key - set in Settings tab")
                self.rx_fire_check.blockSignals(True)
                self.rx_fire_check.setChecked(False)
                self.rx_fire_check.blockSignals(False)
                return
            self._toggle_fire_monitor(Qt.CheckState.Checked.value)
        else:
            self._toggle_fire_monitor(Qt.CheckState.Unchecked.value)
    
    def _rx_toggle_quakes(self, state):
        """Toggle earthquake layer from RX page checkbox"""
        enabled = state == Qt.CheckState.Checked.value
        
        # Sync with Settings tab checkbox
        if hasattr(self, 'quake_enabled'):
            self.quake_enabled.blockSignals(True)
            self.quake_enabled.setChecked(enabled)
            self.quake_enabled.blockSignals(False)
        
        self._toggle_earthquake_monitor(Qt.CheckState.Checked.value if enabled else Qt.CheckState.Unchecked.value)
    
    def _rx_toggle_aqi(self, state):
        """Toggle AQI layer from RX page checkbox"""
        enabled = state == Qt.CheckState.Checked.value
        
        # Sync with Settings tab checkbox
        if hasattr(self, 'aqi_enabled'):
            self.aqi_enabled.blockSignals(True)
            self.aqi_enabled.setChecked(enabled)
            self.aqi_enabled.blockSignals(False)
        
        self._toggle_aqi_monitor(Qt.CheckState.Checked.value if enabled else Qt.CheckState.Unchecked.value)

    # =========================================================================
    # Locations (CSV/Excel load, map display, beacon)
    # =========================================================================

    def _load_locations_menu(self):
        """Show menu to choose file or folder loading"""
        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: #1e3a5f; color: white; border: 1px solid #2d5a87; }
            QMenu::item:selected { background: #2d5a87; }
        """)
        
        file_action = menu.addAction("📄 Load File(s)")
        folder_action = menu.addAction("📁 Load Folder")
        
        action = menu.exec(self.load_locations_btn.mapToGlobal(
            self.load_locations_btn.rect().bottomLeft()))
        
        if action == file_action:
            self._load_locations_files()
        elif action == folder_action:
            self._load_locations_folder()
    
    def _load_locations_files(self):
        """Load locations from multiple CSV/XLS files"""
        filenames, _ = QFileDialog.getOpenFileNames(
            self, "Load Location Files", str(BASE_DIR),
            "CSV/Excel Files (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not filenames:
            return
        
        total_loaded = 0
        for filename in filenames:
            count = self._load_single_location_file(Path(filename))
            total_loaded += count
        
        if total_loaded > 0:
            self._log(f"📍 Loaded {total_loaded} locations from {len(filenames)} files (total: {len(self.custom_locations)})")
            self.location_count_lbl.setText(f"({len(self.custom_locations)})")
            self.beacon_locations_btn.show()
            self.clear_locations_btn.show()
            self._display_locations_on_map()
    
    def _load_locations_folder(self):
        """Load all CSV/XLS files from a folder"""
        folder = QFileDialog.getExistingDirectory(
            self, "Select Locations Folder", str(BASE_DIR)
        )
        if not folder:
            return
        
        folder_path = Path(folder)
        files = list(folder_path.glob("*.csv")) + list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))
        
        if not files:
            QMessageBox.warning(self, "No Files", f"No CSV or Excel files found in:\n{folder}")
            return
        
        total_loaded = 0
        for filepath in sorted(files):
            count = self._load_single_location_file(filepath)
            total_loaded += count
        
        if total_loaded > 0:
            self._log(f"📍 Loaded {total_loaded} locations from {len(files)} files (total: {len(self.custom_locations)})")
            self.location_count_lbl.setText(f"({len(self.custom_locations)})")
            self.beacon_locations_btn.show()
            self.clear_locations_btn.show()
            self._display_locations_on_map()
        
        QMessageBox.information(self, "Folder Loaded",
            f"Loaded {total_loaded} locations from {len(files)} files.")
    
    def _load_single_location_file(self, filepath: Path) -> int:
        """Load locations from a single file. Returns count loaded."""
        try:
            locations = []
            
            if filepath.suffix.lower() == '.csv':
                import csv
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        name = row.get('Name', '').strip()
                        lat_str = row.get('LAT', row.get('Lat', row.get('lat', ''))).strip()
                        lon_str = row.get('Long', row.get('Lon', row.get('lon', ''))).strip()
                        
                        if not name or not lat_str or not lon_str:
                            continue
                        
                        try:
                            lat = float(lat_str)
                            lon = float(lon_str)
                        except ValueError:
                            continue
                        
                        locations.append({
                            'name': name,
                            'address': row.get('Address', '').strip(),
                            'lat': lat,
                            'lon': lon,
                            'symbol': row.get('Symbol', '\\h').strip(),
                            'comment': row.get('Comment', '').strip(),
                            'source': filepath.name
                        })
            
            elif filepath.suffix.lower() in ('.xlsx', '.xls'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(filepath)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        name = str(row_dict.get('Name', '')).strip()
                        lat_str = str(row_dict.get('LAT', row_dict.get('Lat', ''))).strip()
                        lon_str = str(row_dict.get('Long', row_dict.get('Lon', ''))).strip()
                        
                        if not name or not lat_str or not lon_str:
                            continue
                        
                        try:
                            lat = float(lat_str)
                            lon = float(lon_str)
                        except ValueError:
                            continue
                        
                        locations.append({
                            'name': name,
                            'address': str(row_dict.get('Address', '')).strip(),
                            'lat': lat,
                            'lon': lon,
                            'symbol': str(row_dict.get('Symbol', '\\h')).strip(),
                            'comment': str(row_dict.get('Comment', '')).strip(),
                            'source': filepath.name
                        })
                except ImportError:
                    self._log(f"⚠️ Skipped {filepath.name} - openpyxl not installed")
                    return 0
            
            if locations:
                # Deduplicate: don't add if same name+lat+lon already exists
                existing_keys = {(loc['name'], round(loc['lat'], 5), round(loc['lon'], 5)) 
                                 for loc in self.custom_locations}
                new_locations = []
                for loc in locations:
                    key = (loc['name'], round(loc['lat'], 5), round(loc['lon'], 5))
                    if key not in existing_keys:
                        new_locations.append(loc)
                        existing_keys.add(key)
                
                if new_locations:
                    self.custom_locations.extend(new_locations)
                    self._log(f"  📄 {filepath.name}: {len(new_locations)} new locations")
                    return len(new_locations)
                else:
                    self._log(f"  ⚠️ {filepath.name}: all {len(locations)} locations already loaded")
                    return 0
            
            return 0
            
        except Exception as e:
            self._log(f"⚠️ Error loading {filepath.name}: {e}")
            return 0

    def _display_locations_on_map(self):
        """Display custom locations on the map"""
        if not self.map_ready or not self.custom_locations:
            return
        
        # Clear existing custom markers first
        self.map.page().runJavaScript("clearCustomLocations()")
        
        for loc in self.custom_locations:
            name = loc['name']
            lat = loc['lat']
            lon = loc['lon']
            symbol = loc.get('symbol', '\\h')
            comment = loc.get('comment', '')
            address = loc.get('address', '')
            
            # Escape for JavaScript
            name_js = json.dumps(name)
            comment_js = json.dumps(comment)
            address_js = json.dumps(address)
            symbol_js = json.dumps(symbol)
            
            js = f"addCustomLocation({name_js}, {lat}, {lon}, {symbol_js}, {comment_js}, {address_js})"
            self.map.page().runJavaScript(js)
        
        self._log(f"📍 Displayed {len(self.custom_locations)} locations on map")
    
    def _clear_locations(self):
        """Clear all loaded locations"""
        if not self.custom_locations:
            return
        
        count = len(self.custom_locations)
        self.custom_locations = []
        
        # Clear map markers
        if self.map_ready:
            self.map.page().runJavaScript("clearCustomLocations()")
        
        # Hide buttons and clear label
        self.beacon_locations_btn.hide()
        self.clear_locations_btn.hide()
        self.location_count_lbl.setText("")
        
        self._log(f"🗑️ Cleared {count} locations")
    
    def _beacon_locations_menu(self):
        """Show menu to choose RF or APRS-IS beacon"""
        if not self.custom_locations:
            self._log("❌ No locations to beacon")
            return
        
        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: #1e3a5f; color: white; border: 1px solid #2d5a87; }
            QMenu::item:selected { background: #2d5a87; }
        """)
        
        rf_action = menu.addAction("📻 Beacon via RF (Simplex)")
        is_action = menu.addAction("🌐 Beacon via APRS-IS")
        menu.addSeparator()
        cancel_action = menu.addAction("Cancel")
        
        action = menu.exec(self.beacon_locations_btn.mapToGlobal(
            self.beacon_locations_btn.rect().bottomLeft()))
        
        if action == rf_action:
            self._beacon_locations_rf()
        elif action == is_action:
            self._beacon_all_locations()
    
    def _beacon_locations_rf(self):
        """Beacon all locations as APRS objects via RF"""
        if not self.custom_locations:
            self._log("❌ No locations to beacon")
            return
        
        if not HAS_SOUNDDEVICE:
            QMessageBox.warning(self, "RF Disabled", 
                "sounddevice not installed.\nRF AFSK transmit is disabled.")
            return
        
        callsign = self.callsign_edit.text().strip().upper()
        if not callsign or callsign == "N0CALL":
            QMessageBox.warning(self, "No Callsign", "Set your callsign first")
            return
        
        ssid = self.ssid_combo.currentData()
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        path_str = self.path_combo.currentText().strip()
        
        # Parse path
        path_list = []
        if path_str and path_str.upper() != "DIRECT":
            for p in path_str.split(","):
                p = p.strip()
                if "-" in p:
                    pcall, pssid = p.rsplit("-", 1)
                    path_list.append((pcall, int(pssid)))
                else:
                    path_list.append((p, 0))
        
        # Check PTT
        if not self.ptt_serial or not self.ptt_serial.is_open:
            ptt_port = self.settings_ptt_combo.currentData() if hasattr(self, 'settings_ptt_combo') else None
            if ptt_port:
                try:
                    self.ptt_serial = serial.Serial(ptt_port, 9600, timeout=0.1)
                    self._set_ptt(False)
                    self._log(f"✅ Auto-connected PTT: {ptt_port}")
                except Exception as e:
                    QMessageBox.warning(self, "PTT Error", f"Could not connect PTT:\n{e}")
                    return
            else:
                QMessageBox.warning(self, "PTT Not Configured", "Configure PTT in Settings")
                return
        
        # Get TX audio device
        tx_device = self.settings_tx_audio_combo.currentData() if hasattr(self, 'settings_tx_audio_combo') else None
        if tx_device is None:
            QMessageBox.warning(self, "No TX Audio", "Select TX audio device in Settings")
            return
        
        tx_level_pct = self.settings_tx_level.value() if hasattr(self, 'settings_tx_level') else 10
        
        # Timestamp
        from datetime import datetime
        now = datetime.utcnow()
        timestamp = now.strftime("%d%H%M") + "z"
        
        # Confirm
        reply = QMessageBox.question(self, "Beacon Locations via RF",
            f"Beacon {len(self.custom_locations)} locations via RF?\n\n"
            f"Path: {path_str}\n"
            f"This will take ~{len(self.custom_locations) * 3} seconds.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        self._log(f"📻 Starting RF beacon of {len(self.custom_locations)} locations...")
        self.tx_in_progress = True
        sent = 0
        
        try:
            for i, loc in enumerate(self.custom_locations):
                name = loc['name'][:9].ljust(9)
                lat = loc['lat']
                lon = loc['lon']
                symbol = loc.get('symbol', '\\h')
                address = loc.get('address', '')
                comment = loc.get('comment', '')
                # Combine address and comment (address first)
                full_comment = f"{address} {comment}".strip()[:43]
                
                # Parse symbol
                sym_table = symbol[0] if len(symbol) >= 1 else '\\'
                sym_code = symbol[1] if len(symbol) >= 2 else 'h'
                
                # Format position
                lat_deg = int(abs(lat))
                lat_min = (abs(lat) - lat_deg) * 60
                lat_dir = "N" if lat >= 0 else "S"
                lon_deg = int(abs(lon))
                lon_min = (abs(lon) - lon_deg) * 60
                lon_dir = "E" if lon >= 0 else "W"
                
                # APRS Object format
                info = f";{name}*{timestamp}{lat_deg:02d}{lat_min:05.2f}{lat_dir}{sym_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{sym_code}{full_comment}"
                
                self._log(f"📡 TX [{i+1}/{len(self.custom_locations)}]: {name.strip()}")
                
                # Build packet
                packet_data = APRSPacketBuilder.build_ui_packet(
                    src_call=callsign, src_ssid=ssid,
                    dst_call="APPR01", dst_ssid=0,
                    path=path_list,
                    info=info
                )
                fcs = APRSPacketBuilder.compute_fcs(packet_data)
                full_packet = packet_data + bytes([fcs & 0xFF, (fcs >> 8) & 0xFF])
                
                # Generate audio
                modulator = AFSKModulator(TX_SAMPLE_RATE)
                audio = modulator.generate_packet_audio(full_packet, preamble_flags=50, postamble_flags=8)
                
                # Add silence
                silence = np.zeros(int(TX_SAMPLE_RATE * 0.03), dtype=np.float32)
                audio = np.concatenate([silence, audio, silence])
                
                # Apply level
                audio = apply_cosine_ramp(audio, TX_SAMPLE_RATE, ramp_ms=5.0)
                audio = audio * (tx_level_pct / 100.0)
                
                # Soft limit
                if float(np.abs(audio).max()) > 0.9:
                    audio = np.tanh(audio * 1.5) * 0.9
                
                # Get device info for resampling
                device_info = sd.query_devices(tx_device)
                device_sr = int(device_info.get('default_samplerate', 48000))
                max_ch = device_info.get('max_output_channels', 1)
                
                if device_sr != TX_SAMPLE_RATE:
                    from scipy import signal as scipy_signal
                    num_samples = int(len(audio) * device_sr / TX_SAMPLE_RATE)
                    audio = scipy_signal.resample(audio, num_samples).astype(np.float32)
                
                if max_ch >= 2:
                    audio_out = np.column_stack([audio, audio]).astype(np.float32)
                else:
                    audio_out = audio.astype(np.float32)
                
                # Key PTT
                self._set_ptt(True)
                time.sleep(0.5)  # Let radio settle
                
                # Play audio
                sd.play(audio_out, device_sr, device=tx_device)
                sd.wait()
                
                # Unkey PTT
                time.sleep(0.15)
                self._set_ptt(False)
                
                sent += 1
                
                # Pause between packets
                if i < len(self.custom_locations) - 1:
                    time.sleep(1.5)
                
                QApplication.processEvents()
                
        except Exception as e:
            self._log(f"❌ RF beacon error: {e}")
            QMessageBox.warning(self, "Error", f"RF beacon failed:\n{e}")
        finally:
            self.tx_in_progress = False
            self._set_ptt(False)
        
        self._log(f"📻 RF beacon complete: {sent}/{len(self.custom_locations)} sent")
        QMessageBox.information(self, "RF Beacon Complete",
            f"Sent {sent} location objects via RF.")

    def _beacon_all_locations(self):
        """Beacon all locations as APRS objects"""
        if not self.custom_locations:
            self._log("❌ No locations to beacon")
            return
        
        callsign = self.callsign_edit.text().strip().upper()
        if not callsign or callsign == "N0CALL":
            QMessageBox.warning(self, "No Callsign", "Set your callsign first")
            return
        
        # Check if APRS-IS connected
        if not (hasattr(self, 'aprs_is_socket') and self.aprs_is_socket):
            QMessageBox.warning(self, "Not Connected", 
                "Connect to APRS-IS first to beacon objects.\n\n"
                "Go to Settings → APRS-IS and connect.")
            return
        
        ssid = self.ssid_combo.currentData()
        full_call = f"{callsign}-{ssid}" if ssid > 0 else callsign
        
        # Timestamp (DHM format for objects)
        from datetime import datetime
        now = datetime.utcnow()
        timestamp = now.strftime("%d%H%M") + "z"
        
        sent = 0
        for loc in self.custom_locations:
            try:
                name = loc['name'][:9].ljust(9)  # Object names are 9 chars
                lat = loc['lat']
                lon = loc['lon']
                symbol = loc.get('symbol', '\\h')
                address = loc.get('address', '')
                comment = loc.get('comment', '')
                # Combine address and comment (address first)
                full_comment = f"{address} {comment}".strip()[:43]
                
                # Parse symbol table/code
                if len(symbol) >= 2:
                    sym_table = symbol[0]
                    sym_code = symbol[1]
                else:
                    sym_table = '\\'
                    sym_code = 'h'
                
                # Format position
                lat_deg = int(abs(lat))
                lat_min = (abs(lat) - lat_deg) * 60
                lat_dir = "N" if lat >= 0 else "S"
                
                lon_deg = int(abs(lon))
                lon_min = (abs(lon) - lon_deg) * 60
                lon_dir = "E" if lon >= 0 else "W"
                
                # APRS Object format: ;NAME*DDHHMMz/DDMM.MMN/DDDMM.MMWsComment
                # ; = object, * = live object, _ = killed object
                info = f";{name}*{timestamp}{lat_deg:02d}{lat_min:05.2f}{lat_dir}{sym_table}{lon_deg:03d}{lon_min:05.2f}{lon_dir}{sym_code}{full_comment}"
                
                packet = f"{full_call}>APPR01,TCPIP*:{info}\r\n"
                self.aprs_is_socket.send(packet.encode())
                sent += 1
                
                # Small delay between packets
                time.sleep(0.1)
                
            except Exception as e:
                self._log(f"⚠️ Failed to beacon {loc['name']}: {e}")
        
        self._log(f"📡 Beaconed {sent}/{len(self.custom_locations)} location objects")
        QMessageBox.information(self, "Beaconed", 
            f"Sent {sent} location objects via APRS-IS.\n\n"
            f"They should appear on aprs.fi within a minute.")

